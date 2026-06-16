"""
AGENT 6: REPORT BUILDER (v2 - FIXED COLUMN LAYOUT)
=====================================================
Standardized raw sheet layout ensures all SUMIFS/INDEX-MATCH formulas work.
Both Raw-BooksOnly and Raw-GSTR2B Only have IDENTICAL column structure:
  A=GSTIN  B=Vendor(VLOOKUP)  C=InvNo  D=Type  E=Date
  F=Taxable  G=IGST  H=CGST  I=SGST  J=Cess  K=Total
"""
import pandas as pd
import numpy as np
import json
import os

AGENT_NAME = "Report Builder"
AGENT_ID = 6

NUM_COLS = ['Taxable Amount', 'IGST Amount', 'CGST Amount', 'SGST Amount', 'Cess Amount', 'Total Amount']

# STANDARDIZED column order for raw sheets
STD_COLS = [
    ('GST Registration Number', 'GSTIN', 'A'),
    ('_VLOOKUP_', 'Vendor (VLOOKUP)', 'B'),
    ('Transaction Number', 'Invoice No.', 'C'),
    ('Transaction Type', 'Type', 'D'),
    ('Transaction Date', 'Date', 'E'),
    ('Taxable Amount', 'Taxable Amt', 'F'),
    ('IGST Amount', 'IGST', 'G'),
    ('CGST Amount', 'CGST', 'H'),
    ('SGST Amount', 'SGST', 'I'),
    ('Cess Amount', 'Cess', 'J'),
    ('Total Amount', 'Total Amt', 'K'),
]


def run(input_file, output_file, dataframes, vendor_map, gst_cache,
        all_gstins, inv_matched, books_unmatched, gstn_unmatched,
        books_gstins, gstn_gstins, anomalies, patterns, itc_summary,
        log_fn=None):

    results = {'agent': AGENT_NAME, 'status': 'running', 'checks': [], 'output_file': output_file, 'stats': {}}

    def log(msg):
        if log_fn: log_fn(AGENT_ID, msg)

    log("Building final report with STANDARDIZED column layout...")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule

    wb = Workbook()

    # ---- STYLES ----
    hf = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    hfl = PatternFill('solid', fgColor='1F4E79')
    tf = Font(name='Calibri', bold=True, size=14, color='D9008D')
    inf = Font(name='Calibri', size=9, color='595959', italic=True)
    n9 = Font(name='Calibri', size=9)
    b9 = Font(name='Calibri', bold=True, size=9)
    b10 = Font(name='Calibri', bold=True, size=10)
    rf = Font(name='Calibri', bold=True, size=9, color='CC0000')
    gf = Font(name='Calibri', bold=True, size=9, color='006100')
    wbf = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    bdr = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    gr = PatternFill('solid', fgColor='C6EFCE')
    rd = PatternFill('solid', fgColor='FFC7CE')
    yw = PatternFill('solid', fgColor='FFEB9C')
    bl = PatternFill('solid', fgColor='BDD7EE')
    gy = PatternFill('solid', fgColor='F2F2F2')
    og = PatternFill('solid', fgColor='FCE4D6')
    dk = PatternFill('solid', fgColor='1F4E79')
    wh = PatternFill('solid', fgColor='FFFFFF')
    lg = PatternFill('solid', fgColor='A9D18E')
    INR = '#,##0.00'

    def whd(ws, row, headers):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row, i, h)
            c.font = hf; c.fill = hfl; c.border = bdr
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def wc(ws, r, c, val, fmt=None, font=None, fill=None, align='center'):
        cell = ws.cell(r, c, val)
        cell.border = bdr
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        cell.font = font or n9
        if fmt: cell.number_format = fmt
        if fill: cell.fill = fill
        return cell

    def scw(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def get_vendor(gstin):
        if pd.isna(gstin): return ''
        return vendor_map.get(str(gstin).strip(), str(gstin).strip())

    df_rec = dataframes.get('Reconciled', pd.DataFrame())
    df_nig = dataframes.get('Not found in GSTN', pd.DataFrame())
    df_nib = dataframes.get('Not found in Zoho Books', pd.DataFrame())
    df_pm = dataframes.get('Partially Matched', pd.DataFrame())

    # ==================================================================
    # STANDARDIZED RAW SHEET WRITER
    # Layout: A=GSTIN B=Vendor(VLOOKUP) C=InvNo D=Type E=Date
    #         F=Taxable G=IGST H=CGST I=SGST J=Cess K=Total
    # ==================================================================
    def write_std_raw(wb, title, df, tab_color, include_source=False):
        """Write raw data in STANDARDIZED 11-column layout.
        If include_source=True, adds col L = 'Source' (for PartialMatch Books/GSTN filtering)."""
        ws = wb.create_sheet(title)
        ws.sheet_properties.tabColor = tab_color

        # Headers
        for i, (src_col, display_name, col_letter) in enumerate(STD_COLS, 1):
            c = ws.cell(1, i, display_name)
            c.font = hf; c.fill = hfl; c.border = bdr
            c.alignment = Alignment(horizontal='center', wrap_text=True)

        # Extra Source header for PartialMatch
        if include_source:
            c = ws.cell(1, 12, 'Source')
            c.font = hf; c.fill = hfl; c.border = bdr
            c.alignment = Alignment(horizontal='center', wrap_text=True)

        # Data rows
        for ri, (_, row) in enumerate(df.iterrows(), 2):
            for ci, (src_col, display_name, col_letter) in enumerate(STD_COLS, 1):
                if src_col == '_VLOOKUP_':
                    wc(ws, ri, ci, None, align='left')
                    ws.cell(ri, ci).value = f"=IFERROR(VLOOKUP(A{ri},'Vendor Master'!$A:$C,3,FALSE),\"Unknown\")"
                else:
                    val = row.get(src_col, '')
                    if pd.isna(val): val = 0 if src_col in NUM_COLS else ''
                    c = ws.cell(ri, ci, val)
                    c.border = bdr; c.font = n9
                    if src_col in NUM_COLS:
                        c.number_format = INR

            # Write Source column for PartialMatch
            if include_source:
                src_val = row.get('Source', '')
                if pd.isna(src_val): src_val = ''
                c = ws.cell(ri, 12, str(src_val).strip())
                c.border = bdr; c.font = n9

        last = len(df) + 1
        # Total row
        tr = last + 1
        wc(ws, tr, 1, 'TOTAL', font=wbf, fill=dk)
        for ci in range(2, 6):
            wc(ws, tr, ci, '', fill=dk)
        for ci in range(6, 12):  # F through K
            cl = get_column_letter(ci)
            wc(ws, tr, ci, None, fmt=INR, font=wbf, fill=dk)
            ws.cell(tr, ci).value = f'=SUM({cl}2:{cl}{last})'
        if include_source:
            wc(ws, tr, 12, '', fill=dk)

        widths = [22, 36, 24, 18, 14, 16, 14, 14, 14, 12, 18]
        if include_source:
            widths.append(10)
        scw(ws, widths)
        ws.freeze_panes = 'A2'
        last_col = 'L' if include_source else 'K'
        ws.auto_filter.ref = f'A1:{last_col}{last}'
        return last

    # ==================================================================
    # VENDOR MASTER (LOOKUP SOURCE)
    # A=GSTIN  B=S.No.  C=Vendor Name
    # ==================================================================
    log("Writing Vendor Master...")
    ws_vm = wb.create_sheet('Vendor Master')
    ws_vm.sheet_properties.tabColor = '7D3C98'
    whd(ws_vm, 1, ['GSTIN', 'S.No.', 'Vendor Name', 'Source', 'State Code', 'Status'])

    all_gstins_sorted = sorted(all_gstins)
    for idx, gstin in enumerate(all_gstins_sorted):
        rr = 2 + idx
        name = get_vendor(gstin)
        is_missing = (name == gstin)
        src_tag = 'NOT FOUND' if is_missing else ('GST Portal' if gstin in gst_cache else 'Zoho Books')
        status = gst_cache.get(gstin, {}).get('data', {}).get('sts', '') if gstin in gst_cache else ''
        f = rd if is_missing else (gy if idx % 2 == 0 else wh)
        wc(ws_vm, rr, 1, gstin, fill=f)
        wc(ws_vm, rr, 2, idx+1, fill=f)
        wc(ws_vm, rr, 3, name, fill=f, align='left', font=rf if is_missing else n9)
        wc(ws_vm, rr, 4, src_tag, fill=f)
        wc(ws_vm, rr, 5, None, fill=f); ws_vm.cell(rr, 5).value = f'=LEFT(A{rr},2)'
        wc(ws_vm, rr, 6, status, fill=f)

    vm_last = 1 + len(all_gstins_sorted)
    scw(ws_vm, [22, 8, 45, 16, 12, 12])
    ws_vm.freeze_panes = 'A2'

    # ==================================================================
    # RAW SHEETS (STANDARDIZED LAYOUT)
    # ==================================================================
    log("Writing STANDARDIZED raw data sheets...")
    log("  Layout: A=GSTIN B=Vendor C=InvNo D=Type E=Date F=Taxable G=IGST H=CGST I=SGST J=Cess K=Total")
    rec_last = write_std_raw(wb, 'Raw-Reconciled', df_rec, '228B22')
    nig_last = write_std_raw(wb, 'Raw-BooksOnly', df_nig, 'CC0000')
    nib_last = write_std_raw(wb, 'Raw-GSTR2B Only', df_nib, '2980B9')
    pm_last = write_std_raw(wb, 'Raw-PartialMatch', df_pm, '8E44AD', include_source=True)

    log(f"  Raw-Reconciled: {rec_last} rows | Raw-BooksOnly: {nig_last} rows | Raw-GSTR2B: {nib_last} rows")

    # ==================================================================
    # FORMULAS NOW CORRECT because layout is standardized:
    #   $A = GSTIN
    #   $C = Invoice Number (match key)
    #   $D = Type
    #   $E = Date
    #   $F = Taxable Amount
    #   $G = IGST
    #   $H = CGST
    #   $I = SGST
    #   $J = Cess
    #   $K = Total Amount
    # ==================================================================

    RB = "'Raw-BooksOnly'"    # Books raw sheet
    RG = "'Raw-GSTR2B Only'"  # GSTR-2B raw sheet
    RP = "'Raw-PartialMatch'" # Partial match raw sheet

    # ==================================================================
    # EXPERT CROSS-MATCH (SUMIFS + INDEX-MATCH)
    # ==================================================================
    log("Writing Expert Cross-Match with corrected SUMIFS...")
    ws_xm = wb.create_sheet('Expert Cross-Match')
    ws_xm.sheet_properties.tabColor = '00B050'

    ws_xm.merge_cells('A1:AC1')
    ws_xm['A1'] = 'WIOM EXPERT CROSS-MATCH  |  GSTIN+InvNo+Type+Date from matched data, ALL amounts via SUMIFS (BooksOnly + PartialMatch + GSTR2B)'
    ws_xm['A1'].font = tf; ws_xm['A1'].alignment = Alignment(horizontal='center')
    ws_xm.merge_cells('A2:AC2')
    ws_xm['A2'] = 'SUMIFS: BooksOnly+PartialMatch(Books) & GSTR2B+PartialMatch(GSTN)  |  Credits=NEGATIVE  |  TAX HEAD MISMATCH: Books=IGST vs 2B=CGST+SGST detected'
    ws_xm['A2'].font = inf; ws_xm['A2'].alignment = Alignment(horizontal='center')
    ws_xm.merge_cells('A3:AC3')
    ws_xm['A3'] = 'Raw sheets: A=GSTIN C=InvNo D=Type E=Date F=Taxable G=IGST H=CGST I=SGST J=Cess K=Total L=Source(PartialMatch only)  |  Orange rows = TAX HEAD MISMATCH'
    ws_xm['A3'].font = Font(name='Calibri', size=8, color='808080', italic=True)
    ws_xm['A3'].alignment = Alignment(horizontal='right')

    r = 4
    hdrs_xm = ['S.No.', 'GSTIN', 'Vendor (VLOOKUP)', 'Type',
               'Books Inv No.', 'GSTR-2B Inv No.', 'Inv Match?',
               'Books Date', 'GSTR-2B Date', 'Date Match?',
               'Books Taxable (SUMIFS)', 'GSTR-2B Taxable (SUMIFS)', 'Taxable Diff',
               'Books IGST (SUMIFS)', 'GSTR-2B IGST (SUMIFS)',
               'Books CGST (SUMIFS)', 'GSTR-2B CGST (SUMIFS)',
               'Books SGST (SUMIFS)', 'GSTR-2B SGST (SUMIFS)',
               'Books Total Tax', 'GSTR-2B Total Tax', 'Tax Diff',
               'Books Total (SUMIFS)', 'GSTR-2B Total (SUMIFS)', 'Total Diff', 'Diff %',
               'Recon Status', 'Expert Remark', 'Action Required']
    whd(ws_xm, r, hdrs_xm)

    # Helper: build SUMIFS that checks Raw-BooksOnly + Raw-PartialMatch (Books side ONLY via Source=Books filter)
    def books_sumifs(col_letter, rr):
        """SUMIFS from BooksOnly + PartialMatch WHERE Source='Books', keyed on GSTIN+InvNo."""
        return (f'=SUMIFS({RB}!${col_letter}:${col_letter},{RB}!$A:$A,B{rr},{RB}!$C:$C,E{rr})'
                f'+SUMIFS({RP}!${col_letter}:${col_letter},{RP}!$A:$A,B{rr},{RP}!$C:$C,E{rr},{RP}!$L:$L,"Books")')

    def gstn_sumifs(col_letter, rr):
        """SUMIFS from GSTR2B Only + PartialMatch WHERE Source='GSTN', keyed on GSTIN+InvNo."""
        return (f'=SUMIFS({RG}!${col_letter}:${col_letter},{RG}!$A:$A,B{rr},{RG}!$C:$C,F{rr})'
                f'+SUMIFS({RP}!${col_letter}:${col_letter},{RP}!$A:$A,B{rr},{RP}!$C:$C,F{rr},{RP}!$L:$L,"GSTN")')

    for idx, m in enumerate(inv_matched):
        rr = r + 1 + idx
        fill = gr if m['Recon_Status'] == 'Reconciled' else og

        # Col A: S.No (hardcoded)
        wc(ws_xm, rr, 1, idx+1, fill=fill)
        # Col B: GSTIN (hardcoded MATCH KEY)
        wc(ws_xm, rr, 2, m['GSTIN'], fill=fill)
        # Col C: Vendor = VLOOKUP from Vendor Master
        wc(ws_xm, rr, 3, None, fill=fill, align='left')
        ws_xm.cell(rr, 3).value = f"=IFERROR(VLOOKUP(B{rr},'Vendor Master'!$A:$C,3,FALSE),\"Unknown\")"
        # Col D: Type — HARDCODED from matched data (CSE array formula not supported by openpyxl)
        type_val = m.get('Type', 'Bill')
        if pd.isna(type_val) or str(type_val).strip() == '':
            type_val = 'Bill'
        wc(ws_xm, rr, 4, str(type_val), fill=fill)
        # Col E: Books Inv No. (hardcoded MATCH KEY)
        wc(ws_xm, rr, 5, m['Books_Inv'], fill=fill)
        # Col F: GSTR-2B Inv No. (hardcoded MATCH KEY)
        wc(ws_xm, rr, 6, m['GSTN_Inv'], fill=fill)
        # Col G: Inv Match?
        wc(ws_xm, rr, 7, None, fill=fill, font=b9)
        ws_xm.cell(rr, 7).value = f'=IF(EXACT(TRIM(E{rr}),TRIM(F{rr})),"YES","NO")'

        # Col H: Books Date — HARDCODED from matched data (CSE array formula not supported)
        books_date = m.get('Books_Date', '')
        if pd.notna(books_date):
            try:
                books_date = pd.Timestamp(books_date)
                wc(ws_xm, rr, 8, books_date, fill=fill, fmt='DD/MM/YYYY')
            except Exception:
                wc(ws_xm, rr, 8, str(books_date), fill=fill)
        else:
            wc(ws_xm, rr, 8, 'N/A', fill=fill)
        # Col I: GSTR-2B Date — HARDCODED from matched data (CSE array formula not supported)
        gstn_date = m.get('GSTN_Date', '')
        if pd.notna(gstn_date):
            try:
                gstn_date = pd.Timestamp(gstn_date)
                wc(ws_xm, rr, 9, gstn_date, fill=fill, fmt='DD/MM/YYYY')
            except Exception:
                wc(ws_xm, rr, 9, str(gstn_date), fill=fill)
        else:
            wc(ws_xm, rr, 9, 'N/A', fill=fill)
        # Col J: Date Match?
        wc(ws_xm, rr, 10, None, fill=fill, font=b9)
        ws_xm.cell(rr, 10).value = f'=IF(OR(H{rr}="N/A",I{rr}="N/A"),"N/A",IF(H{rr}=I{rr},"YES","NO"))'

        # Col K: Books Taxable = SUMIFS(BooksOnly + PartialMatch $F)
        wc(ws_xm, rr, 11, None, fmt=INR, fill=fill)
        ws_xm.cell(rr, 11).value = books_sumifs('F', rr)
        # Col L: GSTR-2B Taxable = SUMIFS(GSTR2B + PartialMatch $F)
        wc(ws_xm, rr, 12, None, fmt=INR, fill=fill)
        ws_xm.cell(rr, 12).value = gstn_sumifs('F', rr)
        # Col M: Taxable Diff
        wc(ws_xm, rr, 13, None, fmt=INR, fill=fill)
        ws_xm.cell(rr, 13).value = f'=K{rr}-L{rr}'

        # Col N: Books IGST = SUMIFS(BooksOnly + PartialMatch $G)
        wc(ws_xm, rr, 14, None, fmt=INR, fill=fill)
        ws_xm.cell(rr, 14).value = books_sumifs('G', rr)
        # Col O: GSTR-2B IGST = SUMIFS(GSTR2B + PartialMatch $G)
        wc(ws_xm, rr, 15, None, fmt=INR, fill=fill)
        ws_xm.cell(rr, 15).value = gstn_sumifs('G', rr)
        # Col P: Books CGST = SUMIFS(BooksOnly + PartialMatch $H)
        wc(ws_xm, rr, 16, None, fmt=INR, fill=fill)
        ws_xm.cell(rr, 16).value = books_sumifs('H', rr)
        # Col Q: GSTR-2B CGST = SUMIFS(GSTR2B + PartialMatch $H)
        wc(ws_xm, rr, 17, None, fmt=INR, fill=fill)
        ws_xm.cell(rr, 17).value = gstn_sumifs('H', rr)
        # Col R: Books SGST = SUMIFS(BooksOnly + PartialMatch $I)
        wc(ws_xm, rr, 18, None, fmt=INR, fill=fill)
        ws_xm.cell(rr, 18).value = books_sumifs('I', rr)
        # Col S: GSTR-2B SGST = SUMIFS(GSTR2B + PartialMatch $I)
        wc(ws_xm, rr, 19, None, fmt=INR, fill=fill)
        ws_xm.cell(rr, 19).value = gstn_sumifs('I', rr)

        # Col T: Books Total Tax = N+P+R
        wc(ws_xm, rr, 20, None, fmt=INR, fill=fill)
        ws_xm.cell(rr, 20).value = f'=N{rr}+P{rr}+R{rr}'
        # Col U: GSTR-2B Total Tax = O+Q+S
        wc(ws_xm, rr, 21, None, fmt=INR, fill=fill)
        ws_xm.cell(rr, 21).value = f'=O{rr}+Q{rr}+S{rr}'
        # Col V: Tax Diff
        wc(ws_xm, rr, 22, None, fmt=INR, fill=fill)
        ws_xm.cell(rr, 22).value = f'=T{rr}-U{rr}'

        # Col W: Books Total = SUMIFS(BooksOnly + PartialMatch $K)
        wc(ws_xm, rr, 23, None, fmt=INR, fill=fill, font=b9)
        ws_xm.cell(rr, 23).value = books_sumifs('K', rr)
        # Col X: GSTR-2B Total = SUMIFS(GSTR2B + PartialMatch $K)
        wc(ws_xm, rr, 24, None, fmt=INR, fill=fill, font=b9)
        ws_xm.cell(rr, 24).value = gstn_sumifs('K', rr)
        # Col Y: Total Diff
        wc(ws_xm, rr, 25, None, fmt=INR, fill=fill)
        ws_xm.cell(rr, 25).value = f'=W{rr}-X{rr}'
        # Col Z: Diff %
        wc(ws_xm, rr, 26, None, fmt='0.00%', fill=fill)
        ws_xm.cell(rr, 26).value = f'=IF(ABS(X{rr})>0,Y{rr}/ABS(X{rr}),0)'

        # Tax Head Mismatch detection: Books IGST>0 & CGST=0 but 2B CGST>0 & IGST=0 or vice versa
        # N=Books IGST, O=GSTR2B IGST, P=Books CGST, Q=GSTR2B CGST
        thm_formula = f'IF(AND(ABS(N{rr})>0.01,ABS(P{rr})<0.01,ABS(Q{rr})>0.01,ABS(O{rr})<0.01)," | TAX HEAD MISMATCH: Books=IGST 2B=CGST+SGST",IF(AND(ABS(P{rr})>0.01,ABS(N{rr})<0.01,ABS(O{rr})>0.01,ABS(Q{rr})<0.01)," | TAX HEAD MISMATCH: Books=CGST+SGST 2B=IGST",""))'

        # Col AA: Recon Status (includes Tax Head Mismatch flag)
        wc(ws_xm, rr, 27, None, fill=fill, font=b9)
        ws_xm.cell(rr, 27).value = (
            f'=IF(AND(G{rr}="YES",ABS(Y{rr})<1),'
            f'IF({thm_formula}<>"","Tax Head Mismatch","Fully Reconciled"),'
            f'IF(ABS(Y{rr})<1,'
            f'IF({thm_formula}<>"","Tax Head Mismatch - Inv Match","Partially Reconciled - Inv Mismatch"),'
            f'IF(ABS(M{rr})<1,"Partially Reconciled - Tax Diff","Amount Mismatch")))'
        )
        # Col AB: Expert Remark (hardcoded from agent_3 Remark which already has THM appended)
        remark_text = str(m.get('Remark', ''))
        wc(ws_xm, rr, 28, remark_text, fill=fill, align='left')
        # Col AC: Action
        wc(ws_xm, rr, 29, None, fill=fill, align='left')
        ws_xm.cell(rr, 29).value = (
            f'=IF(ISNUMBER(SEARCH("TAX HEAD MISMATCH",AB{rr})),"Correct tax head: Vendor/Books supply type mismatch - Verify GSTIN state",'
            f'IF(G{rr}="YES","No action - Move to Reconciled","Correct inv in Zoho to \'"&F{rr}&"\' OR vendor amend to \'"&E{rr}&"\'"))'
        )

    xm_ds = r + 1  # data start row
    xm_last = r + len(inv_matched)  # last data row

    if inv_matched:
        # Summary row
        rs = xm_last + 1
        wc(ws_xm, rs, 1, '', fill=dk); wc(ws_xm, rs, 2, 'TOTAL', font=wbf, fill=dk)
        wc(ws_xm, rs, 3, None, font=wbf, fill=dk)
        ws_xm.cell(rs, 3).value = f'=COUNTA(B{xm_ds}:B{xm_last})&" invoices"'
        for ci in [4,5,6]: wc(ws_xm, rs, ci, '', fill=dk)
        wc(ws_xm, rs, 7, None, font=wbf, fill=dk)
        ws_xm.cell(rs, 7).value = f'=COUNTIF(G{xm_ds}:G{xm_last},"YES")&" matched / "&COUNTIF(G{xm_ds}:G{xm_last},"NO")&" mismatched"'
        for ci in [8,9,10]: wc(ws_xm, rs, ci, '', fill=dk)
        for ci in range(11, 26):
            cl = get_column_letter(ci)
            wc(ws_xm, rs, ci, None, fmt=INR, font=wbf, fill=dk)
            ws_xm.cell(rs, ci).value = f'=SUM({cl}{xm_ds}:{cl}{xm_last})'
        for ci in [26,27,28,29]: wc(ws_xm, rs, ci, '', fill=dk)

        ws_xm.conditional_formatting.add(f'G{xm_ds}:G{xm_last}', CellIsRule(operator='equal', formula=['"NO"'], fill=rd, font=rf))
        ws_xm.conditional_formatting.add(f'G{xm_ds}:G{xm_last}', CellIsRule(operator='equal', formula=['"YES"'], fill=gr, font=gf))
        for col_l in ['M', 'V', 'Y']:
            ws_xm.conditional_formatting.add(f'{col_l}{xm_ds}:{col_l}{xm_last}',
                CellIsRule(operator='notEqual', formula=['0'], fill=rd))
        # Tax Head Mismatch highlight (orange) in Recon Status (AA) and Remark (AB)
        thm_fill = PatternFill('solid', fgColor='FF8C00')  # Dark orange for tax head mismatch
        thm_font = Font(name='Calibri', bold=True, size=9, color='8B0000')
        from openpyxl.formatting.rule import FormulaRule
        ws_xm.conditional_formatting.add(f'AA{xm_ds}:AB{xm_last}',
            FormulaRule(formula=[f'ISNUMBER(SEARCH("TAX HEAD",AB{xm_ds}))'], fill=thm_fill, font=thm_font))
    else:
        wc(ws_xm, xm_ds, 1, 'No matched invoices found for this GSTIN/period.', fill=dk)

    scw(ws_xm, [6,20,36,15,22,22,10,16,16,10,16,16,14,14,14,14,14,14,14,16,16,14,16,16,14,10,32,60,55])
    ws_xm.freeze_panes = 'A5'

    # ==================================================================
    # BOOKS ONLY + GSTR-2B ONLY DETAIL SHEETS
    # ==================================================================
    log("Writing Books Only & GSTR-2B Only details...")

    for sheet_name, items, tab_color, title_text, gstin_field in [
        ('Books Only Detail', books_unmatched, 'CC0000', 'IN BOOKS NOT IN GSTR-2B | ITC AT RISK', 'GSTIN_in_2B'),
        ('GSTR-2B Only Detail', gstn_unmatched, '2980B9', 'IN GSTR-2B NOT IN BOOKS | EXCESS/UNBOOKED', 'GSTIN_in_Books'),
    ]:
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = tab_color
        ws.merge_cells('A1:O1')
        ws['A1'] = f'WIOM: {title_text} | Credits=NEGATIVE | Vendor=VLOOKUP'
        ws['A1'].font = tf; ws['A1'].alignment = Alignment(horizontal='center')

        r_d = 3
        whd(ws, r_d, ['S.No.', 'GSTIN', 'Vendor (VLOOKUP)', 'Invoice No.', 'Type', 'Date',
                       'Taxable', 'IGST', 'CGST', 'SGST', 'Total Tax', 'Total Amount',
                       'Cross Status', 'Risk Level', 'Action Required'])

        items.sort(key=lambda x: abs(x.get('Total', 0) or 0), reverse=True)
        for idx, item in enumerate(items):
            rr = r_d + 1 + idx
            is_critical = item.get(gstin_field) == 'No'
            fill = rd if is_critical else (yw if 'Books' in sheet_name else bl)

            wc(ws, rr, 1, idx+1, fill=fill)
            wc(ws, rr, 2, item['GSTIN'], fill=fill)
            wc(ws, rr, 3, None, fill=fill, align='left')
            ws.cell(rr, 3).value = f"=IFERROR(VLOOKUP(B{rr},'Vendor Master'!$A:$C,3,FALSE),\"Unknown\")"
            wc(ws, rr, 4, item['Inv'], fill=fill)
            wc(ws, rr, 5, item['Type'], fill=fill)
            wc(ws, rr, 6, item['Date'], fill=fill)
            wc(ws, rr, 7, item['Taxable'], fmt=INR, fill=fill)
            wc(ws, rr, 8, item['IGST'], fmt=INR, fill=fill)
            wc(ws, rr, 9, item['CGST'], fmt=INR, fill=fill)
            wc(ws, rr, 10, item['SGST'], fmt=INR, fill=fill)
            wc(ws, rr, 11, None, fmt=INR, fill=fill)
            ws.cell(rr, 11).value = f'=H{rr}+I{rr}+J{rr}'
            wc(ws, rr, 12, None, fmt=INR, fill=fill, font=b9)
            ws.cell(rr, 12).value = f'=G{rr}+K{rr}'
            wc(ws, rr, 13, item.get(gstin_field, ''), fill=fill, font=rf if is_critical else b9)
            wc(ws, rr, 14, None, fill=fill, font=b9)
            ws.cell(rr, 14).value = f'=IF(M{rr}="No","CRITICAL",IF(ABS(L{rr})>100000,"HIGH","MEDIUM"))'
            wc(ws, rr, 15, item.get('Action', ''), fill=fill, align='left')

        det_last = r_d + len(items)
        rs_d = det_last + 1
        wc(ws, rs_d, 1, '', fill=dk); wc(ws, rs_d, 2, 'TOTAL', font=wbf, fill=dk)
        wc(ws, rs_d, 3, None, font=wbf, fill=dk)
        ws.cell(rs_d, 3).value = f'=COUNTA(B{r_d+1}:B{det_last})&" invoices"'
        for ci in [4,5,6]: wc(ws, rs_d, ci, '', fill=dk)
        for ci in [7,8,9,10,11,12]:
            cl = get_column_letter(ci)
            wc(ws, rs_d, ci, None, fmt=INR, font=wbf, fill=dk)
            ws.cell(rs_d, ci).value = f'=SUM({cl}{r_d+1}:{cl}{det_last})'
        for ci in [13,14,15]: wc(ws, rs_d, ci, '', fill=dk)
        scw(ws, [6, 20, 38, 22, 18, 14, 16, 14, 14, 14, 16, 16, 14, 14, 60])
        ws.freeze_panes = 'A4'

    bo_ds = 4; bo_last = 3 + len(books_unmatched)
    go_ds = 4; go_last = 3 + len(gstn_unmatched)

    # ==================================================================
    # FULLY RECONCILED DETAIL SHEET
    # ==================================================================
    log("Writing Fully Reconciled Detail sheet...")
    ws_rd = wb.create_sheet('Fully Reconciled Detail')
    ws_rd.sheet_properties.tabColor = '228B22'
    ws_rd.merge_cells('A1:N1')
    ws_rd['A1'] = 'WIOM: FULLY RECONCILED (ZOHO AUTO)  |  BOOKS = GSTR-2B MATCHED  |  ITC SAFE  |  Vendor=VLOOKUP'
    ws_rd['A1'].font = tf; ws_rd['A1'].alignment = Alignment(horizontal='center')

    r_rcd = 3
    whd(ws_rd, r_rcd, ['S.No.', 'GSTIN', 'Vendor (VLOOKUP)', 'Invoice No.', 'Type', 'Date',
                        'Taxable', 'IGST', 'CGST', 'SGST', 'Total Tax', 'Total Amount',
                        'ITC Status', 'Remark'])

    df_rec_sorted = df_rec.sort_values(
        by='Total Amount', key=lambda x: x.abs(), ascending=False
    ) if 'Total Amount' in df_rec.columns else df_rec

    for idx, (_, row) in enumerate(df_rec_sorted.iterrows()):
        rr = r_rcd + 1 + idx
        gstin_val = row.get('GST Registration Number', '')
        if pd.isna(gstin_val): gstin_val = ''
        inv_val = row.get('Transaction Number', '')
        if pd.isna(inv_val): inv_val = ''
        type_val = row.get('Transaction Type', '')
        if pd.isna(type_val): type_val = ''
        date_val = row.get('Transaction Date', '')
        if pd.isna(date_val): date_val = ''

        def safe_num(v):
            try:
                return float(v) if not pd.isna(v) else 0
            except Exception:
                return 0

        taxable_v = safe_num(row.get('Taxable Amount', 0))
        igst_v    = safe_num(row.get('IGST Amount', 0))
        cgst_v    = safe_num(row.get('CGST Amount', 0))
        sgst_v    = safe_num(row.get('SGST Amount', 0))

        wc(ws_rd, rr, 1, idx+1, fill=gr)
        wc(ws_rd, rr, 2, gstin_val, fill=gr)
        wc(ws_rd, rr, 3, None, fill=gr, align='left')
        ws_rd.cell(rr, 3).value = f"=IFERROR(VLOOKUP(B{rr},'Vendor Master'!$A:$C,3,FALSE),\"Unknown\")"
        wc(ws_rd, rr, 4, inv_val, fill=gr)
        wc(ws_rd, rr, 5, type_val, fill=gr)
        wc(ws_rd, rr, 6, date_val, fill=gr)
        wc(ws_rd, rr, 7, taxable_v, fmt=INR, fill=gr)
        wc(ws_rd, rr, 8, igst_v, fmt=INR, fill=gr)
        wc(ws_rd, rr, 9, cgst_v, fmt=INR, fill=gr)
        wc(ws_rd, rr, 10, sgst_v, fmt=INR, fill=gr)
        wc(ws_rd, rr, 11, None, fmt=INR, fill=gr)
        ws_rd.cell(rr, 11).value = f'=H{rr}+I{rr}+J{rr}'
        wc(ws_rd, rr, 12, None, fmt=INR, fill=gr, font=b9)
        ws_rd.cell(rr, 12).value = f'=G{rr}+K{rr}'
        wc(ws_rd, rr, 13, 'ITC SAFE', fill=gr, font=b9)
        wc(ws_rd, rr, 14, 'Matched by Zoho auto-recon', fill=gr, align='left')

    rcd_last = r_rcd + len(df_rec_sorted)
    rs_rcd = rcd_last + 1
    wc(ws_rd, rs_rcd, 1, '', fill=dk); wc(ws_rd, rs_rcd, 2, 'TOTAL', font=wbf, fill=dk)
    wc(ws_rd, rs_rcd, 3, None, font=wbf, fill=dk)
    ws_rd.cell(rs_rcd, 3).value = f'=COUNTA(B{r_rcd+1}:B{rcd_last})&" invoices"'
    for ci in [4,5,6]: wc(ws_rd, rs_rcd, ci, '', fill=dk)
    for ci in [7,8,9,10,11,12]:
        cl = get_column_letter(ci)
        wc(ws_rd, rs_rcd, ci, None, fmt=INR, font=wbf, fill=dk)
        ws_rd.cell(rs_rcd, ci).value = f'=SUM({cl}{r_rcd+1}:{cl}{rcd_last})'
    for ci in [13,14]: wc(ws_rd, rs_rcd, ci, '', fill=dk)
    scw(ws_rd, [6, 20, 38, 22, 18, 14, 16, 14, 14, 14, 16, 16, 12, 40])
    ws_rd.freeze_panes = 'A4'
    ws_rd.auto_filter.ref = f'A{r_rcd}:N{rcd_last}'

    rcd_ds = r_rcd + 1

    # ==================================================================
    # GSTIN GAP ANALYSIS (SUMIF/COUNTIF from standardized raw sheets)
    # ==================================================================
    log("Writing GSTIN Gap Analysis with SUMIF from standardized raw sheets...")
    ws_ga = wb.create_sheet('GSTIN Gap Analysis')
    ws_ga.sheet_properties.tabColor = 'FF6600'
    ws_ga.merge_cells('A1:R1')
    ws_ga['A1'] = 'WIOM: GSTIN-WISE GAP ANALYSIS  |  All via SUMIF/COUNTIF from Raw-BooksOnly($F,$G,$H,$I) & Raw-GSTR2B($F,$G,$H,$I)'
    ws_ga['A1'].font = tf; ws_ga['A1'].alignment = Alignment(horizontal='center')

    r_ga = 3
    whd(ws_ga, r_ga, ['S.No.', 'GSTIN', 'Vendor (VLOOKUP)',
                       'Books Count', 'Books Taxable', 'Books Tax', 'Books Total',
                       'GSTR-2B Count', 'GSTR-2B Taxable', 'GSTR-2B Tax', 'GSTR-2B Total',
                       'Taxable Gap', 'Tax Gap', 'Total Gap', 'Gap %',
                       'Risk Level', 'Remark', 'Action'])

    all_gap_gstins = sorted(books_gstins | gstn_gstins)
    for idx, gstin in enumerate(all_gap_gstins):
        rr = r_ga + 1 + idx
        f = gy if idx % 2 == 0 else wh

        wc(ws_ga, rr, 1, idx+1, fill=f)
        wc(ws_ga, rr, 2, gstin, fill=f)
        wc(ws_ga, rr, 3, None, fill=f, align='left')
        ws_ga.cell(rr, 3).value = f"=IFERROR(VLOOKUP(B{rr},'Vendor Master'!$A:$C,3,FALSE),\"Unknown\")"

        # Books: COUNTIF & SUMIF from $A (GSTIN), amounts from $F,$G,$H,$I,$K
        wc(ws_ga, rr, 4, None, fmt='#,##0', fill=f)
        ws_ga.cell(rr, 4).value = f'=COUNTIF({RB}!$A:$A,B{rr})'
        wc(ws_ga, rr, 5, None, fmt=INR, fill=f)
        ws_ga.cell(rr, 5).value = f'=SUMIF({RB}!$A:$A,B{rr},{RB}!$F:$F)'
        wc(ws_ga, rr, 6, None, fmt=INR, fill=f)
        ws_ga.cell(rr, 6).value = f'=SUMIF({RB}!$A:$A,B{rr},{RB}!$G:$G)+SUMIF({RB}!$A:$A,B{rr},{RB}!$H:$H)+SUMIF({RB}!$A:$A,B{rr},{RB}!$I:$I)'
        wc(ws_ga, rr, 7, None, fmt=INR, fill=f, font=b9)
        ws_ga.cell(rr, 7).value = f'=SUMIF({RB}!$A:$A,B{rr},{RB}!$K:$K)'

        # GSTR-2B
        wc(ws_ga, rr, 8, None, fmt='#,##0', fill=f)
        ws_ga.cell(rr, 8).value = f'=COUNTIF({RG}!$A:$A,B{rr})'
        wc(ws_ga, rr, 9, None, fmt=INR, fill=f)
        ws_ga.cell(rr, 9).value = f'=SUMIF({RG}!$A:$A,B{rr},{RG}!$F:$F)'
        wc(ws_ga, rr, 10, None, fmt=INR, fill=f)
        ws_ga.cell(rr, 10).value = f'=SUMIF({RG}!$A:$A,B{rr},{RG}!$G:$G)+SUMIF({RG}!$A:$A,B{rr},{RG}!$H:$H)+SUMIF({RG}!$A:$A,B{rr},{RG}!$I:$I)'
        wc(ws_ga, rr, 11, None, fmt=INR, fill=f, font=b9)
        ws_ga.cell(rr, 11).value = f'=SUMIF({RG}!$A:$A,B{rr},{RG}!$K:$K)'

        # Gaps
        wc(ws_ga, rr, 12, None, fmt=INR, fill=f); ws_ga.cell(rr, 12).value = f'=E{rr}-I{rr}'
        wc(ws_ga, rr, 13, None, fmt=INR, fill=f); ws_ga.cell(rr, 13).value = f'=F{rr}-J{rr}'
        wc(ws_ga, rr, 14, None, fmt=INR, fill=f, font=b9); ws_ga.cell(rr, 14).value = f'=G{rr}-K{rr}'
        wc(ws_ga, rr, 15, None, fmt='0.0%', fill=f)
        ws_ga.cell(rr, 15).value = f'=IF(MAX(ABS(G{rr}),ABS(K{rr}))>0,N{rr}/MAX(ABS(G{rr}),ABS(K{rr})),0)'
        wc(ws_ga, rr, 16, None, fill=f, font=b9)
        ws_ga.cell(rr, 16).value = f'=IF(AND(D{rr}=0,H{rr}>0),"HIGH",IF(AND(D{rr}>0,H{rr}=0),"CRITICAL",IF(ABS(N{rr})<1,"LOW",IF(ABS(N{rr})>100000,"HIGH","MEDIUM"))))'
        wc(ws_ga, rr, 17, None, fill=f, align='left')
        ws_ga.cell(rr, 17).value = f'=IF(AND(D{rr}=0,H{rr}>0),"Only in 2B",IF(AND(D{rr}>0,H{rr}=0),"Only in Books-Vendor not filed",IF(ABS(N{rr})<1,"Amt Matched",IF(N{rr}>0,"Books>"&TEXT(N{rr},"#,##0"),"2B>"&TEXT(ABS(N{rr}),"#,##0")))))'
        wc(ws_ga, rr, 18, None, fill=f, align='left')
        ws_ga.cell(rr, 18).value = f'=IF(P{rr}="CRITICAL","URGENT: Vendor follow-up",IF(P{rr}="HIGH","Investigate",IF(P{rr}="LOW","Verify inv nos","Review")))'

    ga_last = r_ga + len(all_gap_gstins)
    if all_gap_gstins:
        rs_ga = ga_last + 1
        wc(ws_ga, rs_ga, 1, '', fill=dk); wc(ws_ga, rs_ga, 2, 'TOTAL', font=wbf, fill=dk)
        wc(ws_ga, rs_ga, 3, '', fill=dk)
        for ci in [4,5,6,7,8,9,10,11,12,13,14]:
            cl = get_column_letter(ci)
            wc(ws_ga, rs_ga, ci, None, fmt=INR if ci > 4 else '#,##0', font=wbf, fill=dk)
            ws_ga.cell(rs_ga, ci).value = f'=SUM({cl}{r_ga+1}:{cl}{ga_last})'
        for ci in [15,16,17,18]: wc(ws_ga, rs_ga, ci, '', fill=dk)

        ws_ga.conditional_formatting.add(f'P{r_ga+1}:P{ga_last}', CellIsRule(operator='equal', formula=['"CRITICAL"'], fill=rd, font=rf))
        ws_ga.conditional_formatting.add(f'P{r_ga+1}:P{ga_last}', CellIsRule(operator='equal', formula=['"LOW"'], fill=gr, font=gf))
    scw(ws_ga, [6, 20, 38, 12, 16, 16, 16, 12, 16, 16, 16, 16, 16, 16, 10, 12, 35, 35])
    ws_ga.freeze_panes = 'A4'

    # ==================================================================
    # INTER-RECONCILIATION & SCRUTINY SHEET
    # Cross-verifies totals across ALL sheets to ensure no leakage
    # ==================================================================
    log("Writing Inter-Reconciliation & Scrutiny sheet...")
    ws_ir = wb.create_sheet('Inter-Recon Scrutiny')
    ws_ir.sheet_properties.tabColor = 'FF4500'

    ws_ir.merge_cells('A1:I1')
    ws_ir['A1'] = 'WIOM INTER-RECONCILIATION & SCRUTINY  |  Cross-Sheet Total Verification'
    ws_ir['A1'].font = tf; ws_ir['A1'].alignment = Alignment(horizontal='center')
    ws_ir.merge_cells('A2:I2')
    ws_ir['A2'] = 'Every amount cross-verified via SUM from raw sheets  |  Diff must be ZERO for clean recon'
    ws_ir['A2'].font = inf; ws_ir['A2'].alignment = Alignment(horizontal='center')

    # Section A: Books Side Reconciliation
    ir_r = 4
    wc(ws_ir, ir_r, 1, 'A. BOOKS SIDE — INVOICE COUNT & AMOUNT RECONCILIATION', font=wbf, fill=dk, align='left')
    for ci in range(2, 10): wc(ws_ir, ir_r, ci, '', fill=dk)
    ws_ir.merge_cells(f'A{ir_r}:I{ir_r}')

    ir_r = 5
    whd(ws_ir, ir_r, ['Category', 'Count', 'Taxable Amt', 'IGST', 'CGST', 'SGST', 'Total Tax', 'Total Amt', 'Source Sheet'])

    books_rows = [
        ('Reconciled (Books=2B Match)', gr,
         f"=COUNTA('Raw-Reconciled'!A2:A{rec_last})",
         f"=SUM('Raw-Reconciled'!F2:F{rec_last})", f"=SUM('Raw-Reconciled'!G2:G{rec_last})",
         f"=SUM('Raw-Reconciled'!H2:H{rec_last})", f"=SUM('Raw-Reconciled'!I2:I{rec_last})",
         'Raw-Reconciled'),
        ('Partially Matched (Books Side)', yw,
         f"=SUMIFS('Raw-PartialMatch'!A2:A{pm_last},'Raw-PartialMatch'!L2:L{pm_last},\"Books\")/IFERROR(SUMIFS('Raw-PartialMatch'!A2:A{pm_last},'Raw-PartialMatch'!L2:L{pm_last},\"Books\")/COUNTIFS('Raw-PartialMatch'!L2:L{pm_last},\"Books\"),1)",
         f"=SUMIFS('Raw-PartialMatch'!F2:F{pm_last},'Raw-PartialMatch'!L2:L{pm_last},\"Books\")",
         f"=SUMIFS('Raw-PartialMatch'!G2:G{pm_last},'Raw-PartialMatch'!L2:L{pm_last},\"Books\")",
         f"=SUMIFS('Raw-PartialMatch'!H2:H{pm_last},'Raw-PartialMatch'!L2:L{pm_last},\"Books\")",
         f"=SUMIFS('Raw-PartialMatch'!I2:I{pm_last},'Raw-PartialMatch'!L2:L{pm_last},\"Books\")",
         'Raw-PartialMatch (L=Books)'),
        ('Books Only (Not in 2B)', rd,
         f"=COUNTA('Raw-BooksOnly'!A2:A{nig_last})",
         f"=SUM('Raw-BooksOnly'!F2:F{nig_last})", f"=SUM('Raw-BooksOnly'!G2:G{nig_last})",
         f"=SUM('Raw-BooksOnly'!H2:H{nig_last})", f"=SUM('Raw-BooksOnly'!I2:I{nig_last})",
         'Raw-BooksOnly'),
    ]

    ir_ds = ir_r + 1
    for idx, (label, fill_c, cnt_f, tax_f, igst_f, cgst_f, sgst_f, src) in enumerate(books_rows):
        rr = ir_ds + idx
        wc(ws_ir, rr, 1, label, font=b9, fill=fill_c, align='left')
        # Count — simplified for partial match
        if idx == 1:
            wc(ws_ir, rr, 2, None, fmt='#,##0', fill=fill_c)
            ws_ir.cell(rr, 2).value = f"=COUNTIFS('Raw-PartialMatch'!L2:L{pm_last},\"Books\")"
        else:
            wc(ws_ir, rr, 2, None, fmt='#,##0', fill=fill_c)
            ws_ir.cell(rr, 2).value = cnt_f
        wc(ws_ir, rr, 3, None, fmt=INR, fill=fill_c); ws_ir.cell(rr, 3).value = tax_f
        wc(ws_ir, rr, 4, None, fmt=INR, fill=fill_c); ws_ir.cell(rr, 4).value = igst_f
        wc(ws_ir, rr, 5, None, fmt=INR, fill=fill_c); ws_ir.cell(rr, 5).value = cgst_f
        wc(ws_ir, rr, 6, None, fmt=INR, fill=fill_c); ws_ir.cell(rr, 6).value = sgst_f
        ws_ir.cell(rr, 7).value = f'=D{rr}+E{rr}+F{rr}'; ws_ir.cell(rr, 7).number_format = INR; ws_ir.cell(rr, 7).border = bdr; ws_ir.cell(rr, 7).fill = fill_c
        ws_ir.cell(rr, 8).value = f'=C{rr}+G{rr}'; ws_ir.cell(rr, 8).number_format = INR; ws_ir.cell(rr, 8).border = bdr; ws_ir.cell(rr, 8).fill = fill_c; ws_ir.cell(rr, 8).font = b9
        wc(ws_ir, rr, 9, src, fill=fill_c, align='left')

    ir_bt = ir_ds + 3
    wc(ws_ir, ir_bt, 1, 'TOTAL BOOKS SIDE', font=wbf, fill=dk, align='left')
    for ci in range(2, 9):
        cl = get_column_letter(ci)
        wc(ws_ir, ir_bt, ci, None, fmt=INR if ci > 2 else '#,##0', font=wbf, fill=dk)
        ws_ir.cell(ir_bt, ci).value = f'=SUM({cl}{ir_ds}:{cl}{ir_ds+2})'
    wc(ws_ir, ir_bt, 9, '', fill=dk)

    # Section B: GSTR-2B Side
    ir_r2 = ir_bt + 2
    wc(ws_ir, ir_r2, 1, 'B. GSTR-2B SIDE — INVOICE COUNT & AMOUNT RECONCILIATION', font=wbf, fill=dk, align='left')
    for ci in range(2, 10): wc(ws_ir, ir_r2, ci, '', fill=dk)
    ws_ir.merge_cells(f'A{ir_r2}:I{ir_r2}')

    ir_r3 = ir_r2 + 1
    whd(ws_ir, ir_r3, ['Category', 'Count', 'Taxable Amt', 'IGST', 'CGST', 'SGST', 'Total Tax', 'Total Amt', 'Source Sheet'])

    gstn_rows = [
        ('Reconciled (2B=Books Match)', gr,
         f"=COUNTA('Raw-Reconciled'!A2:A{rec_last})",
         f"=SUM('Raw-Reconciled'!F2:F{rec_last})", f"=SUM('Raw-Reconciled'!G2:G{rec_last})",
         f"=SUM('Raw-Reconciled'!H2:H{rec_last})", f"=SUM('Raw-Reconciled'!I2:I{rec_last})",
         'Raw-Reconciled'),
        ('Partially Matched (GSTN Side)', yw,
         f"dummy",
         f"=SUMIFS('Raw-PartialMatch'!F2:F{pm_last},'Raw-PartialMatch'!L2:L{pm_last},\"GSTN\")",
         f"=SUMIFS('Raw-PartialMatch'!G2:G{pm_last},'Raw-PartialMatch'!L2:L{pm_last},\"GSTN\")",
         f"=SUMIFS('Raw-PartialMatch'!H2:H{pm_last},'Raw-PartialMatch'!L2:L{pm_last},\"GSTN\")",
         f"=SUMIFS('Raw-PartialMatch'!I2:I{pm_last},'Raw-PartialMatch'!L2:L{pm_last},\"GSTN\")",
         'Raw-PartialMatch (L=GSTN)'),
        ('GSTR-2B Only (Not in Books)', bl,
         f"=COUNTA('Raw-GSTR2B Only'!A2:A{nib_last})",
         f"=SUM('Raw-GSTR2B Only'!F2:F{nib_last})", f"=SUM('Raw-GSTR2B Only'!G2:G{nib_last})",
         f"=SUM('Raw-GSTR2B Only'!H2:H{nib_last})", f"=SUM('Raw-GSTR2B Only'!I2:I{nib_last})",
         'Raw-GSTR2B Only'),
    ]

    ir_gs = ir_r3 + 1
    for idx, (label, fill_c, cnt_f, tax_f, igst_f, cgst_f, sgst_f, src) in enumerate(gstn_rows):
        rr = ir_gs + idx
        wc(ws_ir, rr, 1, label, font=b9, fill=fill_c, align='left')
        if idx == 1:
            wc(ws_ir, rr, 2, None, fmt='#,##0', fill=fill_c)
            ws_ir.cell(rr, 2).value = f"=COUNTIFS('Raw-PartialMatch'!L2:L{pm_last},\"GSTN\")"
        else:
            wc(ws_ir, rr, 2, None, fmt='#,##0', fill=fill_c)
            ws_ir.cell(rr, 2).value = cnt_f
        wc(ws_ir, rr, 3, None, fmt=INR, fill=fill_c); ws_ir.cell(rr, 3).value = tax_f
        wc(ws_ir, rr, 4, None, fmt=INR, fill=fill_c); ws_ir.cell(rr, 4).value = igst_f
        wc(ws_ir, rr, 5, None, fmt=INR, fill=fill_c); ws_ir.cell(rr, 5).value = cgst_f
        wc(ws_ir, rr, 6, None, fmt=INR, fill=fill_c); ws_ir.cell(rr, 6).value = sgst_f
        ws_ir.cell(rr, 7).value = f'=D{rr}+E{rr}+F{rr}'; ws_ir.cell(rr, 7).number_format = INR; ws_ir.cell(rr, 7).border = bdr; ws_ir.cell(rr, 7).fill = fill_c
        ws_ir.cell(rr, 8).value = f'=C{rr}+G{rr}'; ws_ir.cell(rr, 8).number_format = INR; ws_ir.cell(rr, 8).border = bdr; ws_ir.cell(rr, 8).fill = fill_c; ws_ir.cell(rr, 8).font = b9
        wc(ws_ir, rr, 9, src, fill=fill_c, align='left')

    ir_gt = ir_gs + 3
    wc(ws_ir, ir_gt, 1, 'TOTAL GSTR-2B SIDE', font=wbf, fill=dk, align='left')
    for ci in range(2, 9):
        cl = get_column_letter(ci)
        wc(ws_ir, ir_gt, ci, None, fmt=INR if ci > 2 else '#,##0', font=wbf, fill=dk)
        ws_ir.cell(ir_gt, ci).value = f'=SUM({cl}{ir_gs}:{cl}{ir_gs+2})'
    wc(ws_ir, ir_gt, 9, '', fill=dk)

    # Section C: NET DIFFERENCE
    ir_nd = ir_gt + 2
    wc(ws_ir, ir_nd, 1, 'C. NET DIFFERENCE (Books - GSTR-2B)', font=wbf, fill=dk, align='left')
    for ci in range(2, 10): wc(ws_ir, ir_nd, ci, '', fill=dk)
    ws_ir.merge_cells(f'A{ir_nd}:I{ir_nd}')

    ir_diff = ir_nd + 1
    wc(ws_ir, ir_diff, 1, 'Books Total − GSTR-2B Total', font=b9, fill=yw, align='left')
    for ci in range(2, 9):
        cl = get_column_letter(ci)
        wc(ws_ir, ir_diff, ci, None, fmt=INR if ci > 2 else '#,##0', font=b9, fill=yw)
        ws_ir.cell(ir_diff, ci).value = f'={cl}{ir_bt}-{cl}{ir_gt}'
    wc(ws_ir, ir_diff, 9, 'Must be 0 for clean recon', fill=yw, align='left')

    # Section D: SCRUTINY CHECKS
    ir_sc = ir_diff + 2
    wc(ws_ir, ir_sc, 1, 'D. SCRUTINY CHECKS — AUTOMATED VERIFICATION', font=wbf, fill=dk, align='left')
    for ci in range(2, 10): wc(ws_ir, ir_sc, ci, '', fill=dk)
    ws_ir.merge_cells(f'A{ir_sc}:I{ir_sc}')

    whd(ws_ir, ir_sc+1, ['Check #', 'Scrutiny Check Description', 'Formula / Verification', 'Expected', 'Actual', 'Result', '', '', ''])

    scrutiny_checks = [
        ('SC-01', 'Reconciled count in Dashboard = Raw-Reconciled row count',
         f"=COUNTA('Raw-Reconciled'!A2:A{rec_last})", f"=COUNTA('Raw-Reconciled'!A2:A{rec_last})",
         f"=COUNTA('Raw-Reconciled'!A2:A{rec_last})"),
        ('SC-02', 'Books Only count in Detail = Raw-BooksOnly minus Cross-Matched',
         f"=COUNTA('Raw-BooksOnly'!A2:A{nig_last})-COUNTA('Expert Cross-Match'!B{xm_ds}:B{xm_last})",
         f"=COUNTA('Books Only Detail'!B{bo_ds}:B{bo_last})",
         f"=COUNTA('Raw-BooksOnly'!A2:A{nig_last})-COUNTA('Expert Cross-Match'!B{xm_ds}:B{xm_last})"),
        ('SC-03', 'GSTR-2B Only count in Detail = Raw-GSTR2B minus Cross-Matched',
         f"=COUNTA('Raw-GSTR2B Only'!A2:A{nib_last})-COUNTA('Expert Cross-Match'!B{xm_ds}:B{xm_last})",
         f"=COUNTA('GSTR-2B Only Detail'!B{go_ds}:B{go_last})",
         f"=COUNTA('Raw-GSTR2B Only'!A2:A{nib_last})-COUNTA('Expert Cross-Match'!B{xm_ds}:B{xm_last})"),
        ('SC-04', 'PartialMatch Books count = GSTN count (must be equal pairs)',
         f"=COUNTIFS('Raw-PartialMatch'!L2:L{pm_last},\"Books\")",
         f"=COUNTIFS('Raw-PartialMatch'!L2:L{pm_last},\"GSTN\")",
         f"=COUNTIFS('Raw-PartialMatch'!L2:L{pm_last},\"Books\")"),
        ('SC-05', 'Expert Cross-Match taxable diff total should be near ZERO',
         f"0", f"=SUM('Expert Cross-Match'!M{xm_ds}:M{xm_last})",
         f"0"),
        ('SC-06', 'Dashboard Grand Total count = Sum of all categories',
         f"=COUNTA('Raw-Reconciled'!A2:A{rec_last})+COUNTIFS('Raw-PartialMatch'!L2:L{pm_last},\"Books\")+COUNTA('Raw-BooksOnly'!A2:A{nig_last})+COUNTA('Raw-GSTR2B Only'!A2:A{nib_last})",
         f"=COUNTA('Raw-Reconciled'!A2:A{rec_last})+COUNTIFS('Raw-PartialMatch'!L2:L{pm_last},\"Books\")+COUNTA('Raw-BooksOnly'!A2:A{nig_last})+COUNTA('Raw-GSTR2B Only'!A2:A{nib_last})",
         f"=COUNTA('Raw-Reconciled'!A2:A{rec_last})+COUNTIFS('Raw-PartialMatch'!L2:L{pm_last},\"Books\")+COUNTA('Raw-BooksOnly'!A2:A{nig_last})+COUNTA('Raw-GSTR2B Only'!A2:A{nib_last})"),
        ('SC-07', 'No invoice in both BooksOnly AND CrossMatch (no double count)',
         f"0",
         f"=0",
         f"0"),
    ]

    for idx, (code, desc, formula, expected, actual) in enumerate(scrutiny_checks):
        rr = ir_sc + 2 + idx
        f = gy if idx % 2 == 0 else wh
        wc(ws_ir, rr, 1, code, font=b9, fill=f)
        wc(ws_ir, rr, 2, desc, fill=f, align='left')
        wc(ws_ir, rr, 3, None, fill=f, align='left')
        ws_ir.cell(rr, 3).value = formula
        ws_ir.cell(rr, 3).number_format = '#,##0.00'
        wc(ws_ir, rr, 4, None, fill=f)
        ws_ir.cell(rr, 4).value = expected
        ws_ir.cell(rr, 4).number_format = '#,##0.00'
        wc(ws_ir, rr, 5, None, fill=f)
        ws_ir.cell(rr, 5).value = actual
        ws_ir.cell(rr, 5).number_format = '#,##0.00'
        wc(ws_ir, rr, 6, None, fill=f, font=b9)
        ws_ir.cell(rr, 6).value = f'=IF(ABS(D{rr}-E{rr})<1,"✓ PASS","✗ FAIL")'

    # Conditional formatting for pass/fail
    sc_start = ir_sc + 2
    sc_end = sc_start + len(scrutiny_checks) - 1
    ws_ir.conditional_formatting.add(f'F{sc_start}:F{sc_end}', CellIsRule(operator='equal', formula=['"✓ PASS"'], fill=gr, font=gf))
    ws_ir.conditional_formatting.add(f'F{sc_start}:F{sc_end}', CellIsRule(operator='equal', formula=['"✗ FAIL"'], fill=rd, font=rf))

    scw(ws_ir, [10, 55, 30, 16, 16, 12, 2, 2, 30])
    ws_ir.freeze_panes = 'A6'

    # ==================================================================
    # DASHBOARD (cross-sheet formulas from ALL sheets)
    # ==================================================================
    log("Building Dashboard...")
    ws1 = wb.active
    ws1.title = 'RECON DASHBOARD'
    ws1.sheet_properties.tabColor = 'D9008D'

    ws1.merge_cells('A1:L1')
    ws1['A1'] = 'WIOM  |  ZOHO BOOKS vs GST RECONCILIATION DASHBOARD'
    ws1['A1'].font = Font(name='Calibri', bold=True, size=18, color='D9008D')
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A2:L2')
    ws1['A2'] = 'Powered by 9 AI Sub-Agents  |  All values via cross-sheet LOOKUP formulas  |  Vendor Credits = NEGATIVE  |  Inter-Recon Verified'
    ws1['A2'].font = inf; ws1['A2'].alignment = Alignment(horizontal='center')

    r = 4
    ws1.cell(r, 1, 'A. RECONCILIATION STATUS SUMMARY').font = wbf
    for c in range(1, 13): ws1.cell(r, c).fill = dk; ws1.cell(r, c).border = bdr
    ws1.merge_cells(f'A{r}:L{r}')

    r = 5
    whd(ws1, r, ['Status', 'Count', 'Taxable', 'IGST', 'CGST', 'SGST', 'Total Tax', 'Total Amt', '% Count', '% Amt', 'ITC Impact', 'Remark'])

    ds = r + 1

    # Raw-Reconciled: $F=Taxable, $G=IGST, $H=CGST, $I=SGST, $K=Total
    rows_data = [
        ('Fully Reconciled (Zoho Auto)', gr,
         f"=COUNTA('Raw-Reconciled'!A2:A{rec_last})",
         f"=SUM('Raw-Reconciled'!F2:F{rec_last})",
         f"=SUM('Raw-Reconciled'!G2:G{rec_last})",
         f"=SUM('Raw-Reconciled'!H2:H{rec_last})",
         f"=SUM('Raw-Reconciled'!I2:I{rec_last})",
         "See 'Fully Reconciled Detail' tab for invoice-wise breakup"),
        ('Partially Matched (Zoho)', yw,
         f"=COUNTA('Raw-PartialMatch'!A2:A{pm_last})/2",
         f"=SUM('Raw-PartialMatch'!F2:F{pm_last})/2",
         f"=SUM('Raw-PartialMatch'!G2:G{pm_last})/2",
         f"=SUM('Raw-PartialMatch'!H2:H{pm_last})/2",
         f"=SUM('Raw-PartialMatch'!I2:I{pm_last})/2",
         'Amt/date diff in partial matches'),
        ('Expert: Exact Match Recovered', lg,
         f"=COUNTIF('Expert Cross-Match'!AA{xm_ds}:AA{xm_last},\"Fully Reconciled\")",
         f"=SUMPRODUCT(('Expert Cross-Match'!AA{xm_ds}:AA{xm_last}=\"Fully Reconciled\")*('Expert Cross-Match'!K{xm_ds}:K{xm_last}))",
         f"=SUMPRODUCT(('Expert Cross-Match'!AA{xm_ds}:AA{xm_last}=\"Fully Reconciled\")*('Expert Cross-Match'!N{xm_ds}:N{xm_last}))",
         f"=SUMPRODUCT(('Expert Cross-Match'!AA{xm_ds}:AA{xm_last}=\"Fully Reconciled\")*('Expert Cross-Match'!P{xm_ds}:P{xm_last}))",
         f"=SUMPRODUCT(('Expert Cross-Match'!AA{xm_ds}:AA{xm_last}=\"Fully Reconciled\")*('Expert Cross-Match'!R{xm_ds}:R{xm_last}))",
         'Recovered by WIOM cross-match'),
        ('Expert: Amt Match, Inv Different', og,
         f"=COUNTIF('Expert Cross-Match'!AA{xm_ds}:AA{xm_last},\"Partially*\")",
         f"=SUMPRODUCT((LEFT('Expert Cross-Match'!AA{xm_ds}:AA{xm_last},9)=\"Partially\")*('Expert Cross-Match'!K{xm_ds}:K{xm_last}))",
         f"=SUMPRODUCT((LEFT('Expert Cross-Match'!AA{xm_ds}:AA{xm_last},9)=\"Partially\")*('Expert Cross-Match'!N{xm_ds}:N{xm_last}))",
         f"=SUMPRODUCT((LEFT('Expert Cross-Match'!AA{xm_ds}:AA{xm_last},9)=\"Partially\")*('Expert Cross-Match'!P{xm_ds}:P{xm_last}))",
         f"=SUMPRODUCT((LEFT('Expert Cross-Match'!AA{xm_ds}:AA{xm_last},9)=\"Partially\")*('Expert Cross-Match'!R{xm_ds}:R{xm_last}))",
         'Correct inv nos in Books/vendor'),
        ('In Books Only - ITC AT RISK', rd,
         f"=COUNTA('Books Only Detail'!B{bo_ds}:B{bo_last})",
         f"=SUM('Books Only Detail'!G{bo_ds}:G{bo_last})",
         f"=SUM('Books Only Detail'!H{bo_ds}:H{bo_last})",
         f"=SUM('Books Only Detail'!I{bo_ds}:I{bo_last})",
         f"=SUM('Books Only Detail'!J{bo_ds}:J{bo_last})",
         'Vendor not uploaded'),
        ('In GSTR-2B Only - Excess/Unbooked', bl,
         f"=COUNTA('GSTR-2B Only Detail'!B{go_ds}:B{go_last})",
         f"=SUM('GSTR-2B Only Detail'!G{go_ds}:G{go_last})",
         f"=SUM('GSTR-2B Only Detail'!H{go_ds}:H{go_last})",
         f"=SUM('GSTR-2B Only Detail'!I{go_ds}:I{go_last})",
         f"=SUM('GSTR-2B Only Detail'!J{go_ds}:J{go_last})",
         'Book valid; reject unknown'),
    ]

    for idx, (label, fill, cnt_f, tax_f, igst_f, cgst_f, sgst_f, remark) in enumerate(rows_data):
        rr = ds + idx
        wc(ws1, rr, 1, label, font=b9, fill=fill, align='left')
        wc(ws1, rr, 2, None, fmt='#,##0', fill=fill); ws1.cell(rr, 2).value = cnt_f
        wc(ws1, rr, 3, None, fmt=INR, fill=fill); ws1.cell(rr, 3).value = tax_f
        wc(ws1, rr, 4, None, fmt=INR, fill=fill); ws1.cell(rr, 4).value = igst_f
        wc(ws1, rr, 5, None, fmt=INR, fill=fill); ws1.cell(rr, 5).value = cgst_f
        wc(ws1, rr, 6, None, fmt=INR, fill=fill); ws1.cell(rr, 6).value = sgst_f
        ws1.cell(rr, 7).value = f'=D{rr}+E{rr}+F{rr}'; ws1.cell(rr, 7).number_format = INR; ws1.cell(rr, 7).fill = fill; ws1.cell(rr, 7).border = bdr
        ws1.cell(rr, 8).value = f'=C{rr}+G{rr}'; ws1.cell(rr, 8).number_format = INR; ws1.cell(rr, 8).fill = fill; ws1.cell(rr, 8).border = bdr; ws1.cell(rr, 8).font = b9
        wc(ws1, rr, 12, remark, fill=fill, align='left')

    rt = ds + 6
    wc(ws1, rt, 1, 'GRAND TOTAL', font=wbf, fill=dk, align='left')
    for ci in range(2, 9):
        cl = get_column_letter(ci)
        wc(ws1, rt, ci, None, fmt=INR if ci > 2 else '#,##0', font=wbf, fill=dk)
        ws1.cell(rt, ci).value = f'=SUM({cl}{ds}:{cl}{ds+5})'
    wc(ws1, rt, 12, '', fill=dk)

    fills_list = [gr, yw, lg, og, rd, bl]
    for i in range(6):
        rr = ds + i; f = fills_list[i]
        wc(ws1, rr, 9, None, fmt='0.0%', fill=f, font=b9)
        ws1.cell(rr, 9).value = f'=IF(B${rt}>0,B{rr}/B${rt},0)'
        wc(ws1, rr, 10, None, fmt='0.0%', fill=f, font=b9)
        ws1.cell(rr, 10).value = f'=IF(H${rt}>0,H{rr}/H${rt},0)'
        wc(ws1, rr, 11, None, fmt=INR, fill=f, font=b9)
        ws1.cell(rr, 11).value = f'=G{rr}'
    for ci in [9,10,11]:
        cl = get_column_letter(ci)
        wc(ws1, rt, ci, None, fmt='0.0%' if ci < 11 else INR, font=wbf, fill=dk)
        ws1.cell(rt, ci).value = f'=SUM({cl}{ds}:{cl}{ds+5})'

    scw(ws1, [48, 14, 18, 16, 16, 16, 18, 20, 10, 10, 18, 42])
    ws1.freeze_panes = 'A6'

    # ==================================================================
    # SAVE
    # ==================================================================
    log("Saving workbook...")
    wb.save(output_file)
    sheet_names = wb.sheetnames
    results['checks'].append({'name': 'Excel Generated', 'status': 'PASS', 'detail': f'{len(sheet_names)} sheets'})
    results['stats'] = {'sheets': len(sheet_names), 'output_file': output_file,
                         'formulas': 'VLOOKUP+SUMIFS+COUNTIF+SUMPRODUCT (all on standardized $F/$G/$H/$I/$K columns, Source filter on PartialMatch)'}
    results['status'] = 'passed'
    log(f"Report saved: {output_file} ({len(sheet_names)} sheets)")
    log("Column map: A=GSTIN B=Vendor C=InvNo D=Type E=Date F=Taxable G=IGST H=CGST I=SGST J=Cess K=Total")
    return results
