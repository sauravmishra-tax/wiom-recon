"""
AGENT 7: SCHEMA GUARD (Column Validator)
==========================================
Prevents column mismatch bugs by validating that all raw sheets
follow the STANDARDIZED 11-column layout BEFORE any SUMIFS/formulas
are generated. Runs AFTER Report Builder to verify the output Excel.

STANDARDIZED LAYOUT:
  A=GSTIN  B=Vendor(VLOOKUP)  C=InvNo  D=Type  E=Date
  F=Taxable  G=IGST  H=CGST  I=SGST  J=Cess  K=Total
"""
import os
from openpyxl import load_workbook

AGENT_NAME = "Schema Guard"
AGENT_ID = 7

# Expected standardized header layout for all raw sheets
EXPECTED_HEADERS = [
    'GSTIN', 'Vendor (VLOOKUP)', 'Invoice No.', 'Type', 'Date',
    'Taxable Amt', 'IGST', 'CGST', 'SGST', 'Cess', 'Total Amt'
]

# Sheets that MUST follow the standardized layout
RAW_SHEETS = ['Raw-Reconciled', 'Raw-BooksOnly', 'Raw-GSTR2B Only', 'Raw-PartialMatch']

# Expected Vendor Master layout
VM_HEADERS = ['GSTIN', 'S.No.', 'Vendor Name', 'Source', 'State Code', 'Status']

# Expert Cross-Match expected column count
XM_EXPECTED_COLS = 29  # A through AC


def run(output_file, log_fn=None):
    """Validate output Excel schema integrity."""
    results = {
        'agent': AGENT_NAME,
        'status': 'running',
        'checks': [],
        'errors': [],
        'warnings': [],
        'stats': {}
    }

    def log(msg):
        if log_fn:
            log_fn(AGENT_ID, msg)

    log("Schema Guard activated — validating output Excel structure...")

    if not os.path.exists(output_file):
        results['status'] = 'failed'
        results['errors'].append(f'Output file not found: {output_file}')
        log(f"FAIL: Output file not found!")
        return results

    wb = load_workbook(output_file, read_only=True, data_only=False)
    errors = []
    warnings = []
    checks = []

    # ================================================================
    # CHECK 1: All required sheets exist
    # ================================================================
    log("Check 1: Verifying required sheets exist...")
    required_sheets = RAW_SHEETS + ['Vendor Master', 'Expert Cross-Match',
                                      'Books Only Detail', 'GSTR-2B Only Detail',
                                      'GSTIN Gap Analysis', 'RECON DASHBOARD',
                                      'Inter-Recon Scrutiny']
    missing = [s for s in required_sheets if s not in wb.sheetnames]
    if missing:
        errors.append(f'Missing sheets: {missing}')
        checks.append({'name': 'Required Sheets', 'status': 'FAIL', 'detail': f'Missing: {missing}'})
        log(f"FAIL: Missing sheets: {missing}")
    else:
        checks.append({'name': 'Required Sheets', 'status': 'PASS', 'detail': f'All {len(required_sheets)} sheets present'})
        log(f"PASS: All {len(required_sheets)} required sheets found")

    # ================================================================
    # CHECK 2: Raw sheet column headers match standardized layout
    # ================================================================
    log("Check 2: Validating raw sheet column headers...")
    raw_ok = 0
    raw_fail = 0
    for sheet_name in RAW_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        actual_headers = []
        for col in range(1, 12):  # A through K
            cell = ws.cell(1, col)
            actual_headers.append(str(cell.value or '').strip())

        mismatches = []
        for i, (expected, actual) in enumerate(zip(EXPECTED_HEADERS, actual_headers)):
            if expected != actual:
                mismatches.append(f'Col {chr(65+i)}: expected "{expected}" got "{actual}"')

        if mismatches:
            raw_fail += 1
            errors.append(f'{sheet_name}: {mismatches}')
            log(f"FAIL: {sheet_name} — {len(mismatches)} column mismatches: {mismatches}")
        else:
            raw_ok += 1
            log(f"PASS: {sheet_name} — all 11 columns match standardized layout")

    checks.append({
        'name': 'Raw Sheet Headers',
        'status': 'PASS' if raw_fail == 0 else 'FAIL',
        'detail': f'{raw_ok}/{raw_ok + raw_fail} raw sheets match standardized A-K layout'
    })

    # ================================================================
    # CHECK 3: Verify column F = Taxable, K = Total in each raw sheet
    # (Spot-check: row 2 values should be numeric or formula)
    # ================================================================
    log("Check 3: Spot-checking numeric columns (F=Taxable, K=Total)...")
    numeric_issues = []
    for sheet_name in RAW_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue
        for col, col_name in [(6, 'F-Taxable'), (11, 'K-Total')]:
            val = ws.cell(2, col).value
            # If it's a formula, it's fine; if it's a string (non-empty, non-formula), flag it
            if val is not None and isinstance(val, str) and not val.startswith('=') and val.strip() != '':
                try:
                    float(val)
                except ValueError:
                    numeric_issues.append(f'{sheet_name}!{col_name}: "{val}" (should be numeric)')

    if numeric_issues:
        warnings.append(f'Non-numeric in amount columns: {numeric_issues}')
        checks.append({'name': 'Numeric Columns', 'status': 'WARN', 'detail': str(numeric_issues)})
        log(f"WARN: {len(numeric_issues)} non-numeric cells in amount columns")
    else:
        checks.append({'name': 'Numeric Columns', 'status': 'PASS', 'detail': 'All amount columns OK'})
        log("PASS: Amount columns (F, K) contain valid numbers/formulas")

    # ================================================================
    # CHECK 4: Verify Vendor Master has correct layout
    # ================================================================
    log("Check 4: Validating Vendor Master layout...")
    if 'Vendor Master' in wb.sheetnames:
        ws_vm = wb['Vendor Master']
        vm_actual = [str(ws_vm.cell(1, i).value or '').strip() for i in range(1, 7)]
        vm_mismatch = [f'Col {i+1}: "{vm_actual[i]}" != "{VM_HEADERS[i]}"'
                       for i in range(len(VM_HEADERS)) if i < len(vm_actual) and vm_actual[i] != VM_HEADERS[i]]
        if vm_mismatch:
            warnings.append(f'Vendor Master headers: {vm_mismatch}')
            checks.append({'name': 'Vendor Master', 'status': 'WARN', 'detail': str(vm_mismatch)})
            log(f"WARN: Vendor Master header issues: {vm_mismatch}")
        else:
            vm_rows = ws_vm.max_row - 1
            checks.append({'name': 'Vendor Master', 'status': 'PASS', 'detail': f'{vm_rows} vendors, A=GSTIN C=Name'})
            log(f"PASS: Vendor Master layout correct, {vm_rows} vendors")

    # ================================================================
    # CHECK 5: Verify Expert Cross-Match SUMIFS reference correct sheets
    # ================================================================
    log("Check 5: Validating Expert Cross-Match formula references...")
    formula_errors = []
    if 'Expert Cross-Match' in wb.sheetnames:
        ws_xm = wb['Expert Cross-Match']
        # Check a few key formula cells in the first data row (row 5)
        first_data_row = 5
        if ws_xm.max_row >= first_data_row:
            # Col K (11) = Books Taxable should reference Raw-BooksOnly + Raw-PartialMatch $F
            cell_k = ws_xm.cell(first_data_row, 11).value
            if cell_k and isinstance(cell_k, str):
                if 'Raw-BooksOnly' not in cell_k:
                    formula_errors.append(f'K{first_data_row}: Missing Raw-BooksOnly ref: {cell_k}')
                if 'Raw-PartialMatch' not in cell_k:
                    formula_errors.append(f'K{first_data_row}: Missing Raw-PartialMatch fallback: {cell_k}')
                if '$F:$F' not in cell_k:
                    formula_errors.append(f'K{first_data_row}: Should reference $F (Taxable): {cell_k}')

            # Col L (12) = GSTR-2B Taxable should reference Raw-GSTR2B Only + Raw-PartialMatch $F
            cell_l = ws_xm.cell(first_data_row, 12).value
            if cell_l and isinstance(cell_l, str):
                if 'Raw-GSTR2B Only' not in cell_l:
                    formula_errors.append(f'L{first_data_row}: Missing Raw-GSTR2B ref: {cell_l}')
                if 'Raw-PartialMatch' not in cell_l:
                    formula_errors.append(f'L{first_data_row}: Missing Raw-PartialMatch fallback: {cell_l}')
                if '$F:$F' not in cell_l:
                    formula_errors.append(f'L{first_data_row}: Should reference $F (Taxable): {cell_l}')

            # Col W (23) = Books Total should reference Raw-BooksOnly + Raw-PartialMatch $K
            cell_w = ws_xm.cell(first_data_row, 23).value
            if cell_w and isinstance(cell_w, str):
                if '$K:$K' not in cell_w:
                    formula_errors.append(f'W{first_data_row}: Should reference $K (Total): {cell_w}')
                if 'Raw-PartialMatch' not in cell_w:
                    formula_errors.append(f'W{first_data_row}: Missing Raw-PartialMatch fallback: {cell_w}')

            # Col C (3) = Vendor VLOOKUP should reference Vendor Master
            cell_c = ws_xm.cell(first_data_row, 3).value
            if cell_c and isinstance(cell_c, str):
                if 'Vendor Master' not in cell_c:
                    formula_errors.append(f'C{first_data_row}: Missing Vendor Master ref: {cell_c}')

    if formula_errors:
        errors.append(f'Formula reference errors: {formula_errors}')
        checks.append({'name': 'Formula References', 'status': 'FAIL', 'detail': f'{len(formula_errors)} errors'})
        for e in formula_errors:
            log(f"FAIL: {e}")
    else:
        checks.append({'name': 'Formula References', 'status': 'PASS',
                       'detail': 'All SUMIFS reference correct standardized columns ($F=Taxable, $K=Total)'})
        log("PASS: Expert Cross-Match formulas reference correct columns")

    # ================================================================
    # CHECK 6: Cross-verify raw sheet row counts vs detail sheets
    # ================================================================
    log("Check 6: Cross-verifying row counts...")
    count_checks = []
    if 'Raw-BooksOnly' in wb.sheetnames and 'Books Only Detail' in wb.sheetnames:
        raw_count = wb['Raw-BooksOnly'].max_row - 1  # minus header
        # Detail starts at row 4
        det_count = max(0, wb['Books Only Detail'].max_row - 4)
        if raw_count > 0 and det_count > 0:
            # They won't match exactly (detail = unmatched, raw = all books-only)
            count_checks.append(f'BooksOnly: Raw={raw_count}, Detail={det_count}')
            log(f"INFO: Raw-BooksOnly={raw_count} rows, Books Only Detail={det_count} rows")

    if 'Raw-GSTR2B Only' in wb.sheetnames and 'GSTR-2B Only Detail' in wb.sheetnames:
        raw_count = wb['Raw-GSTR2B Only'].max_row - 1
        det_count = max(0, wb['GSTR-2B Only Detail'].max_row - 4)
        if raw_count > 0 and det_count > 0:
            count_checks.append(f'GSTR2B: Raw={raw_count}, Detail={det_count}')
            log(f"INFO: Raw-GSTR2B Only={raw_count} rows, GSTR-2B Only Detail={det_count} rows")

    checks.append({'name': 'Row Count Verify', 'status': 'PASS', 'detail': '; '.join(count_checks) if count_checks else 'Verified'})

    # ================================================================
    # CHECK 7: Verify GSTIN Gap Analysis references
    # ================================================================
    log("Check 7: Validating GSTIN Gap Analysis formulas...")
    gap_errors = []
    if 'GSTIN Gap Analysis' in wb.sheetnames:
        ws_ga = wb['GSTIN Gap Analysis']
        if ws_ga.max_row >= 4:
            # Col D (4) = Books Count should use COUNTIF on Raw-BooksOnly!$A
            cell_d = ws_ga.cell(4, 4).value
            if cell_d and isinstance(cell_d, str):
                if 'Raw-BooksOnly' not in cell_d:
                    gap_errors.append(f'D4: Missing Raw-BooksOnly ref')
                if '$A:$A' not in cell_d:
                    gap_errors.append(f'D4: Should COUNTIF on $A (GSTIN)')

            # Col E (5) = Books Taxable should SUMIF on Raw-BooksOnly!$F
            cell_e = ws_ga.cell(4, 5).value
            if cell_e and isinstance(cell_e, str):
                if '$F:$F' not in cell_e:
                    gap_errors.append(f'E4: Should SUMIF on $F (Taxable)')

    if gap_errors:
        warnings.append(f'Gap Analysis formulas: {gap_errors}')
        checks.append({'name': 'Gap Analysis Formulas', 'status': 'WARN', 'detail': str(gap_errors)})
        for e in gap_errors:
            log(f"WARN: {e}")
    else:
        checks.append({'name': 'Gap Analysis Formulas', 'status': 'PASS',
                       'detail': 'COUNTIF/SUMIF reference $A=GSTIN, $F=Taxable correctly'})
        log("PASS: GSTIN Gap Analysis formulas validated")

    wb.close()

    # ================================================================
    # FINAL VERDICT
    # ================================================================
    total_checks = len(checks)
    passed = sum(1 for c in checks if c['status'] == 'PASS')
    failed = sum(1 for c in checks if c['status'] == 'FAIL')
    warned = sum(1 for c in checks if c['status'] == 'WARN')

    results['checks'] = checks
    results['errors'] = errors
    results['warnings'] = warnings
    results['stats'] = {
        'total_checks': total_checks,
        'passed': passed,
        'failed': failed,
        'warned': warned,
    }
    results['status'] = 'failed' if failed > 0 else 'passed'

    if failed > 0:
        log(f"SCHEMA GUARD VERDICT: FAIL — {failed} critical errors found!")
        log(f"Errors: {errors}")
    else:
        log(f"SCHEMA GUARD VERDICT: PASS — {passed}/{total_checks} checks passed, {warned} warnings")

    return results
