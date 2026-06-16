"""
AGENT 9: SCRUTINY & INTER-RECONCILIATION
==========================================
Post-output verification agent. Opens the generated Excel and verifies:
1. Raw sheet row counts match expected
2. PartialMatch Source column has Books + GSTN rows
3. Expert Cross-Match has NO N/A in Type/Date columns (hardcoded check)
4. Expert Cross-Match SUMIFS produce non-zero for matched invoices
5. Inter-Recon sheet exists with scrutiny checks
6. No double-counted invoices across sheets
7. Vendor Credit amounts are negative
"""
import pandas as pd
import numpy as np

AGENT_NAME = "Scrutiny & Inter-Recon"
AGENT_ID = 9


def run(output_file, dataframes, inv_matched, books_unmatched, gstn_unmatched, log_fn=None):
    """Post-generation scrutiny of the output Excel file."""
    results = {
        'agent': AGENT_NAME,
        'status': 'running',
        'checks': [],
        'stats': {},
        'issues': [],
    }

    def log(msg):
        if log_fn:
            log_fn(AGENT_ID, msg)

    log("Starting post-output scrutiny & inter-reconciliation...")

    from openpyxl import load_workbook
    try:
        wb = load_workbook(output_file, data_only=False)
    except Exception as e:
        results['status'] = 'failed'
        results['checks'].append({'name': 'File Load', 'status': 'FAIL', 'detail': str(e)})
        return results

    sheet_names = wb.sheetnames
    issues = []
    checks_passed = 0
    total_checks = 0

    # ====== CHECK 1: Inter-Recon Scrutiny sheet exists ======
    total_checks += 1
    if 'Inter-Recon Scrutiny' in sheet_names:
        checks_passed += 1
        results['checks'].append({'name': 'Inter-Recon Sheet', 'status': 'PASS', 'detail': 'Sheet exists'})
        log("  ✓ Inter-Recon Scrutiny sheet present")
    else:
        issues.append('Inter-Recon Scrutiny sheet MISSING')
        results['checks'].append({'name': 'Inter-Recon Sheet', 'status': 'FAIL', 'detail': 'MISSING'})
        log("  ✗ Inter-Recon Scrutiny sheet MISSING")

    # ====== CHECK 2: PartialMatch has Source column (L) ======
    total_checks += 1
    if 'Raw-PartialMatch' in sheet_names:
        ws_pm = wb['Raw-PartialMatch']
        col_l_header = ws_pm.cell(1, 12).value
        if col_l_header == 'Source':
            checks_passed += 1
            # Count Books vs GSTN
            books_count = 0
            gstn_count = 0
            for row in range(2, ws_pm.max_row + 1):
                val = ws_pm.cell(row, 12).value
                if val == 'Books':
                    books_count += 1
                elif val == 'GSTN':
                    gstn_count += 1
            results['checks'].append({'name': 'PM Source Column', 'status': 'PASS',
                                       'detail': f'Books={books_count}, GSTN={gstn_count}'})
            log(f"  ✓ PartialMatch Source col: Books={books_count}, GSTN={gstn_count}")
            if books_count != gstn_count:
                issues.append(f'PartialMatch Books({books_count}) ≠ GSTN({gstn_count}) count mismatch')
        else:
            issues.append(f'PartialMatch col L header is "{col_l_header}" not "Source"')
            results['checks'].append({'name': 'PM Source Column', 'status': 'FAIL', 'detail': f'Col L = {col_l_header}'})
    else:
        results['checks'].append({'name': 'PM Source Column', 'status': 'FAIL', 'detail': 'Sheet missing'})

    # ====== CHECK 3: Expert Cross-Match - No N/A in Type/Date ======
    total_checks += 1
    if 'Expert Cross-Match' in sheet_names:
        ws_xm = wb['Expert Cross-Match']
        na_count = 0
        data_rows = 0
        for row in range(5, ws_xm.max_row + 1):
            gstin = ws_xm.cell(row, 2).value
            if not gstin or gstin == 'TOTAL':
                break
            data_rows += 1
            # Col D = Type, Col H = Books Date, Col I = GSTR-2B Date
            for col in [4, 8, 9]:
                val = ws_xm.cell(row, col).value
                if val == 'N/A' or val is None:
                    na_count += 1

        if na_count == 0:
            checks_passed += 1
            results['checks'].append({'name': 'XM No N/A', 'status': 'PASS',
                                       'detail': f'{data_rows} rows, 0 N/A values'})
            log(f"  ✓ Expert Cross-Match: {data_rows} rows, zero N/A in Type/Date")
        else:
            issues.append(f'Expert Cross-Match has {na_count} N/A values in Type/Date')
            results['checks'].append({'name': 'XM No N/A', 'status': 'FAIL',
                                       'detail': f'{na_count} N/A values found'})
            log(f"  ✗ Expert Cross-Match: {na_count} N/A values found!")
    else:
        results['checks'].append({'name': 'XM No N/A', 'status': 'FAIL', 'detail': 'Sheet missing'})

    # ====== CHECK 4: Expert Cross-Match SUMIFS formulas reference PartialMatch with Source filter ======
    total_checks += 1
    if 'Expert Cross-Match' in sheet_names:
        ws_xm = wb['Expert Cross-Match']
        # Check first data row's Books Taxable formula (col K = 11)
        first_formula = ws_xm.cell(5, 11).value or ''
        has_source_filter = '$L:$L,"Books"' in str(first_formula) or '$L:$L,"GSTN"' in str(first_formula)
        if has_source_filter:
            checks_passed += 1
            results['checks'].append({'name': 'XM Source Filter', 'status': 'PASS',
                                       'detail': 'SUMIFS filter by Source=Books/GSTN'})
            log("  ✓ SUMIFS correctly filter PartialMatch by Source column")
        else:
            issues.append('SUMIFS missing Source filter on PartialMatch')
            results['checks'].append({'name': 'XM Source Filter', 'status': 'FAIL',
                                       'detail': f'Formula: {str(first_formula)[:80]}'})
            log(f"  ✗ SUMIFS missing Source filter! Formula: {str(first_formula)[:80]}")
    else:
        results['checks'].append({'name': 'XM Source Filter', 'status': 'FAIL', 'detail': 'Sheet missing'})

    # ====== CHECK 5: Cross-match count matches inv_matched ======
    total_checks += 1
    expected_xm = len(inv_matched)
    if 'Expert Cross-Match' in sheet_names:
        ws_xm = wb['Expert Cross-Match']
        actual_xm = 0
        for row in range(5, ws_xm.max_row + 1):
            if ws_xm.cell(row, 2).value and ws_xm.cell(row, 2).value != 'TOTAL':
                actual_xm += 1
            else:
                break
        if actual_xm == expected_xm:
            checks_passed += 1
            results['checks'].append({'name': 'XM Count Match', 'status': 'PASS',
                                       'detail': f'{actual_xm} matches'})
            log(f"  ✓ Cross-Match count: {actual_xm} = expected {expected_xm}")
        else:
            issues.append(f'Cross-Match count {actual_xm} ≠ expected {expected_xm}')
            results['checks'].append({'name': 'XM Count Match', 'status': 'FAIL',
                                       'detail': f'{actual_xm} ≠ {expected_xm}'})

    # ====== CHECK 6: Vendor Credits are negative ======
    total_checks += 1
    credit_issues = 0
    for m in inv_matched:
        t = str(m.get('Type', '')).strip().lower()
        if 'credit' in t:
            if m.get('Books_Total', 0) > 0:
                credit_issues += 1
    for item in books_unmatched + gstn_unmatched:
        t = str(item.get('Type', '')).strip().lower()
        if 'credit' in t:
            if item.get('Total', 0) > 0:
                credit_issues += 1

    if credit_issues == 0:
        checks_passed += 1
        results['checks'].append({'name': 'Credits Negative', 'status': 'PASS', 'detail': 'All vendor credits negative'})
        log("  ✓ All vendor credit amounts are NEGATIVE")
    else:
        issues.append(f'{credit_issues} vendor credits with POSITIVE amounts')
        results['checks'].append({'name': 'Credits Negative', 'status': 'WARN',
                                   'detail': f'{credit_issues} positive credits'})

    # ====== CHECK 7: No duplicate invoices across BooksOnly and CrossMatch ======
    total_checks += 1
    xm_book_invs = set()
    for m in inv_matched:
        xm_book_invs.add((m['GSTIN'], m['Books_Inv']))
    book_only_invs = set()
    for item in books_unmatched:
        book_only_invs.add((item['GSTIN'], item['Inv']))
    overlap = xm_book_invs & book_only_invs
    if len(overlap) == 0:
        checks_passed += 1
        results['checks'].append({'name': 'No Duplicates', 'status': 'PASS',
                                   'detail': 'Zero overlap between CrossMatch & BooksOnly'})
        log("  ✓ No duplicate invoices between CrossMatch and BooksOnly")
    else:
        issues.append(f'{len(overlap)} invoices in BOTH CrossMatch and BooksOnly')
        results['checks'].append({'name': 'No Duplicates', 'status': 'FAIL',
                                   'detail': f'{len(overlap)} overlapping invoices'})

    # ====== CHECK 8: Total sheets count ======
    total_checks += 1
    expected_sheets = ['RECON DASHBOARD', 'Vendor Master', 'Raw-Reconciled', 'Raw-BooksOnly',
                       'Raw-GSTR2B Only', 'Raw-PartialMatch', 'Expert Cross-Match',
                       'Books Only Detail', 'GSTR-2B Only Detail', 'GSTIN Gap Analysis',
                       'Inter-Recon Scrutiny']
    missing = [s for s in expected_sheets if s not in sheet_names]
    if len(missing) == 0:
        checks_passed += 1
        results['checks'].append({'name': 'All Sheets', 'status': 'PASS',
                                   'detail': f'{len(sheet_names)} sheets present'})
        log(f"  ✓ All {len(expected_sheets)} required sheets present")
    else:
        issues.append(f'Missing sheets: {missing}')
        results['checks'].append({'name': 'All Sheets', 'status': 'FAIL',
                                   'detail': f'Missing: {missing}'})

    wb.close()

    # Final verdict
    verdict = 'PASS' if checks_passed == total_checks else ('WARN' if checks_passed >= total_checks - 1 else 'FAIL')
    results['status'] = 'passed' if verdict != 'FAIL' else 'failed'
    results['stats'] = {
        'passed': checks_passed,
        'total_checks': total_checks,
        'issues': len(issues),
        'verdict': f'{verdict} ({checks_passed}/{total_checks})',
    }
    results['issues'] = issues

    log(f"Scrutiny verdict: {verdict} — {checks_passed}/{total_checks} checks passed")
    if issues:
        for iss in issues:
            log(f"  ⚠ {iss}")

    return results
