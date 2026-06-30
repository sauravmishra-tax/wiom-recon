"""
WIOM Zoho Books vs GST Recon — Multi-user workflow platform
=============================================================
- 9 AI sub-agents produce the reconciliation (unchanged engine).
- Results are stored monthly in a database, auto-split state-wise.
- Accounts team logs in and writes remarks/reasons on their state's rows.
- Manager (admin) approves/resolves; every change is audit-logged (who + when).
- Local-first (SQLite) and cloud-portable (set DATABASE_URL to switch).
"""
import os
import sys
import json
import time
import threading
import base64
import hashlib
from datetime import datetime, timedelta
from cryptography.fernet import Fernet

from flask import (Flask, render_template, request, jsonify, send_file,
                   redirect, url_for, abort)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from sqlalchemy import func

def _get_fernet():
    key = (os.environ.get('SECRET_KEY') or 'wiom-recon-default-secret-key-change-me').encode()
    derived = base64.urlsafe_b64encode(hashlib.sha256(key).digest())
    return Fernet(derived)

def _encrypt(val):
    if not val: return val
    return _get_fernet().encrypt(val.encode()).decode()

def _decrypt(val):
    if not val: return val
    try: return _get_fernet().decrypt(val.encode()).decode()
    except Exception: return val  # fallback for pre-encryption plain values


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)

from config import Config
from models import (db, User, ReconRun, ReconRow, VendorMaster, AuditLog,
                    Setting, RowComment, RowAttachment, VendorEmail, LoginEvent,
                    log_audit, now_ist, get_setting, set_setting)
from auth import login_manager, auth_bp, admin_required, superadmin_required, write_required
from persist import persist_run, derive_period, derive_period_from_file
from state_codes import STATE_CODES, WIOM_STATES
import zoho
import slack_notify

from agents import agent_1_validator
from agents import agent_2_vendor_resolver
from agents import agent_3_cross_matcher
from agents import agent_4_itc_analyzer
from agents import agent_5_anomaly_detector
from agents import agent_6_report_builder
from agents import agent_7_schema_guard
from agents import agent_8_qa_reviewer
from agents import agent_9_scrutiny


def create_app():
    app = Flask(__name__, template_folder='templates', static_folder='static')
    app.config.from_object(Config)
    app.config['TEMPLATES_AUTO_RELOAD'] = True  # reflect template edits without restart

    db.init_app(app)
    login_manager.init_app(app)
    app.register_blueprint(auth_bp)

    with app.app_context():
        db.create_all()
        _run_migrations()
        _seed_admin(app)
        _auto_approve_fully_reconciled()

    _start_scheduler(app)
    return app


def _start_scheduler(app):
    """Monthly CFO-summary email (feature 7). Best-effort: only fires while the
    app is running. Sends on the configured day-of-month at 09:00 IST."""
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
    except ImportError:
        print("  [scheduler] apscheduler not installed — monthly CFO email disabled (manual send still works).")
        return

    def monthly_cfo():
        with app.app_context():
            try:
                day = int(get_setting('cfo_send_day', '1') or '1')
            except ValueError:
                day = 1
            if now_ist().day == day and _smtp_configured() and get_setting('cfo_email'):
                ok, msg = _send_cfo_email()
                print(f"  [scheduler] Monthly CFO email: {ok} — {msg}")

    def daily_slack():
        with app.app_context():
            day = get_setting('slack_send_day', '*') or '*'   # '*' = every day, or a day-of-month
            if not get_setting('slack_webhook'):
                return
            if day == '*' or str(now_ist().day) == str(day):
                ok, msg = _send_slack_report()
                print(f"  [scheduler] Daily Slack report: {ok} — {msg}")

    import pytz
    ist = pytz.timezone('Asia/Kolkata')
    sched = BackgroundScheduler(daemon=True, timezone=ist)
    def approval_reminder():
        with app.app_context():
            since = now_ist() - timedelta(days=3)
            stale = ReconRow.query.filter(ReconRow.status == 'remarked',
                                          ReconRow.remarked_at <= since).count()
            if stale == 0:
                return
            last = get_setting('reminder_last_sent', '')
            today = now_ist().strftime('%Y-%m-%d')
            if last == today:
                return
            set_setting('reminder_last_sent', today)
            db.session.commit()
            admins = User.query.filter(User.role.in_(['admin', 'superadmin']), User.active == True).all()
            for admin in admins:
                try:
                    from email_util import send_email
                    send_email(admin.email, f'[WIOM Recon] {stale} rows pending approval (3+ days)',
                               f'<p>{stale} remarked rows are waiting for approval for 3+ days.</p>'
                               f'<p><a href="https://web-production-bf681c.up.railway.app/detail">Review now →</a></p>')
                except Exception as e:
                    print(f'  [reminder] email failed: {e}')

    sched.add_job(monthly_cfo, 'cron', hour=9, minute=0, id='monthly_cfo')
    sched.add_job(daily_slack, 'cron', hour=9, minute=30, id='daily_slack')
    sched.add_job(approval_reminder, 'cron', hour=8, minute=30, id='approval_reminder')

    # Slack Bot daily jobs
    import slack_notify
    sched.add_job(slack_notify.notify_pending_summary, 'cron', hour=7,  minute=0, id='slack_pending')
    sched.add_job(slack_notify.notify_cfo_summary,     'cron', hour=19, minute=0, id='slack_cfo')

    sched.start()
    print("  [scheduler] Started — Slack pending @7:00 IST, Slack CFO summary @19:00 IST")
    return app


def _auto_approve_fully_reconciled():
    """On every startup: ensure all Fully Reconciled rows are status=approved.
    Idempotent — only touches rows still stuck on 'open'."""
    updated = ReconRow.query.filter(
        ReconRow.recon_status.like('%Fully Reconciled%'),
        ReconRow.status == 'open'
    ).update({'status': 'approved'}, synchronize_session=False)
    if updated:
        db.session.commit()
        print(f"  [startup] Auto-approved {updated} Fully Reconciled rows.")


def _run_migrations():
    """Add any missing columns to existing tables (safe to run on every start)."""
    cols_to_add = [
        ("recon_rows", "itc_table4", "VARCHAR(20) DEFAULT ''"),
        # FIX 2 — soft-delete audit logs on re-upload
        ("recon_runs", "archived", "BOOLEAN DEFAULT 0"),
        # FIX 3 — period lock/unlock
        ("recon_runs", "locked", "BOOLEAN DEFAULT 0"),
        ("recon_runs", "locked_by_id", "INTEGER"),
        ("recon_runs", "locked_at", "DATETIME"),
    ]
    for table, col, col_def in cols_to_add:
        try:
            db.session.execute(db.text(f"ALTER TABLE {table} ADD COLUMN {col} {col_def}"))
            db.session.commit()
            print(f"  [migration] Added column {table}.{col}")
        except Exception:
            db.session.rollback()  # column already exists — ignore

    # FIX 6 — seed default email settings if not already set
    for k, v in [
        ('email_cc', 'mahesh.thakur@wiom.in, wiomfinance@wiom.in, tushar.gupta@wiom.in, ap@wiom.in'),
        ('email_sign', 'Regards,\nFinance and Taxation Team\nOmnia Information Private Limited'),
    ]:
        if not get_setting(k):
            set_setting(k, v)
            db.session.commit()

    # One-time: backfill existing RowComments into AuditLog
    if not get_setting('comments_backfilled_v1'):
        already = {(e.row_id, e.new_value) for e in
                   AuditLog.query.filter_by(action='comment').all()}
        backfilled = 0
        for c in RowComment.query.all():
            if (c.row_id, c.text) not in already:
                db.session.add(AuditLog(
                    row_id=c.row_id, user_id=c.user_id, user_name=c.user_name,
                    action='comment', field='discussion',
                    old_value='', new_value=c.text,
                    created_at=c.created_at,
                ))
                backfilled += 1
        db.session.commit()
        set_setting('comments_backfilled_v1', '1')
        db.session.commit()
        if backfilled:
            print(f'  [migration] Backfilled {backfilled} discussion comments into audit log')

    # One-time: reset all users to default password Wiom@123 if not already done
    if not get_setting('default_pwd_reset_done'):
        for u in User.query.all():
            u.set_password('Wiom@123')
        db.session.commit()
        set_setting('default_pwd_reset_done', '1')
        db.session.commit()
        print('  [migration] All user passwords reset to Wiom@123')


def _seed_admin(app):
    """Create the first manager account if no users exist."""
    if User.query.count() == 0:
        admin = User(
            name=app.config['ADMIN_NAME'],
            email=app.config['ADMIN_EMAIL'].lower(),
            role='superadmin', title='Platform Owner',
        )
        admin.set_password(app.config['ADMIN_PASSWORD'])
        db.session.add(admin)
        db.session.commit()
        print(f"  [seed] Admin created: {admin.email} / {app.config['ADMIN_PASSWORD']}")


app = create_app()

# ----------------------------------------------------------------------
# LIVE PROCESSING STATE (one upload at a time)
# ----------------------------------------------------------------------
processing_state = {
    'active': False, 'progress': 0, 'current_agent': '',
    'logs': [], 'agent_results': {}, 'run_id': None,
    'output_file': None, 'error': None, 'start_time': None,
}


def reset_state():
    processing_state.update({
        'active': False, 'progress': 0, 'current_agent': '',
        'logs': [], 'agent_results': {}, 'run_id': None,
        'output_file': None, 'error': None, 'start_time': None,
    })


def add_log(agent_id, msg):
    names = {0: 'SYSTEM', 1: 'Validator', 2: 'Vendor Resolver', 3: 'Cross-Matcher',
             4: 'ITC Analyzer', 5: 'Anomaly Detector', 6: 'Report Builder',
             7: 'Schema Guard', 8: 'QA Reviewer', 9: 'Scrutiny'}
    processing_state['logs'].append({
        'time': datetime.now().strftime('%H:%M:%S'),
        'agent': names.get(agent_id, f'Agent {agent_id}'),
        'agent_id': agent_id, 'msg': msg,
    })


# ----------------------------------------------------------------------
# ENGINE ORCHESTRATOR (9 agents) — runs in a background thread
# ----------------------------------------------------------------------
def run_reconciliation(file_path, period, label, user_id, run_state=None):
    with app.app_context():
        try:
            processing_state['active'] = True
            processing_state['start_time'] = time.time()
            processing_state['state'] = run_state or ''
            processing_state['period'] = period or ''
            user_obj = db.session.get(User, user_id)
            processing_state['uploaded_by'] = user_obj.name if user_obj else 'System'

            processing_state['current_agent'] = 'Agent 1: Data Validator'
            processing_state['progress'] = 5
            add_log(0, '═══ AGENT 1: DATA VALIDATOR ═══')
            r1 = agent_1_validator.run(file_path, log_fn=add_log)
            processing_state['agent_results']['validator'] = {
                'status': r1['status'], 'checks': r1['checks'],
                'warnings': r1['warnings'], 'errors': r1['errors'], 'stats': r1['stats']}
            if r1['status'] == 'failed':
                processing_state['error'] = f"Validation failed: {r1['errors']}"
                processing_state['active'] = False
                return
            processing_state['progress'] = 15

            processing_state['current_agent'] = 'Agent 2: Vendor Resolver'
            processing_state['progress'] = 20
            add_log(0, '═══ AGENT 2: VENDOR RESOLVER ═══')
            r2 = agent_2_vendor_resolver.run(r1['data'], log_fn=add_log)
            processing_state['agent_results']['vendor_resolver'] = {
                'status': r2['status'], 'checks': r2['checks'], 'stats': r2['stats']}
            processing_state['progress'] = 35

            processing_state['current_agent'] = 'Agent 3: Cross-Match Engine'
            processing_state['progress'] = 40
            add_log(0, '═══ AGENT 3: CROSS-MATCH ENGINE ═══')
            r3 = agent_3_cross_matcher.run(r2['data'], r2['vendor_map'], log_fn=add_log)
            processing_state['agent_results']['cross_matcher'] = {
                'status': r3['status'], 'checks': r3['checks'], 'stats': r3['stats']}
            processing_state['progress'] = 55

            # Period comes from the user's month pick at upload. Only auto-detect
            # (from invoice dates) as a fallback if none was provided.
            if not period:
                period = derive_period(r3['inv_matched'], r3['books_unmatched'],
                                       r3['gstn_unmatched'])
                add_log(0, f'Auto-detected period from invoice dates: {period}')
            else:
                add_log(0, f'Reconciliation period (selected): {period}')

            processing_state['current_agent'] = 'Agent 4: ITC Risk Analyzer'
            processing_state['progress'] = 60
            add_log(0, '═══ AGENT 4: ITC RISK ANALYZER ═══')
            r4 = agent_4_itc_analyzer.run(r2['data'], r3['inv_matched'],
                                          r3['books_unmatched'], r3['gstn_unmatched'], log_fn=add_log)
            processing_state['agent_results']['itc_analyzer'] = {
                'status': r4['status'], 'checks': r4['checks'], 'stats': r4['stats'],
                'itc_summary': r4['itc_summary']}
            processing_state['progress'] = 70

            processing_state['current_agent'] = 'Agent 5: Anomaly Detector'
            processing_state['progress'] = 75
            add_log(0, '═══ AGENT 5: ANOMALY DETECTOR ═══')
            r5 = agent_5_anomaly_detector.run(r2['data'], r3['inv_matched'],
                                              r3['books_unmatched'], r3['gstn_unmatched'], log_fn=add_log)
            processing_state['agent_results']['anomaly_detector'] = {
                'status': r5['status'], 'checks': r5['checks'], 'stats': r5['stats'],
                'anomalies': [{'type': a['type'], 'severity': a['severity'], 'detail': a['detail']}
                              for a in r5['anomalies']]}
            processing_state['progress'] = 85

            processing_state['current_agent'] = 'Agent 6: Report Builder'
            processing_state['progress'] = 88
            add_log(0, '═══ AGENT 6: REPORT BUILDER ═══')
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = os.path.join(app.config['OUTPUT_FOLDER'],
                                       f'WIOM_GST_Recon_{period}_{ts}.xlsx')
            r6 = agent_6_report_builder.run(
                input_file=file_path, output_file=output_file,
                dataframes=r2['data'], vendor_map=r2['vendor_map'],
                gst_cache=r2.get('gst_cache', {}), all_gstins=r2.get('all_gstins', []),
                inv_matched=r3['inv_matched'], books_unmatched=r3['books_unmatched'],
                gstn_unmatched=r3['gstn_unmatched'],
                books_gstins=r3.get('books_gstins', set()),
                gstn_gstins=r3.get('gstn_gstins', set()),
                anomalies=r5.get('anomalies', []), patterns=r5.get('patterns', []),
                itc_summary=r4.get('itc_summary', {}), log_fn=add_log)
            processing_state['agent_results']['report_builder'] = {
                'status': r6['status'], 'checks': r6['checks'], 'stats': r6['stats']}
            processing_state['progress'] = 90

            processing_state['current_agent'] = 'Agent 7: Schema Guard'
            add_log(0, '═══ AGENT 7: SCHEMA GUARD ═══')
            r7 = agent_7_schema_guard.run(output_file, log_fn=add_log)
            processing_state['agent_results']['schema_guard'] = {
                'status': r7['status'], 'checks': r7['checks'], 'stats': r7['stats']}
            processing_state['progress'] = 93

            processing_state['current_agent'] = 'Agent 8: QA Reviewer'
            add_log(0, '═══ AGENT 8: QA REVIEWER ═══')
            r8 = agent_8_qa_reviewer.run(
                agent_results=processing_state['agent_results'], dataframes=r2['data'],
                inv_matched=r3['inv_matched'], books_unmatched=r3['books_unmatched'],
                gstn_unmatched=r3['gstn_unmatched'], vendor_map=r2['vendor_map'],
                itc_summary=r4.get('itc_summary', {}), log_fn=add_log)
            processing_state['agent_results']['qa_reviewer'] = {
                'status': r8['status'], 'checks': r8['checks'], 'stats': r8['stats']}
            processing_state['progress'] = 96

            processing_state['current_agent'] = 'Agent 9: Scrutiny'
            add_log(0, '═══ AGENT 9: SCRUTINY & INTER-RECON ═══')
            r9 = agent_9_scrutiny.run(
                output_file=output_file, dataframes=r2['data'],
                inv_matched=r3['inv_matched'], books_unmatched=r3['books_unmatched'],
                gstn_unmatched=r3['gstn_unmatched'], log_fn=add_log)
            processing_state['agent_results']['scrutiny'] = {
                'status': r9['status'], 'checks': r9['checks'], 'stats': r9['stats']}

            # ---- Persist into DB (with rollback on any failure) ----
            add_log(0, 'Storing results into database for team review...')
            run = ReconRun(
                period=period, state=run_state or '', label=label,
                source_filename=os.path.basename(file_path),
                output_file=output_file, uploaded_by_id=user_id,
                summary_json=json.dumps({
                    'matched': len(r3['inv_matched']),
                    'books_only': len(r3['books_unmatched']),
                    'gstn_only': len(r3['gstn_unmatched']),
                    'itc_summary': r4.get('itc_summary', {}),
                    'qa_verdict': r8['stats'].get('verdict', 'N/A'),
                }))
            db.session.add(run)
            db.session.commit()
            try:
                reconciled_dfs = [r2['data'].get('Reconciled'), r2['data'].get('Matched')]
                reconciled_dfs = [d for d in reconciled_dfs if d is not None]
                n, carried = persist_run(run, r3['inv_matched'], r3['books_unmatched'],
                                r3['gstn_unmatched'], vendor_map=r2['vendor_map'],
                                gst_cache=r2.get('gst_cache', {}), run_state=run_state,
                                reconciled_dfs=reconciled_dfs)
                log_audit(None, current_user, 'upload',
                          'file_upload', '',
                          f'State: {run_state} | Period: {run.period} | Rows: {n} | Run ID: {run.id}')
                db.session.commit()
            except Exception as persist_err:
                db.session.rollback()
                ReconRun.query.filter_by(id=run.id).delete()
                db.session.commit()
                raise persist_err
            if carried:
                add_log(0, f'Smart re-upload: replaced previous {run_state} snapshot, '
                           f'carried over remarks/status on {carried} matching invoices')

            # Auto-sync vendor master from Zoho if configured (best-effort)
            if _zoho_configured():
                try:
                    res = sync_zoho_master()
                    if res.get('ok'):
                        add_log(0, f"Zoho auto-sync: {res['count']} vendors, {res['updated_rows']} names updated")
                except Exception as e:
                    add_log(0, f"Zoho auto-sync skipped: {e}")

            processing_state['run_id'] = run.id
            processing_state['output_file'] = output_file
            processing_state['progress'] = 100
            elapsed = time.time() - processing_state['start_time']
            add_log(0, f'═══ COMPLETE in {elapsed:.1f}s | {n} rows stored | '
                       f"QA: {r8['stats'].get('verdict','N/A')} ═══")

        except Exception as e:
            processing_state['error'] = str(e)
            add_log(0, f'ERROR: {e}')
            import traceback
            add_log(0, traceback.format_exc())
        finally:
            processing_state['active'] = False


# ======================================================================
# PAGE ROUTES
@app.route('/health')
def health():
    return 'ok', 200


# ======================================================================
@app.route('/')
@login_required
def dashboard():
    periods = [p[0] for p in db.session.query(ReconRow.period)
               .distinct().order_by(ReconRow.period.desc()).all()]
    runs = ReconRun.query.filter(ReconRun.archived != True).order_by(ReconRun.created_at.desc()).limit(12).all()
    stats = _summary_stats(current_user)
    return render_template('dashboard.html', periods=periods, runs=runs,
                           stats=stats, states=WIOM_STATES)


def _distinct_periods():
    return [p[0] for p in db.session.query(ReconRow.period)
            .distinct().order_by(ReconRow.period.desc()).all()]


def _fy_list(periods):
    """Distinct Indian financial years (start year): from data + current + next 10 years."""
    fys = set()
    for p in periods:
        try:
            y, m = int(p[:4]), int(p[5:7])
            fys.add(y if m >= 4 else y - 1)
        except (ValueError, IndexError):
            continue
    now = now_ist()
    current_fy = now.year if now.month >= 4 else now.year - 1
    for i in range(11):          # current FY + 10 future FYs
        fys.add(current_fy + i)
    return sorted(fys, reverse=True)


@app.route('/workflow')
@login_required
def workflow_page():
    return render_template('workflow.html')


@app.route('/detail')
@login_required
def detail():
    periods = _distinct_periods()
    fys = _fy_list(periods)
    now = now_ist()
    current_fy = now.year if now.month >= 4 else now.year - 1
    return render_template('detail.html', periods=periods, fys=fys,
                           default_fy=current_fy, states=WIOM_STATES,
                           is_admin=current_user.is_admin,
                           is_viewer=current_user.is_viewer)


@app.route('/upload-page')
@admin_required
def upload_page():
    zoho_ready = bool(get_setting('zoho_client_id') and get_setting('zoho_refresh_token')
                      and get_setting('zoho_org_id'))
    return render_template('upload.html', zoho_ready=zoho_ready)


@app.route('/upload/from-zoho', methods=['POST'])
@admin_required
def upload_from_zoho():
    """Fetch GSTR-2B reconciliation data directly from Zoho Books and persist it.
    No Excel needed — replaces the manual export+upload flow."""
    import zoho as zoho_mod
    from persist import persist_run
    from state_codes import state_from_gstin, code_for_state

    data = request.get_json(force=True)
    from_period = (data.get('from_period') or '').strip()   # YYYY-MM
    to_period   = (data.get('to_period')   or '').strip()   # YYYY-MM
    states      = data.get('states') or []
    label       = (data.get('label') or '').strip()

    if not from_period or not to_period:
        return jsonify({'ok': False, 'error': 'from_period and to_period required (YYYY-MM)'}), 400
    if from_period > to_period:
        return jsonify({'ok': False, 'error': 'from_period cannot be after to_period'}), 400

    def _months_between(f, t):
        fy, fm = int(f[:4]), int(f[5:7])
        ty, tm = int(t[:4]), int(t[5:7])
        months = []
        while (fy, fm) <= (ty, tm):
            months.append(f'{fy}-{fm:02d}')
            fm += 1
            if fm > 12:
                fm = 1; fy += 1
        return months

    periods_list = _months_between(from_period, to_period)

    cfg = {k: get_setting(k) for k in ('zoho_client_id', 'zoho_client_secret',
                                         'zoho_refresh_token', 'zoho_org_id', 'zoho_region')}
    if not cfg.get('zoho_client_id') or not cfg.get('zoho_refresh_token'):
        return jsonify({'ok': False, 'error': 'Zoho not configured in Settings.'}), 400

    all_results = []
    for period in periods_list:
        try:
            tok = zoho_mod.get_access_token(cfg['zoho_client_id'], cfg['zoho_client_secret'],
                                             cfg['zoho_refresh_token'], cfg.get('zoho_region', 'in'))
            gstr2b = zoho_mod.fetch_gstr2b(tok, cfg['zoho_org_id'],
                                            cfg.get('zoho_region', 'in'), period)
        except Exception as e:
            return jsonify({'ok': False, 'error': f'Zoho fetch failed for {period}: {e}'}), 500

        all_entries = (
            [(e, 'matched')    for e in gstr2b['matched']]   +
            [(e, 'books_only') for e in gstr2b['books_only']] +
            [(e, 'gstn_only')  for e in gstr2b['gstn_only']]  +
            [(e, 'reconciled') for e in gstr2b['reconciled']]
        )
        if not all_entries:
            continue

        state_map = {}
        for e, cat in all_entries:
            gstin = e.get('GSTIN', '')
            _, sname = state_from_gstin(gstin)
            if not sname:
                sname = 'Unknown'
            if states and sname not in states:
                continue
            if sname not in state_map:
                state_map[sname] = {'matched': [], 'books_only': [], 'gstn_only': [],
                                     'reconciled': [], 'vendor_map': {}}
            state_map[sname][cat].append(e)
            if e.get('GSTIN') and e.get('Vendor'):
                state_map[sname]['vendor_map'][e['GSTIN']] = e['Vendor']

        for sname, rows in state_map.items():
            run = ReconRun(
                period=period, label=label or f'Zoho import {period}',
                state=sname, uploaded_by_id=current_user.id)
            db.session.add(run); db.session.flush()

            inv_matched     = rows['matched'] + rows['reconciled']
            books_unmatched = rows['books_only']
            gstn_unmatched  = rows['gstn_only']

            n, carried = persist_run(run, inv_matched, books_unmatched, gstn_unmatched,
                                     vendor_map=rows['vendor_map'], run_state=sname)
            approved = ReconRow.query.filter(
                ReconRow.run_id == run.id,
                ReconRow.recon_status.like('%Fully Reconciled%'),
                ReconRow.status == 'open'
            ).update({'status': 'approved'}, synchronize_session=False)
            db.session.commit()

            all_results.append({'state': sname, 'period': period, 'run_id': run.id,
                                 'rows': n, 'carried': carried, 'auto_approved': approved,
                                 'matched': len(rows['matched']),
                                 'reconciled': len(rows['reconciled']),
                                 'books_only': len(books_unmatched),
                                 'gstn_only': len(gstn_unmatched)})

    if not all_results:
        return jsonify({'ok': False, 'error': 'No data found for selected period range and states.'}), 400

    return jsonify({'ok': True, 'from_period': from_period, 'to_period': to_period, 'states': all_results})


# ======================================================================
# ZOHO BROWSER AGENT  (server-side headless Playwright)
# ======================================================================
@app.route('/api/zoho/browser-fetch', methods=['POST'])
@login_required
def zoho_browser_fetch():
    """Headless browser automation: login to Zoho Books, export GSTR-2B recon
    Excel for the given period, then persist it as ReconRuns — one per state.
    Runs entirely on the server; no local browser needed."""
    data = request.get_json(force=True)
    from_period = (data.get('from_period') or '').strip()   # YYYY-MM
    to_period   = (data.get('to_period')   or '').strip()   # YYYY-MM
    states  = data.get('states') or []
    label   = (data.get('label') or '').strip()

    if not from_period or not to_period:
        return jsonify({'ok': False, 'error': 'from_period and to_period required (YYYY-MM)'}), 400

    zoho_email    = get_setting('zoho_browser_email') or os.environ.get('ZOHO_EMAIL', '')
    zoho_password = get_setting('zoho_browser_password') or os.environ.get('ZOHO_PASSWORD', '')
    zoho_org_id   = get_setting('zoho_org_id') or os.environ.get('ZOHO_ORG_ID', '')

    if not zoho_email or not zoho_password:
        return jsonify({'ok': False,
                        'error': 'Zoho login credentials not set. Add ZOHO_EMAIL and ZOHO_PASSWORD in Railway env vars or Settings.'}), 400

    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    except ImportError:
        return jsonify({'ok': False, 'error': 'Playwright not installed on server.'}), 500

    import tempfile, threading
    from persist import persist_run
    from state_codes import state_from_gstin

    fy, fm = from_period.split('-')
    ty, tm = to_period.split('-')
    zoho_from = f'{fm}-{fy}'
    zoho_to   = f'{tm}-{ty}'

    download_dir = tempfile.mkdtemp(prefix='wiom_zoho_')
    excel_path   = None

    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(
                headless=True,
                args=['--no-sandbox', '--disable-dev-shm-usage'],
            )
            ctx = browser.new_context(accept_downloads=True)
            page = ctx.new_page()

            # Intercept all network requests to Zoho APIs for diagnostics
            _api_calls = []
            def _on_request(req):
                u = req.url
                if ('zoho' in u or 'zohoapis' in u) and req.resource_type in ('xhr','fetch','document'):
                    _api_calls.append(f'{req.method} {u[:120]}')
            page.on('request', _on_request)

            # --- Login ---
            print(f'[zoho] goto login page, org={zoho_org_id}')
            page.goto(f'https://books.zoho.in/app/{zoho_org_id}', timeout=60000)
            page.wait_for_timeout(3000)
            print(f'[zoho] after goto, url={page.url[:80]}')

            if 'accounts.zoho' in page.url or 'login' in page.url.lower():
                print('[zoho] login page detected, filling credentials')
                # Log page title / visible error text for diagnosis
                try:
                    page_diag = page.evaluate("""() => ({
                        title: document.title,
                        h1: (document.querySelector('h1,h2,.signIn-form-title')||{}).innerText||'',
                        err: (document.querySelector('.error-msg,.disp-err,.alert,.za-error-msg')||{}).innerText||''
                    })""")
                    print(f'[zoho] login page diag: {page_diag}')
                except Exception:
                    pass
                page.fill('#login_id', zoho_email, timeout=10000)
                page.click('#nextbtn', timeout=5000)
                page.wait_for_timeout(2000)
                # Log what appeared after email submit
                try:
                    after_email_diag = page.evaluate("""() => ({
                        url: location.href.slice(0,80),
                        title: document.title,
                        err: (document.querySelector('.error-msg,.disp-err,.alert,.za-error-msg')||{}).innerText||'',
                        has_pwd: !!document.querySelector('#password')
                    })""")
                    print(f'[zoho] after email submit: {after_email_diag}')
                except Exception:
                    pass
                page.fill('#password', zoho_password, timeout=8000)
                page.click('#nextbtn', timeout=5000)
                page.wait_for_timeout(3000)
                # Log what appeared after password submit
                try:
                    after_pwd_diag = page.evaluate("""() => ({
                        url: location.href.slice(0,80),
                        title: document.title,
                        err: (document.querySelector('.error-msg,.disp-err,.alert,.za-error-msg,.error')||{}).innerText||'',
                        body_text: document.body.innerText.slice(0,300)
                    })""")
                    print(f'[zoho] after pwd submit: {after_pwd_diag}')
                except Exception:
                    pass
                print('[zoho] waiting for post-login redirect...')
                # Wait up to 60s for any Zoho page (org ID or intermediate auth pages)
                try:
                    page.wait_for_url(f'**{zoho_org_id}**', timeout=30000)
                except PWTimeout:
                    # Might be on trust-device or other intermediate page
                    print(f'[zoho] wait_for_url timeout, current url={page.url[:80]}')
                print(f'[zoho] post-login url={page.url[:80]}')

            # Handle intermediate Zoho auth pages: trust-device, announcements, etc.
            for _attempt in range(6):
                cur = page.url
                print(f'[zoho] inter-page check attempt={_attempt} url={cur[:80]}')
                if 'books.zoho.in/app' in cur and zoho_org_id in cur:
                    print('[zoho] reached app, done with auth pages')
                    break
                # "Trust this device?" — click "Not now"
                try:
                    page.click('.notnowbtn, button:has-text("Not now")', timeout=2000)
                    print('[zoho] clicked Not now (trust device)')
                    page.wait_for_timeout(2000)
                    continue
                except PWTimeout:
                    pass
                # "Trust" button — click it to proceed
                try:
                    page.click('.trustbtn, button:has-text("Trust")', timeout=1500)
                    print('[zoho] clicked Trust')
                    page.wait_for_timeout(2000)
                    continue
                except PWTimeout:
                    pass
                # Announcement/notice page — navigate directly
                if 'announcement' in cur or ('accounts.zoho' in cur and zoho_org_id not in cur):
                    print(f'[zoho] announcement/redirect page, navigating directly')
                    page.wait_for_timeout(1000)
                    break
                # Any other accounts.zoho page — wait and retry
                if 'accounts.zoho' in cur:
                    page.wait_for_timeout(3000)
                    continue
                break

            # --- Navigate to GSTR-2B Reconciliation and Export (one month at a time) ---
            # Zoho Books reconciliation works per-month; we iterate each month in the range.
            import datetime as _dt
            def _month_range(fp, tp):
                y0, m0 = int(fp[:4]), int(fp[5:])
                y1, m1 = int(tp[:4]), int(tp[5:])
                months = []
                y, m = y0, m0
                while (y, m) <= (y1, m1):
                    months.append(f'{y:04d}-{m:02d}')
                    m += 1
                    if m > 12:
                        m = 1; y += 1
                return months

            months = _month_range(from_period, to_period)
            print(f'[zoho] will export months: {months}')

            def _click_export(pg):
                """Click Export button then Excel option; returns True if download triggered."""
                # Find export button (case-insensitive, checks text/title/aria)
                found = pg.evaluate("""() => {
                    const all = document.querySelectorAll('button,a,[role="button"]');
                    for (const el of all) {
                        const t = ((el.innerText||el.textContent||'').trim() + ' ' +
                                   (el.title||'') + ' ' +
                                   (el.getAttribute('aria-label')||'')).toLowerCase();
                        if (/\\bexport\\b/.test(t) || /\\bdownload\\b/.test(t)) {
                            el.click();
                            return (el.innerText||el.title||el.getAttribute('aria-label')||el.className||'found').slice(0,60);
                        }
                    }
                    return null;
                }""")
                if found:
                    print(f'[zoho] export btn clicked: {found}')
                    pg.wait_for_timeout(1500)
                    # Log all visible dropdown items for diagnosis
                    try:
                        dropdown_items = pg.evaluate("""() => {
                            const items = document.querySelectorAll(
                                '.dropdown-menu a, .dropdown-menu li, .dropdown-menu .dropdown-item, '
                                + '.dropdown-menu button, [role="menuitem"], [role="option"], '
                                + '.ember-power-select-option, .dropdown-menu .disabled'
                            );
                            return Array.from(items).map(el => ({
                                t:(el.innerText||el.textContent||'').trim().slice(0,40),
                                cls:(el.className||'').slice(0,40),
                                tag: el.tagName
                            }));
                        }""")
                        print(f'[zoho] dropdown_items: {dropdown_items}')
                    except Exception:
                        pass
                    # Click Excel option in dropdown (try multiple patterns)
                    excel = pg.evaluate("""() => {
                        const selectors = [
                            '.dropdown-menu a',
                            '.dropdown-menu li',
                            '.dropdown-menu .dropdown-item',
                            '.dropdown-menu button',
                            '[role="menuitem"]',
                            '[role="option"]',
                            'li,a,button'
                        ];
                        for (const sel of selectors) {
                            for (const el of document.querySelectorAll(sel)) {
                                const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                                if (t && (t === 'excel' || /\\bexcel\\b/.test(t) || t.includes('.xlsx') || t === 'xlsx')) {
                                    el.click();
                                    return (el.innerText||'excel').slice(0,30);
                                }
                            }
                        }
                        // If only one dropdown item, click it regardless
                        const ddItems = document.querySelectorAll('.dropdown-menu a, .dropdown-menu .dropdown-item');
                        if (ddItems.length === 1) {
                            ddItems[0].click();
                            return 'only-item:' + (ddItems[0].innerText||'').trim().slice(0,20);
                        }
                        return null;
                    }""")
                    print(f'[zoho] excel option: {excel}')
                    if not excel:
                        # Last resort: try Playwright locators for visible dropdown items
                        for xsel in ['text=Excel', 'text=xlsx', 'text=XLSX',
                                     '.dropdown-menu a >> nth=0',
                                     '.dropdown-menu .dropdown-item >> nth=0']:
                            try:
                                pg.click(xsel, timeout=1000)
                                print(f'[zoho] fallback excel click: {xsel}')
                                break
                            except PWTimeout:
                                continue
                    return True
                # CSS fallbacks
                for sel in ['button:has-text("Export")', '[title="Export"]', '[aria-label="Export"]',
                            '[title*="Export"]', '.export-btn']:
                    try:
                        pg.click(sel, timeout=1500)
                        print(f'[zoho] CSS export: {sel}')
                        pg.wait_for_timeout(1000)
                        for xsel in ['text=Excel', 'li:has-text("Excel")', '[data-value="xlsx"]']:
                            try:
                                pg.click(xsel, timeout=1500)
                                print(f'[zoho] CSS excel: {xsel}')
                                return True
                            except PWTimeout:
                                continue
                        return True
                    except PWTimeout:
                        continue
                return False

            # --- Navigate via SPA router (not direct page.goto with hash) ---
            # First land on the GST filings page via proper navigation so SPA sets up state
            gst_filings_url = f'https://books.zoho.in/app/{zoho_org_id}#/gstfiling/tax/filings'
            print(f'[zoho] navigating to GST filings: {gst_filings_url[:80]}')
            page.goto(gst_filings_url, timeout=60000)
            page.wait_for_timeout(6000)
            print(f'[zoho] GST filings url: {page.url[:80]}')

            # Log all links/buttons on GST filings page to find reconciliation entry
            try:
                filing_links = page.evaluate("""() => {
                    const els = document.querySelectorAll('a,button,[role="button"]');
                    return Array.from(els).slice(0,60).map(e => ({
                        tag: e.tagName,
                        text: (e.innerText||e.textContent||'').trim().slice(0,60),
                        href: e.href||'',
                        cls: (e.className||'').slice(0,50)
                    })).filter(e => e.text || e.href);
                }""")
                for fl in (filing_links or []):
                    if 'gst' in fl.get('href','').lower() or 'reconcil' in fl.get('href','').lower() or 'reconcil' in fl.get('text','').lower() or 'gstr' in fl.get('text','').lower():
                        print(f'[zoho] FILING_LINK: tag={fl["tag"]} text="{fl["text"]}" href="{fl["href"][:80]}"')
            except Exception as fle:
                print(f'[zoho] filing links eval failed: {fle}')

            # Log FULL content area of GST filings page
            try:
                content_html = page.evaluate("""() => {
                    // Find the main content area (not sidebar)
                    const mainArea = document.querySelector('.main-area, .content-area, [class*="main-content"], .ember-application > .ember-view > div > div:last-child');
                    if (mainArea) return mainArea.innerHTML.slice(0, 3000);
                    // Fallback: find elements with gst/reconciliation text
                    const allText = document.body.innerHTML;
                    const idx = allText.toLowerCase().indexOf('reconcil');
                    if (idx >= 0) return allText.slice(Math.max(0,idx-100), idx+500);
                    return 'No reconciliation text found in page';
                }""")
                print(f'[zoho] CONTENT_HTML: {content_html}')
                # Also find any links with gst/reconcil/gstr2b in their href
                gst_links = page.evaluate("""() => {
                    return Array.from(document.querySelectorAll('a')).filter(a =>
                        /gst|reconcil|gstr/i.test(a.href || a.getAttribute('href') || '')
                    ).map(a => ({text:(a.innerText||'').trim().slice(0,50), href:(a.href||'').slice(0,100)}));
                }""")
                print(f'[zoho] GST_LINKS: {gst_links}')
            except Exception as che:
                print(f'[zoho] content html failed: {che}')

            # Find the correct reconciliation URL by looking at the filings page links
            # The URL needs tax_return_id and tax_settings_id which change per period
            print(f'[zoho] finding reconciliation links from filings page...')
            page.wait_for_timeout(3000)
            recon_links = page.evaluate("""() => {
                return Array.from(document.querySelectorAll('a')).filter(a =>
                    /reconciliation/i.test(a.href || a.getAttribute('href') || '')
                ).map(a => ({
                    text: (a.innerText||'').trim().slice(0,80),
                    href: a.href || a.getAttribute('href') || ''
                }));
            }""")
            print(f'[zoho] reconciliation links found: {recon_links}')

            # Determine the single best reconciliation link to use
            best_recon_url = None
            if recon_links:
                # Prefer link matching from_period month
                fm_zoho = zoho_from  # e.g. '04-2026'
                for rl in recon_links:
                    if fm_zoho in rl['href']:
                        best_recon_url = rl['href']
                        break
                if not best_recon_url:
                    best_recon_url = recon_links[0]['href']
                print(f'[zoho] best_recon_url: {best_recon_url}')
            else:
                # Fallback: use URL pattern from user (with known IDs for this org)
                # These IDs are specific to org 60036724867 - may need updating for other periods
                best_recon_url = (
                    f'https://books.zoho.in/app/{zoho_org_id}'
                    f'#/gstfiling/tax/filings/reconciliation'
                    f'?from_date={zoho_from}&to_date={zoho_to}'
                    f'&per_page=100'
                    f'&tax_return_type=in_gstr2b_return'
                    f'&tax_return_id=2295010000011795035'
                    f'&tax_settings_id=2295010000000000271'
                    f'&txn_from_date=&txn_to_date='
                )
                print(f'[zoho] using fallback recon URL with known IDs')

            # Navigate to the reconciliation page
            print(f'[zoho] navigating to: {best_recon_url[:120]}')
            page.goto(best_recon_url, timeout=60000)
            page.wait_for_timeout(8000)
            print(f'[zoho] url after recon nav: {page.url[:100]}')

            # Wait for toolbar buttons to appear (reconciliation data loaded)
            try:
                page.wait_for_selector('.btn-toolbar button, .list-header button', timeout=20000)
                print('[zoho] reconciliation toolbar loaded')
            except PWTimeout:
                print('[zoho] toolbar wait timed out')

            page.wait_for_timeout(2000)

            # Log API calls to confirm data loaded
            _api_calls.clear()
            page.wait_for_timeout(3000)
            print(f'[zoho] API calls after recon nav ({len(_api_calls)}):')
            for c in _api_calls[-10:]:
                print(f'[zoho]   {c}')

            # Log toolbar buttons
            try:
                toolbar_btns = page.evaluate("""() => {
                    const btns = document.querySelectorAll('.btn-toolbar button, .list-header button, .btn-group button');
                    return Array.from(btns).map(b => ({
                        text:(b.innerText||'').trim().slice(0,40),
                        title:(b.title||'').slice(0,40),
                        aria:(b.getAttribute('aria-label')||'').slice(0,40),
                        cls:(b.className||'').slice(0,50)
                    }));
                }""")
                print(f'[zoho] toolbar_btns: {toolbar_btns}')
            except Exception:
                pass

            # Now export — single attempt (no month loop since we use the range URL)
            months = ['single']  # just one iteration
            for month_period in months:
                print(f'[zoho] url after nav: {page.url[:100]}')

                # Log ALL buttons + their inner SVG/icon text for this month
                try:
                    all_btns_info = page.evaluate("""() => {
                        const btns = document.querySelectorAll('button,[role="button"]');
                        return Array.from(btns).map(b => ({
                            text:(b.innerText||b.textContent||'').trim().slice(0,40),
                            title:(b.title||'').slice(0,40),
                            aria:(b.getAttribute('aria-label')||'').slice(0,40),
                            cls:(b.className||'').slice(0,60),
                            html: b.outerHTML.slice(0,150)
                        }));
                    }""")
                    for bi in (all_btns_info or []):
                        print(f'[zoho] BTN2 text="{bi["text"]}" title="{bi["title"]}" aria="{bi["aria"]}" cls="{bi["cls"][:40]}" html="{bi["html"][:100]}"')
                except Exception:
                    pass

                # Log accessibility snapshot to find export button
                try:
                    snap = page.accessibility.snapshot()
                    import json as _json
                    print(f'[zoho] A11Y: {_json.dumps(snap)[:3000]}')
                except Exception as ae:
                    print(f'[zoho] a11y failed: {ae}')

                # Try export: first try _click_export, then try ALL anonymous icon buttons
                def _try_all_icon_btns_for_download(pg, timeout_ms=30000):
                    """Try clicking each anonymous icon button and see if download fires."""
                    icon_btns = pg.evaluate("""() => {
                        const btns = document.querySelectorAll('button');
                        return Array.from(btns)
                            .filter(b => !(b.innerText||b.textContent||'').trim() && !b.title && !b.getAttribute('aria-label'))
                            .map((b, i) => i);
                    }""")
                    print(f'[zoho] anonymous icon btns: {len(icon_btns or [])}')
                    for idx in (icon_btns or []):
                        try:
                            with pg.expect_download(timeout=3000) as dl_chk:
                                pg.evaluate(f"""() => {{
                                    const btns = Array.from(document.querySelectorAll('button'))
                                        .filter(b => !(b.innerText||b.textContent||'').trim() && !b.title && !b.getAttribute('aria-label'));
                                    if (btns[{idx}]) btns[{idx}].click();
                                }}""")
                            return dl_chk.value
                        except PWTimeout:
                            continue
                    return None

                # Log main content area buttons specifically (not sidebar)
                try:
                    content_btns = page.evaluate("""() => {
                        // Exclude sidebar nav and chat buttons
                        const sidebar_cls = ['accordion-button','add-new','chat-button','rhs-sidebar','collapse-expand'];
                        const btns = document.querySelectorAll('button,[role="button"]');
                        return Array.from(btns).filter(b => {
                            const cls = b.className || '';
                            return !sidebar_cls.some(s => cls.includes(s)) &&
                                   !(b.getAttribute('aria-label')||'').includes('Close this side') &&
                                   !(b.getAttribute('aria-label')||'').includes('Chat');
                        }).map((b, i) => ({
                            i, text:(b.innerText||'').trim().slice(0,30),
                            title:(b.title||'').slice(0,30),
                            aria:(b.getAttribute('aria-label')||'').slice(0,30),
                            cls:(b.className||'').slice(0,50)
                        }));
                    }""")
                    print(f'[zoho] content_btns ({len(content_btns or [])}): {content_btns}')
                except Exception as cbe:
                    print(f'[zoho] content_btns failed: {cbe}')

                try:
                    with page.expect_download(timeout=40000) as dl_info:
                        clicked = _click_export(page)
                        if not clicked:
                            print(f'[zoho] _click_export failed for {month_period}')
                            # Try ONLY the d-flex icon-button position-relative buttons (main toolbar)
                            page.evaluate("""() => {
                                const btns = Array.from(document.querySelectorAll('button.d-flex.icon-button'));
                                console.log('[zoho] clicking', btns.length, 'd-flex icon buttons');
                                btns.forEach((b, i) => setTimeout(() => b.click(), i * 400));
                            }""")
                            page.wait_for_timeout(len(content_btns or []) * 400 + 2000)
                    dl = dl_info.value
                    excel_path = os.path.join(download_dir, dl.suggested_filename or f'gstr2b_{month_period}.xlsx')
                    dl.save_as(excel_path)
                    print(f'[zoho] downloaded: {excel_path}')
                    break
                except PWTimeout:
                    print(f'[zoho] download timeout for {month_period}, trying next month')
                    continue

            browser.close()

    except Exception as e:
        import shutil, traceback
        tb = traceback.format_exc()
        print(f'[zoho] EXCEPTION: {e}\n{tb}')
        shutil.rmtree(download_dir, ignore_errors=True)
        err_msg = str(e) or repr(e) or 'Browser automation exception (see server logs)'
        return jsonify({'ok': False, 'error': f'Browser automation failed: {err_msg}',
                        'detail': tb[-800:]}), 500

    if not excel_path or not os.path.exists(excel_path):
        return jsonify({'ok': False, 'error': 'Download did not complete — file not found.'}), 500

    # --- Run reconciliation engine on downloaded file ---
    try:
        result = run_reconciliation(excel_path, from_period,
                                    label or f'Zoho browser {from_period} to {to_period}',
                                    current_user.id,
                                    run_state=states[0] if len(states) == 1 else None)
    except Exception as e:
        import shutil
        shutil.rmtree(download_dir, ignore_errors=True)
        return jsonify({'ok': False, 'error': f'Recon engine error: {e}'}), 500

    import shutil
    shutil.rmtree(download_dir, ignore_errors=True)
    return jsonify({'ok': True, 'from_period': from_period, 'to_period': to_period,
                    'run_id': result.get('run_id'),
                    'rows': result.get('rows', 0), 'file': os.path.basename(excel_path)})


# ======================================================================
# SETTINGS + ZOHO LIVE INTEGRATION
# ======================================================================
def _zoho_creds():
    return {
        'client_id': get_setting('zoho_client_id'),
        'client_secret': get_setting('zoho_client_secret'),
        'refresh_token': get_setting('zoho_refresh_token'),
        'org_id': get_setting('zoho_org_id'),
        'region': get_setting('zoho_region', 'in'),
    }


def _zoho_configured():
    c = _zoho_creds()
    return all([c['client_id'], c['client_secret'], c['refresh_token'], c['org_id']])


def sync_zoho_master():
    """Fetch vendors from Zoho Books -> upsert VendorMaster + backfill row names.
    Returns dict {ok, count, updated_rows, error}."""
    c = _zoho_creds()
    try:
        tok = zoho.get_access_token(c['client_id'], c['client_secret'],
                                    c['refresh_token'], c['region'])
        vendors = zoho.fetch_vendors(tok, c['org_id'], c['region'])
    except Exception as e:
        set_setting('zoho_last_status', f'error: {e}')
        db.session.commit()
        return {'ok': False, 'error': str(e)}

    from state_codes import state_from_gstin
    updated_rows = 0
    for v in vendors:
        g = v['gstin'].strip()
        if not g:
            continue
        code, state = state_from_gstin(g)
        vm = db.session.get(VendorMaster, g)
        if vm is None:
            vm = VendorMaster(gstin=g)
            db.session.add(vm)
        if v['name']:
            vm.name = v['name']
        if v.get('email'):
            vm.email = v['email']
        vm.state_code, vm.state_name, vm.source = code, state, 'zoho_api'
        # backfill rows whose vendor is blank or equals the GSTIN (unknown)
        if v['name']:
            rows = ReconRow.query.filter(ReconRow.gstin == g,
                                         (ReconRow.vendor == None) | (ReconRow.vendor == '') |
                                         (ReconRow.vendor == g)).all()
            for r in rows:
                r.vendor = v['name']; updated_rows += 1
    set_setting('zoho_last_sync', now_ist().strftime('%d-%b-%Y %H:%M'))
    set_setting('zoho_last_status', f'ok: {len(vendors)} vendors')
    db.session.commit()
    return {'ok': True, 'count': len(vendors), 'updated_rows': updated_rows}


@app.route('/settings')
@superadmin_required
def settings_page():
    c = _zoho_creds()
    return render_template('settings.html',
        zoho_client_id=c['client_id'], zoho_org_id=c['org_id'],
        zoho_region=c['region'] or 'in',
        has_secret=bool(c['client_secret']), has_refresh=bool(c['refresh_token']),
        configured=_zoho_configured(),
        last_sync=get_setting('zoho_last_sync', '—'),
        last_status=get_setting('zoho_last_status', '—'),
        vendor_count=VendorMaster.query.count(),
        zoho_count=VendorMaster.query.filter_by(source='zoho_api').count(),
        # SMTP / email
        smtp_host=get_setting('smtp_host'), smtp_port=get_setting('smtp_port', '587'),
        smtp_user=get_setting('smtp_user'), smtp_from=get_setting('smtp_from'),
        smtp_tls=get_setting('smtp_tls', '1') != '0', smtp_configured=_smtp_configured(),
        has_smtp_pw=bool(get_setting('smtp_password')),
        cfo_email=get_setting('cfo_email'), cfo_send_day=get_setting('cfo_send_day', '1'),
        # Slack
        slack_configured=bool(get_setting('slack_webhook')),
        slack_send_day=get_setting('slack_send_day', '*'),
        # GSP / GSTR-2B source
        gsp_provider=get_setting('gsp_provider'), gsp_configured=bool(get_setting('gsp_provider')),
        # FIX 6 — dynamic email settings
        email_cc=get_setting('email_cc', 'mahesh.thakur@wiom.in, wiomfinance@wiom.in, tushar.gupta@wiom.in, ap@wiom.in'),
        email_sign=get_setting('email_sign', 'Regards,\nFinance and Taxation Team\nOmnia Information Private Limited'))


@app.route('/settings/zoho', methods=['POST'])
@superadmin_required
def save_zoho():
    set_setting('zoho_client_id', request.form.get('client_id', '').strip())
    set_setting('zoho_org_id', request.form.get('org_id', '').strip())
    set_setting('zoho_region', request.form.get('region', 'in').strip().lstrip('.'))
    # secrets: only overwrite when a new value is supplied (blank = keep existing)
    sec = request.form.get('client_secret', '').strip()
    ref = request.form.get('refresh_token', '').strip()
    if sec:
        set_setting('zoho_client_secret', sec)
    if ref:
        set_setting('zoho_refresh_token', ref)
    db.session.commit()
    from flask import flash
    flash('Zoho settings saved.', 'success')
    return redirect(url_for('settings_page'))


@app.route('/settings/zoho/test', methods=['POST'])
@superadmin_required
def test_zoho():
    if not _zoho_configured():
        return jsonify({'ok': False, 'error': 'Enter all credentials first.'})
    c = _zoho_creds()
    ok, msg = zoho.test_connection(c['client_id'], c['client_secret'],
                                   c['refresh_token'], c['org_id'], c['region'])
    return jsonify({'ok': ok, 'message': msg})


@app.route('/settings/zoho/sync', methods=['POST'])
@superadmin_required
def sync_zoho():
    if not _zoho_configured():
        return jsonify({'ok': False, 'error': 'Enter all credentials first.'})
    return jsonify(sync_zoho_master())


# ======================================================================
# EMAIL (SMTP) + scheduled CFO summary  (features 2 & 7)
# ======================================================================
def _smtp_cfg():
    return {
        'host': get_setting('smtp_host'), 'port': get_setting('smtp_port', '587'),
        'user': get_setting('smtp_user'), 'password': _decrypt(get_setting('smtp_password')),
        'from_addr': get_setting('smtp_from') or get_setting('smtp_user'),
        'use_tls': get_setting('smtp_tls', '1'),
    }


def _smtp_configured():
    return bool(get_setting('smtp_host') and get_setting('smtp_user'))


@app.route('/settings/smtp', methods=['POST'])
@superadmin_required
def save_smtp():
    set_setting('smtp_host', request.form.get('host', '').strip())
    set_setting('smtp_port', request.form.get('port', '587').strip())
    set_setting('smtp_user', request.form.get('user', '').strip())
    set_setting('smtp_from', request.form.get('from_addr', '').strip())
    set_setting('smtp_tls', '1' if request.form.get('use_tls') == 'on' else '0')
    pw = request.form.get('password', '').strip()
    if pw:
        set_setting('smtp_password', _encrypt(pw))
    set_setting('cfo_email', request.form.get('cfo_email', '').strip())
    set_setting('cfo_send_day', request.form.get('cfo_send_day', '1').strip())
    set_setting('email_cc', request.form.get('email_cc', '').strip())
    set_setting('email_sign', request.form.get('email_sign', '').strip())
    db.session.commit()
    from flask import flash
    flash('Email settings saved.', 'success')
    return redirect(url_for('settings_page'))


@app.route('/settings/smtp/test', methods=['POST'])
@superadmin_required
def test_smtp_route():
    import email_util
    ok, msg = email_util.test_smtp(_smtp_cfg())
    return jsonify({'ok': ok, 'message': msg})


def _send_cfo_email(recipient=None):
    """Email the CFO summary (HTML body + cumulative Excel attached). Returns (ok,msg)."""
    import email_util, export_excel
    to = recipient or get_setting('cfo_email')
    if not to:
        return False, 'No CFO recipient configured.'
    if not _smtp_configured():
        return False, 'SMTP not configured.'
    rows = ReconRow.query.order_by(ReconRow.state_name, ReconRow.period).all()
    ctx = _cfo_context(rows, 'all')
    gap = gap_from_rows(rows)
    xls = export_excel.build_cumulative_excel(rows, gap, 'All states · cumulative')
    body = render_template('cfo_email.html', **ctx)
    return email_util.send_email(_smtp_cfg(), to,
        f"WIOM GST Recon — CFO Summary ({ctx['generated']})", body,
        attachment=xls.read(), attachment_name=f"WIOM_Cumulative_{now_ist().strftime('%Y%m%d')}.xlsx")


@app.route('/cfo-email/send', methods=['POST'])
@admin_required
def send_cfo_email():
    ok, msg = _send_cfo_email(request.form.get('to') or None)
    return jsonify({'ok': ok, 'message': msg})


# ======================================================================
# SLACK daily status report
# ======================================================================
@app.route('/settings/slack', methods=['POST'])
@superadmin_required
def save_slack():
    set_setting('slack_send_day', request.form.get('slack_send_day', '*').strip() or '*')
    url = request.form.get('webhook', '').strip()
    if url:
        set_setting('slack_webhook', url)
    if request.form.get('clear_webhook') == 'on':
        set_setting('slack_webhook', '')
    db.session.commit()
    from flask import flash
    flash('Slack settings saved.', 'success')
    return redirect(url_for('settings_page'))


def _send_slack_report():
    """Build + post the daily recon status report to Slack. Returns (ok, msg)."""
    import slack_util
    url = get_setting('slack_webhook')
    if not url:
        return False, 'Slack webhook not configured.'
    rows = ReconRow.query.all()
    stats = _summary_stats_all()
    bd = _breakdown_data(rows)
    gap = gap_from_rows(rows)
    # state-wise gap
    st = {}
    for r in rows:
        s = st.setdefault(r.state_name, {'c': 0, 'v': 0.0})
        s['c'] += 1; s['v'] += r.total_diff or 0
    by_state = [(k, v['c'], v['v']) for k, v in sorted(st.items())]
    by_reason = [(k, v['count'], v['value']) for k, v in list(bd['by_reason'].items())[:6]]
    worst = sorted([g for g in gap if g['b_cnt'] and not g['g_cnt'] or g['risk'] in ('CRITICAL', 'HIGH')],
                   key=lambda g: -g['total_gap'])[:5]
    top_vendors = [(g['gstin'], (g['vendor'] or '')[:28], g['risk'],
                    (g['b_tax'] if g['b_cnt'] and not g['g_cnt'] else abs(g['total_gap'])))
                   for g in worst]
    kpis = [
        ('ITC at risk', f"₹{stats['itc_risk']:,}"),
        ('Total rows', f"{stats['total']:,}"),
        ('Open / Remarked / Done', f"{stats['open']} / {stats['remarked']} / {stats['done']}"),
    ]
    blocks = slack_util.build_report_blocks(
        f"WIOM GST Recon — Daily Status ({now_ist().strftime('%d-%b-%Y')})",
        kpis, by_state, by_reason, top_vendors)
    text = f"WIOM GST Recon daily: ITC at risk ₹{stats['itc_risk']:,}, {stats['open']} open, {stats['done']} resolved"
    return slack_util.post_message(url, text, blocks)


@app.route('/slack/send', methods=['POST'])
@admin_required
def slack_send():
    ok, msg = _send_slack_report()
    return jsonify({'ok': ok, 'message': msg})


def _summary_stats_all():
    """Org-wide stats (no user state restriction) for scheduled reports. Single query."""
    from sqlalchemy import func, case
    row = db.session.query(
        func.count().label('total'),
        func.sum(case((ReconRow.status == 'open', 1), else_=0)).label('open'),
        func.sum(case((ReconRow.status == 'remarked', 1), else_=0)).label('remarked'),
        func.sum(case((ReconRow.status.in_(['approved', 'resolved']), 1), else_=0)).label('done'),
        func.sum(case((ReconRow.category == 'books_only',
                       func.coalesce(ReconRow.books_igst, 0) +
                       func.coalesce(ReconRow.books_cgst, 0) +
                       func.coalesce(ReconRow.books_sgst, 0)), else_=0)).label('itc_risk'),
    ).one()
    return {
        'total': row.total or 0,
        'open': row.open or 0,
        'remarked': row.remarked or 0,
        'done': row.done or 0,
        'itc_risk': round(row.itc_risk or 0),
    }


# ======================================================================
# AUDIT LOG
# ======================================================================
@app.route('/audit')
@admin_required
def audit_page():
    periods = [p[0] for p in db.session.query(ReconRow.period)
               .distinct().order_by(ReconRow.period.desc()).all()]
    return render_template('audit.html', periods=periods, states=WIOM_STATES)


@app.route('/api/audit')
@admin_required
def api_audit():
    action = request.args.get('action')
    state = request.args.get('state')
    search = request.args.get('q', '').strip()
    # join audit -> row for GSTIN/state context
    q = db.session.query(AuditLog, ReconRow).outerjoin(
        ReconRow, AuditLog.row_id == ReconRow.id)
    if action and action != 'all':
        q = q.filter(AuditLog.action == action)
    if state and state != 'all':
        q = q.filter(db.or_(ReconRow.state_name == state, AuditLog.row_id == None))
    if search:
        like = f'%{search}%'
        q = q.filter(db.or_(ReconRow.gstin.like(like), ReconRow.vendor.like(like),
                     AuditLog.user_name.like(like), AuditLog.new_value.like(like)))
    rows = q.order_by(AuditLog.created_at.desc()).limit(1000).all()
    return jsonify([{
        'at': l.created_at.strftime('%d-%b-%Y %H:%M'),
        'user': l.user_name, 'action': l.action,
        'gstin': r.gstin if r else '',
        'vendor': r.vendor if r else '',
        'inv': (r.books_inv or r.gstn_inv or '') if r else '',
        'txn_type': r.txn_type if r else '',
        'state': r.state_name if r else '',
        'period': r.period if r else '',
        'new': l.new_value, 'old': l.old_value,
    } for l, r in rows])


# ======================================================================
# UPLOAD + ENGINE
# ======================================================================
@app.route('/upload', methods=['POST'])
@admin_required
def upload():
    if processing_state['active']:
        return jsonify({'error': 'A reconciliation is already running.'}), 400

    label = request.form.get('label', '').strip()
    run_state = request.form.get('state', '').strip()
    period = request.form.get('period', '').strip()  # 'YYYY-MM' from the month picker
    if run_state not in WIOM_STATES:
        return jsonify({'error': 'Select a valid state (Delhi / Haryana / Maharashtra / Uttar Pradesh).'}), 400
    import re as _re
    if not _re.match(r'^\d{4}-\d{2}$', period):
        return jsonify({'error': 'Select the return month (period).'}), 400

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'No file selected'}), 400
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Only .xlsx/.xls files accepted'}), 400

    filename = secure_filename(file.filename)
    ts = now_ist().strftime('%Y%m%d_%H%M%S')
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], f'{ts}_{filename}')
    file.save(filepath)

    reset_state()
    add_log(0, f'File uploaded for {run_state} ({period}): {filename}')
    threading.Thread(target=run_reconciliation,
                     args=(filepath, period, label, current_user.id, run_state),
                     daemon=True).start()
    return jsonify({'status': 'started', 'filename': filename, 'period': period})


@app.route('/status')
@login_required
def status():
    return jsonify({
        'active': processing_state['active'], 'progress': processing_state['progress'],
        'current_agent': processing_state['current_agent'],
        'logs': processing_state['logs'][-60:],
        'agent_results': processing_state['agent_results'],
        'run_id': processing_state['run_id'],
        'output_file': processing_state['output_file'],
        'error': processing_state['error'],
        'elapsed': round(time.time() - processing_state['start_time'], 1)
                   if processing_state['start_time'] else 0,
    })


@app.route('/download')
@login_required
def download():
    run_id = request.args.get('run_id', type=int)
    if run_id:
        run = db.session.get(ReconRun, run_id)
        path = run.output_file if run else None
    else:
        path = processing_state['output_file']
    if path and os.path.exists(path):
        return send_file(path, as_attachment=True)
    return jsonify({'error': 'No output file available'}), 404


# ======================================================================
# WORKFLOW API
# ======================================================================
@app.route('/api/rows')
@login_required
def api_rows():
    q = ReconRow.query
    period = request.args.get('period')
    state = request.args.get('state')
    category = request.args.get('category')
    wf_status = request.args.get('status')
    search = request.args.get('q', '').strip()
    only_mismatch = request.args.get('mismatch') == '1'

    if period and period != 'all':
        q = q.filter(ReconRow.period <= period)  # cumulative: show all months up to selected
    if state and state != 'all':
        q = q.filter(ReconRow.state_name == state)
    if category and category != 'all':
        q = q.filter(ReconRow.category == category)
    if request.args.get('recon') == 'fully':
        q = q.filter(ReconRow.category == 'matched',
                     ReconRow.recon_status.like('%Fully Reconciled%'))
    if request.args.get('recon') == 'cross':
        q = q.filter(ReconRow.category == 'matched',
                     ~ReconRow.recon_status.like('%Fully Reconciled%'))
    if request.args.get('recon') == 'rejected':
        # Rejected tab: always show only rejected, ignore wf_status filter
        q = q.filter(ReconRow.status == 'rejected')
    else:
        if request.args.get('exclude_rejected') == '1':
            q = q.filter(ReconRow.status != 'rejected')
        if wf_status and wf_status != 'all':
            q = q.filter(ReconRow.status == wf_status)
    q = _apply_fy(q, request.args.get('fy'))
    if only_mismatch:
        # everything except cleanly matched + fully reconciled rows
        q = q.filter(~((ReconRow.category == 'matched') &
                       (ReconRow.recon_status.like('%Fully Reconciled%'))))
    # My Queue: Book-Keeping → rows assigned to me (or, if none assigned, my-state pending);
    #           Admin+ → rows pending approval (remarked).
    if request.args.get('my') == '1':
        if current_user.is_admin:
            q = q.filter(ReconRow.status == 'remarked')
        else:
            assigned_n = ReconRow.query.filter_by(assigned_to_id=current_user.id).count()
            if assigned_n:
                q = q.filter(ReconRow.assigned_to_id == current_user.id)
            else:
                q = q.filter(ReconRow.status.in_(['open', 'remarked']),
                             ReconRow.category != 'matched')
    if search:
        like = f'%{search}%'
        q = q.filter(ReconRow.gstin.like(like) | ReconRow.vendor.like(like) |
                     ReconRow.books_inv.like(like) | ReconRow.gstn_inv.like(like))

    # Book-Keeping users only see their assigned states (blank = all)
    if not current_user.is_admin and current_user.state_list():
        q = q.filter(ReconRow.state_name.in_(current_user.state_list()))

    rows = q.order_by(ReconRow.state_name, ReconRow.gstin).all()
    # Batch load attachment counts (single query instead of N+1)
    if rows:
        row_ids = [r.id for r in rows]
        counts = dict(db.session.execute(
            db.text('SELECT row_id, COUNT(*) FROM row_attachments WHERE row_id IN :ids GROUP BY row_id'),
            {'ids': tuple(row_ids) if len(row_ids) > 1 else (row_ids[0], row_ids[0])}
        ).fetchall())
        for r in rows:
            r._attach_count = counts.get(r.id, 0)
    return jsonify({'rows': [r.to_dict() for r in rows], 'count': len(rows)})


def _apply_fy(q, fy):
    """Filter to an Indian financial year (Apr YYYY – Mar YYYY+1). fy = 'YYYY'."""
    if not fy or fy == 'all':
        return q
    try:
        y = int(fy)
    except (TypeError, ValueError):
        return q
    return q.filter(ReconRow.period >= f'{y}-04', ReconRow.period <= f'{y + 1}-03')


@app.route('/api/row/<int:row_id>/remark', methods=['POST'])
@write_required
def api_remark(row_id):
    row = db.session.get(ReconRow, row_id)
    if not row:
        abort(404)
    if not current_user.can_see_state(row.state_name):
        abort(403)
    data = request.get_json(force=True)
    old = row.team_remark
    row.team_remark = data.get('remark', '').strip()
    row.team_reason = data.get('reason', '').strip()
    row.remarked_by_id = current_user.id
    row.remarked_at = now_ist()
    if row.status == 'open':
        row.status = 'remarked'
    log_audit(row.id, current_user, 'remark', 'team_remark', old, row.team_remark)
    db.session.commit()
    return jsonify({'ok': True, 'row': row.to_dict()})


@app.route('/api/row/<int:row_id>/approve', methods=['POST'])
@admin_required
def api_approve(row_id):
    row = db.session.get(ReconRow, row_id)
    if not row:
        abort(404)
    data = request.get_json(force=True)
    action = data.get('action', 'approve')  # 'approve' | 'resolve' | 'reopen'
    if action == 'reopen':
        row.status = 'remarked' if row.team_remark else 'open'
        row.approved_by_id = None
        row.approved_at = None
    else:
        row.status = 'resolved' if action == 'resolve' else 'approved'
        row.approved_by_id = current_user.id
        row.approved_at = now_ist()
        row.resolution_note = data.get('note', '').strip()
    log_audit(row.id, current_user, action, 'status', '', row.status)
    db.session.commit()
    return jsonify({'ok': True, 'row': row.to_dict()})


@app.route('/api/counts')
@login_required
def api_counts():
    """Row counts per tab (respecting period/state + user state restriction)."""
    def base():
        q = ReconRow.query
        period = request.args.get('period')
        state = request.args.get('state')
        search = request.args.get('q', '').strip()
        if period and period != 'all':
            q = q.filter(ReconRow.period <= period)  # cumulative: show all months up to selected
        if state and state != 'all':
            q = q.filter(ReconRow.state_name == state)
        q = _apply_fy(q, request.args.get('fy'))
        if search:
            like = f'%{search}%'
            q = q.filter(ReconRow.gstin.like(like) | ReconRow.vendor.like(like) |
                         ReconRow.books_inv.like(like) | ReconRow.gstn_inv.like(like))
        if not current_user.is_admin and current_user.state_list():
            q = q.filter(ReconRow.state_name.in_(current_user.state_list()))
        return q
    def amt(q, col):
        return round(q.with_entities(func.coalesce(func.sum(col), 0)).scalar() or 0)
    fully_q    = base().filter(ReconRow.category == 'matched',
                               ReconRow.recon_status.like('%Fully Reconciled%'))
    cross_q    = base().filter(ReconRow.category == 'matched',
                               ~ReconRow.recon_status.like('%Fully Reconciled%'))
    books_q    = base().filter(ReconRow.category == 'books_only', ReconRow.status != 'rejected')
    gstn_q     = base().filter(ReconRow.category == 'gstn_only',  ReconRow.status != 'rejected')
    rejected_q = base().filter(ReconRow.status == 'rejected')
    pending_q  = base().filter(ReconRow.status == 'remarked')
    return jsonify({
        'cross': cross_q.count(),         'cross_amt': amt(cross_q, ReconRow.books_total),
        'books': books_q.count(),         'books_amt': amt(books_q, ReconRow.books_total),
        'gstn': gstn_q.count(),           'gstn_amt': amt(gstn_q, ReconRow.gstn_total),
        'fully': fully_q.count(),         'fully_amt': amt(fully_q, ReconRow.books_total),
        'rejected': rejected_q.count(),   'rejected_amt': amt(rejected_q, ReconRow.gstn_total),
        'pending': pending_q.count(),
        'gap': base().with_entities(ReconRow.gstin).distinct().count(),
        'gap_amt': amt(base(), ReconRow.total_diff),
        'total': base().count(),
    })


@app.route('/api/gstr3b')
@login_required
def api_gstr3b():
    """GSTR-3B Table 4 summary — row counts + ITC amounts by bucket."""
    from sqlalchemy import func
    fy = request.args.get('fy', '')
    period = request.args.get('period', '')

    def base():
        q = ReconRow.query
        if not current_user.is_admin:
            states = current_user.state_list()
            if states:
                q = q.filter(ReconRow.state_name.in_(states))
        if period:
            q = q.filter(ReconRow.period <= period)
        return q

    rows = base().with_entities(
        ReconRow.itc_table4,
        func.count(ReconRow.id).label('cnt'),
        func.sum(ReconRow.books_igst + ReconRow.books_cgst + ReconRow.books_sgst).label('itc'),
        func.sum(ReconRow.books_taxable).label('taxable'),
    ).group_by(ReconRow.itc_table4).all()

    buckets = {}
    untagged_cnt = 0; untagged_itc = 0
    for r in rows:
        tag = (r.itc_table4 or '').strip()
        if not tag:
            untagged_cnt += int(r.cnt or 0)
            untagged_itc += float(r.itc or 0)
        else:
            buckets[tag] = {'cnt': int(r.cnt or 0), 'itc': float(r.itc or 0), 'taxable': float(r.taxable or 0)}

    # also compute totals per section
    def section_total(keys):
        return {
            'cnt': sum(buckets.get(k, {}).get('cnt', 0) for k in keys),
            'itc': sum(buckets.get(k, {}).get('itc', 0) for k in keys),
        }

    return jsonify({
        'buckets': buckets,
        'untagged': {'cnt': untagged_cnt, 'itc': untagged_itc},
        'total_tagged_cnt': sum(v['cnt'] for v in buckets.values()),
        'total_tagged_itc': sum(v['itc'] for v in buckets.values()),
        '4A_total': section_total(['4A1','4A2','4A3','4A4','4A5']),
        '4B_total': section_total(['4B1','4B2']),
        '4D_total': section_total(['4D1','4D2']),
        'labels': _TABLE4_OPTIONS,
    })


@app.route('/api/rows/bulk', methods=['POST'])
@write_required
def api_bulk():
    """Apply an action to many rows at once.
    remark -> any logged-in user (own states); approve/resolve/reopen -> admin only."""
    data = request.get_json(force=True)
    ids = data.get('ids', [])
    action = data.get('action', '')
    if not ids:
        return jsonify({'ok': False, 'error': 'No rows selected.'}), 400

    rows = ReconRow.query.filter(ReconRow.id.in_(ids)).all()
    changed = 0

    # FIX 3 — check if any of these rows belong to a locked period
    for r in rows:
        run = db.session.get(ReconRun, r.run_id)
        if run and run.locked:
            return jsonify({'ok': False, 'error': 'Period is locked. Contact Super Admin to unlock.'}), 403

    if action == 'remark':
        remark = data.get('remark', '').strip()
        reason = data.get('reason', '').strip()
        for row in rows:
            if not current_user.can_see_state(row.state_name):
                continue
            old = row.team_remark
            if remark:
                row.team_remark = remark
            if reason:
                row.team_reason = reason
            row.remarked_by_id = current_user.id
            row.remarked_at = now_ist()
            if row.status == 'open':
                row.status = 'remarked'
            log_audit(row.id, current_user, 'remark', 'team_remark', old, row.team_remark)
            changed += 1
    elif action == 'reject':
        if not current_user.is_admin:
            abort(403)
        reason = data.get('reason', '').strip() or 'ITC Ineligible — Rejected'
        for row in rows:
            if not current_user.can_see_state(row.state_name):
                continue
            old_status = row.status
            row.status = 'rejected'
            row.team_reason = reason
            row.approved_by_id = current_user.id
            row.approved_at = now_ist()
            log_audit(row.id, current_user, 'reject', 'status', old_status, 'rejected')
            changed += 1
    elif action in ('approve', 'resolve', 'reopen'):
        if not current_user.is_admin:
            abort(403)
        note = data.get('note', '').strip()
        for row in rows:
            if action == 'reopen':
                row.status = 'remarked' if row.team_remark else 'open'
                row.approved_by_id = None
                row.approved_at = None
            else:
                row.status = 'resolved' if action == 'resolve' else 'approved'
                row.approved_by_id = current_user.id
                row.approved_at = now_ist()
                if note:
                    row.resolution_note = note
            log_audit(row.id, current_user, action, 'status', '', row.status)
            changed += 1
    elif action == 'assign':
        if not current_user.is_admin:
            abort(403)
        aid = data.get('assignee_id')
        assignee = db.session.get(User, int(aid)) if aid else None
        if aid and not assignee:
            return jsonify({'ok': False, 'error': 'Unknown assignee.'}), 400
        for row in rows:
            row.assigned_to_id = assignee.id if assignee else None
            log_audit(row.id, current_user, 'assign', 'assigned_to', '',
                      assignee.name if assignee else '(unassigned)')
            changed += 1
    elif action == 'table4':
        val = (data.get('table4') or '').strip()
        if val and val not in _TABLE4_OPTIONS:
            return jsonify({'ok': False, 'error': 'Invalid Table 4 value.'}), 400
        for row in rows:
            old = row.itc_table4 or ''
            row.itc_table4 = val
            log_audit(row.id, current_user, 'table4', 'itc_table4', old, val)
            changed += 1
    else:
        return jsonify({'ok': False, 'error': 'Unknown action.'}), 400

    db.session.commit()
    return jsonify({'ok': True, 'changed': changed})


@app.route('/api/assignees')
@login_required
def api_assignees():
    """Book-Keeping users a row can be assigned to."""
    users = User.query.filter_by(role='user', active=True).order_by(User.name).all()
    return jsonify([{'id': u.id, 'name': u.name} for u in users])


def _breakdown_data(rows):
    """Split rows by workflow status and by team reason (RCM / ineligible / etc.),
    with count + amount at stake. Feeds dashboard, CFO summary and the Slack report."""
    def amt(r):
        # exposure: ITC tax for books-only (at risk), else |total diff|
        if r.category == 'books_only':
            return (r.books_igst or 0) + (r.books_cgst or 0) + (r.books_sgst or 0)
        return abs(r.total_diff or 0)
    by_status, by_reason, other_breakdown = {}, {}, {}
    for r in rows:
        st = r.status or 'open'
        s = by_status.setdefault(st, {'count': 0, 'value': 0.0})
        s['count'] += 1; s['value'] += amt(r)
        rsn = (r.team_reason or '').strip()
        if rsn and rsn != 'Other':
            d = by_reason.setdefault(rsn, {'count': 0, 'value': 0.0})
            d['count'] += 1; d['value'] += amt(r)
        else:
            # "Other" or untagged — break down by system recon_status for visibility
            sys_key = (r.recon_status or r.category or 'Unknown').strip()
            # Shorten long auto labels
            sys_key = sys_key.replace(' (Zoho Auto)', '').replace('GSTIN NOT in ', '').strip()
            if len(sys_key) > 55:
                sys_key = sys_key[:52] + '…'
            bucket = 'Other' if rsn == 'Other' else 'Untagged'
            ob = other_breakdown.setdefault(bucket, {})
            od = ob.setdefault(sys_key, {'count': 0, 'value': 0.0})
            od['count'] += 1; od['value'] += amt(r)

    # Rejected rows: breakdown by team_reason
    rejected_breakdown = {}
    for r in rows:
        if (r.status or '') != 'rejected':
            continue
        rsn = (r.team_reason or 'No reason given').strip()
        d = rejected_breakdown.setdefault(rsn, {'count': 0, 'value': 0.0})
        d['count'] += 1; d['value'] += amt(r)

    rnd = lambda d: {k: {'count': v['count'], 'value': round(v['value'])} for k, v in d.items()}
    rnd_ob = {bucket: rnd(subs) for bucket, subs in other_breakdown.items()}
    return {'by_status': rnd(by_status),
            'by_reason': dict(sorted(rnd(by_reason).items(), key=lambda kv: -kv[1]['value'])),
            'other_breakdown': rnd_ob,
            'rejected_breakdown': dict(sorted(rnd(rejected_breakdown).items(), key=lambda kv: -kv[1]['count']))}


@app.route('/api/breakdown')
@login_required
def api_breakdown():
    return jsonify(_breakdown_data(_filtered_rows_query().all()))


def _filtered_rows_query():
    """ReconRow query filtered by period/state/fy/search/gstin/category/recon/status args + user's state restriction."""
    q = ReconRow.query
    period = request.args.get('period')
    state = request.args.get('state')
    search = request.args.get('q', '').strip()
    gstin_exact = request.args.get('gstin', '').strip()
    category = request.args.get('category')
    recon = request.args.get('recon')
    wf_status = request.args.get('status')
    if period and period != 'all':
        q = q.filter(ReconRow.period <= period)
    if state and state != 'all':
        q = q.filter(ReconRow.state_name == state)
    q = _apply_fy(q, request.args.get('fy'))
    if gstin_exact:
        q = q.filter(ReconRow.gstin == gstin_exact)
    elif search:
        like = f'%{search}%'
        q = q.filter(ReconRow.gstin.like(like) | ReconRow.vendor.like(like) |
                     ReconRow.books_inv.like(like) | ReconRow.gstn_inv.like(like))
    # Tab-specific filters (used by filtered export)
    if recon == 'fully':
        q = q.filter(ReconRow.category == 'matched', ReconRow.recon_status.like('%Fully Reconciled%'))
    elif recon == 'cross':
        q = q.filter(ReconRow.category == 'matched', ~ReconRow.recon_status.like('%Fully Reconciled%'))
    elif recon == 'rejected':
        q = q.filter(ReconRow.status == 'rejected')
    else:
        if category and category != 'all':
            q = q.filter(ReconRow.category == category)
        if wf_status and wf_status not in ('all', ''):
            q = q.filter(ReconRow.status == wf_status)
    if not current_user.is_admin and current_user.state_list():
        q = q.filter(ReconRow.state_name.in_(current_user.state_list()))
    return q


def gap_from_rows(reconrows):
    """Aggregate ReconRows per GSTIN into the gap-analysis rows (Books vs 2B)."""
    agg = {}
    for r in reconrows:
        a = agg.get(r.gstin)
        if a is None:
            a = agg[r.gstin] = {
                'gstin': r.gstin, 'vendor': r.vendor, 'state': r.state_name,
                'b_cnt': 0, 'b_taxable': 0.0, 'b_tax': 0.0, 'b_total': 0.0,
                'g_cnt': 0, 'g_taxable': 0.0, 'g_tax': 0.0, 'g_total': 0.0,
            }
        if r.vendor and (not a['vendor'] or a['vendor'] == r.gstin):
            a['vendor'] = r.vendor
        b_tax = (r.books_igst or 0) + (r.books_cgst or 0) + (r.books_sgst or 0)
        g_tax = (r.gstn_igst or 0) + (r.gstn_cgst or 0) + (r.gstn_sgst or 0)
        if r.category in ('matched', 'books_only'):
            a['b_cnt'] += 1; a['b_taxable'] += r.books_taxable or 0
            a['b_tax'] += b_tax; a['b_total'] += r.books_total or 0
        if r.category in ('matched', 'gstn_only'):
            a['g_cnt'] += 1; a['g_taxable'] += r.gstn_taxable or 0
            a['g_tax'] += g_tax; a['g_total'] += r.gstn_total or 0

    rows = []
    for a in agg.values():
        total_gap = a['b_total'] - a['g_total']
        if a['b_cnt'] == 0 and a['g_cnt'] > 0:
            risk, remark = 'HIGH', 'Only in GSTR-2B'
        elif a['b_cnt'] > 0 and a['g_cnt'] == 0:
            risk, remark = 'CRITICAL', 'Only in Books — vendor not filed'
        elif abs(total_gap) < 1:
            risk, remark = 'LOW', 'Amount matched'
        elif abs(total_gap) > 100000:
            risk, remark = 'HIGH', ('Books > 2B' if total_gap > 0 else '2B > Books')
        else:
            risk, remark = 'MEDIUM', ('Books > 2B' if total_gap > 0 else '2B > Books')
        action = {'CRITICAL': 'URGENT: Vendor follow-up', 'HIGH': 'Investigate',
                  'LOW': 'Verify invoice nos'}.get(risk, 'Review')
        denom = max(abs(a['b_total']), abs(a['g_total']))
        rows.append({
            'gstin': a['gstin'], 'vendor': a['vendor'], 'state': a['state'],
            'b_cnt': a['b_cnt'], 'b_taxable': round(a['b_taxable'], 2),
            'b_tax': round(a['b_tax'], 2), 'b_total': round(a['b_total'], 2),
            'g_cnt': a['g_cnt'], 'g_taxable': round(a['g_taxable'], 2),
            'g_tax': round(a['g_tax'], 2), 'g_total': round(a['g_total'], 2),
            'taxable_gap': round(a['b_taxable'] - a['g_taxable'], 2),
            'tax_gap': round(a['b_tax'] - a['g_tax'], 2),
            'total_gap': round(total_gap, 2),
            'gap_pct': round(total_gap / denom, 4) if denom else 0,
            'risk': risk, 'remark': remark, 'action': action,
        })
    rows.sort(key=lambda x: abs(x['total_gap']), reverse=True)
    return rows


@app.route('/api/gap-analysis')
@login_required
def api_gap_analysis():
    """Per-GSTIN gap analysis using SQL aggregation — no Python-level row loading."""
    from sqlalchemy import func, case
    q = _filtered_rows_query()
    agg = q.with_entities(
        ReconRow.gstin,
        func.max(ReconRow.vendor).label('vendor'),
        func.max(ReconRow.state_name).label('state'),
        func.sum(case((ReconRow.category.in_(['matched', 'books_only']), 1), else_=0)).label('b_cnt'),
        func.sum(case((ReconRow.category.in_(['matched', 'books_only']),
                       func.coalesce(ReconRow.books_taxable, 0)), else_=0)).label('b_taxable'),
        func.sum(case((ReconRow.category.in_(['matched', 'books_only']),
                       func.coalesce(ReconRow.books_igst, 0) +
                       func.coalesce(ReconRow.books_cgst, 0) +
                       func.coalesce(ReconRow.books_sgst, 0)), else_=0)).label('b_tax'),
        func.sum(case((ReconRow.category.in_(['matched', 'books_only']),
                       func.coalesce(ReconRow.books_total, 0)), else_=0)).label('b_total'),
        func.sum(case((ReconRow.category.in_(['matched', 'gstn_only']), 1), else_=0)).label('g_cnt'),
        func.sum(case((ReconRow.category.in_(['matched', 'gstn_only']),
                       func.coalesce(ReconRow.gstn_taxable, 0)), else_=0)).label('g_taxable'),
        func.sum(case((ReconRow.category.in_(['matched', 'gstn_only']),
                       func.coalesce(ReconRow.gstn_igst, 0) +
                       func.coalesce(ReconRow.gstn_cgst, 0) +
                       func.coalesce(ReconRow.gstn_sgst, 0)), else_=0)).label('g_tax'),
        func.sum(case((ReconRow.category.in_(['matched', 'gstn_only']),
                       func.coalesce(ReconRow.gstn_total, 0)), else_=0)).label('g_total'),
    ).group_by(ReconRow.gstin).all()

    rows = []
    for a in agg:
        b_total = float(a.b_total or 0); g_total = float(a.g_total or 0)
        b_tax = float(a.b_tax or 0);     g_tax = float(a.g_tax or 0)
        b_taxable = float(a.b_taxable or 0); g_taxable = float(a.g_taxable or 0)
        total_gap = b_total - g_total
        if a.b_cnt == 0 and a.g_cnt > 0:
            risk, remark = 'HIGH', 'Only in GSTR-2B'
        elif a.b_cnt > 0 and a.g_cnt == 0:
            risk, remark = 'CRITICAL', 'Only in Books — vendor not filed'
        elif abs(total_gap) < 1:
            risk, remark = 'LOW', 'Amount matched'
        elif abs(total_gap) > 100000:
            risk, remark = 'HIGH', ('Books > 2B' if total_gap > 0 else '2B > Books')
        else:
            risk, remark = 'MEDIUM', ('Books > 2B' if total_gap > 0 else '2B > Books')
        action = {'CRITICAL': 'URGENT: Vendor follow-up', 'HIGH': 'Investigate',
                  'LOW': 'Verify invoice nos'}.get(risk, 'Review')
        denom = max(abs(b_total), abs(g_total))
        rows.append({
            'gstin': a.gstin, 'vendor': a.vendor, 'state': a.state,
            'b_cnt': a.b_cnt, 'b_taxable': round(b_taxable, 2),
            'b_tax': round(b_tax, 2), 'b_total': round(b_total, 2),
            'g_cnt': a.g_cnt, 'g_taxable': round(g_taxable, 2),
            'g_tax': round(g_tax, 2), 'g_total': round(g_total, 2),
            'taxable_gap': round(b_taxable - g_taxable, 2),
            'tax_gap': round(b_tax - g_tax, 2),
            'total_gap': round(total_gap, 2),
            'gap_pct': round(total_gap / denom, 4) if denom else 0,
            'risk': risk, 'remark': remark, 'action': action,
        })
    rows.sort(key=lambda x: abs(x['total_gap']), reverse=True)
    return jsonify({'rows': rows, 'count': len(rows)})


# ---- Feature 5: Vendor compliance scorecard ----
def _grade(score):
    return 'A' if score >= 90 else 'B' if score >= 75 else 'C' if score >= 50 else 'D'


@app.route('/scorecard')
@login_required
def scorecard_page():
    return render_template('scorecard.html', states=WIOM_STATES)


@app.route('/api/scorecard')
@login_required
def api_scorecard():
    """Per-vendor compliance: how reliably they appear in GSTR-2B vs Books."""
    from sqlalchemy import func, case
    base = _filtered_rows_query()
    agg_rows = base.with_entities(
        ReconRow.gstin,
        func.max(ReconRow.vendor).label('vendor'),
        func.count(func.distinct(ReconRow.period)).label('months'),
        func.sum(case((ReconRow.category == 'matched', 1), else_=0)).label('matched'),
        func.sum(case((ReconRow.category == 'books_only', 1), else_=0)).label('books_only'),
        func.sum(case((ReconRow.category == 'gstn_only', 1), else_=0)).label('gstn_only'),
        func.sum(case((ReconRow.category == 'books_only',
                       ReconRow.books_igst + ReconRow.books_cgst + ReconRow.books_sgst), else_=0)).label('itc_risk'),
        func.sum(case(((ReconRow.category == 'books_only') & ReconRow.status.notin_(['approved','resolved']), 1), else_=0)).label('open'),
    ).group_by(ReconRow.gstin).all()
    out = []
    for a in agg_rows:
        filed = (a.matched or 0) + (a.gstn_only or 0)
        not_filed = a.books_only or 0
        denom = filed + not_filed
        score = round(100 * filed / denom) if denom else 100
        out.append({
            'gstin': a.gstin, 'vendor': a.vendor, 'months': a.months or 0,
            'matched': a.matched or 0, 'books_only': not_filed, 'gstn_only': a.gstn_only or 0,
            'reliability': score, 'grade': _grade(score),
            'itc_risk': round(a.itc_risk or 0), 'open': a.open or 0,
        })
    out.sort(key=lambda x: (x['reliability'], -x['itc_risk']))
    return jsonify({'rows': out, 'count': len(out)})


@app.route('/export')
@login_required
def export_cumulative():
    """Download a CUMULATIVE multi-sheet Excel (all months to date) from the DB,
    optionally scoped to a state. Mirrors the report tabs + workflow columns."""
    import export_excel
    rows = _filtered_rows_query().order_by(
        ReconRow.state_name, ReconRow.period, ReconRow.gstin).all()
    gap = gap_from_rows(rows)
    state = request.args.get('state')
    scope = (state if state and state != 'all' else 'All states') + ' · cumulative to date'
    bio = export_excel.build_cumulative_excel(rows, gap, scope)
    recon = request.args.get('recon', '')
    category = request.args.get('category', '')
    tab_map = {'fully': 'FullyRecon', 'cross': 'CrossMatch', 'rejected': 'Rejected'}
    tab_label = tab_map.get(recon, '') or ('BooksOnly' if category == 'books' else '')
    fname = f"WIOM_{(tab_label+'_') if tab_label else 'Cumulative_'}{(state or 'AllStates').replace(' ', '')}_{now_ist().strftime('%Y%m%d')}.xlsx"
    return send_file(bio, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ---- Feature 8: comment thread per row ----
@app.route('/api/row/<int:row_id>/comments', methods=['GET', 'POST'])
@login_required
def api_comments(row_id):
    row = db.session.get(ReconRow, row_id)
    if not row:
        abort(404)
    if not current_user.can_see_state(row.state_name):
        abort(403)
    if request.method == 'POST':
        text = (request.get_json(force=True).get('text') or '').strip()
        if text:
            db.session.add(RowComment(row_id=row_id, user_id=current_user.id,
                user_name=current_user.name, user_role=current_user.role_display, text=text))
            log_audit(row_id, current_user, 'comment', 'discussion', '', text)
            db.session.commit()
    cs = RowComment.query.filter_by(row_id=row_id).order_by(RowComment.created_at).all()
    return jsonify([{'user': c.user_name, 'role': c.user_role, 'text': c.text,
                     'at': c.created_at.strftime('%d-%b-%Y %H:%M')} for c in cs])


# ---- Row history (audit log + comments merged, newest first) ----
@app.route('/api/row/<int:row_id>/history')
@login_required
def api_row_history(row_id):
    row = db.session.get(ReconRow, row_id)
    if not row:
        abort(404)
    if not current_user.can_see_state(row.state_name):
        abort(403)
    result = []
    for e in AuditLog.query.filter_by(row_id=row_id).all():
        result.append({
            'type': 'audit',
            'user': e.user_name,
            'action': e.action,
            'old': e.old_value,
            'new': e.new_value,
            'at': e.created_at.strftime('%d-%b-%Y %H:%M'),
            '_ts': e.created_at,
        })
    for c in RowComment.query.filter_by(row_id=row_id).all():
        result.append({
            'type': 'comment',
            'user': c.user_name,
            'action': 'comment',
            'old': '',
            'new': c.text,
            'at': c.created_at.strftime('%d-%b-%Y %H:%M'),
            '_ts': c.created_at,
        })
    result.sort(key=lambda x: x['_ts'], reverse=True)
    for r in result:
        del r['_ts']
    return jsonify(result)


# ---- Attachments (invoice images, email screenshots, PDFs) ----
ATTACH_DIR = os.path.join(BASE_DIR, 'attachments')
os.makedirs(ATTACH_DIR, exist_ok=True)
ALLOWED_ATTACH = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf', 'xlsx', 'xls', 'msg', 'eml'}


def _ext_ok(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_ATTACH


@app.route('/api/row/<int:row_id>/attachments', methods=['GET'])
@login_required
def api_attachments_get(row_id):
    row = db.session.get(ReconRow, row_id)
    if not row or not current_user.can_see_state(row.state_name):
        abort(403)
    atts = RowAttachment.query.filter_by(row_id=row_id).order_by(RowAttachment.uploaded_at).all()
    return jsonify([{
        'id': a.id, 'name': a.original_name, 'note': a.note or '',
        'size_kb': a.size_kb, 'mime': a.mime_type,
        'by': a.uploaded_by_name, 'at': a.uploaded_at.strftime('%d-%b-%Y %H:%M'),
    } for a in atts])


@app.route('/api/row/<int:row_id>/attachments', methods=['POST'])
@write_required
def api_attachments_post(row_id):
    row = db.session.get(ReconRow, row_id)
    if not row or not current_user.can_see_state(row.state_name):
        abort(403)
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': 'No file sent.'}), 400
    if not _ext_ok(f.filename):
        return jsonify({'error': f'File type not allowed. Use: {", ".join(sorted(ALLOWED_ATTACH))}'}), 400
    note = (request.form.get('note') or '').strip()[:300]
    safe = secure_filename(f.filename)
    stored = f'{row_id}_{now_ist().strftime("%Y%m%d%H%M%S")}_{safe}'
    path = os.path.join(ATTACH_DIR, stored)
    f.save(path)
    size_kb = max(1, os.path.getsize(path) // 1024)
    db.session.add(RowAttachment(
        row_id=row_id, filename=stored, original_name=f.filename,
        mime_type=f.mimetype or '', size_kb=size_kb, note=note,
        uploaded_by_id=current_user.id, uploaded_by_name=current_user.name,
    ))
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/attachments/<int:att_id>')
@login_required
def api_attachment_view(att_id):
    att = db.session.get(RowAttachment, att_id)
    if not att:
        abort(404)
    row = db.session.get(ReconRow, att.row_id)
    if not row or not current_user.can_see_state(row.state_name):
        abort(403)
    path = os.path.join(ATTACH_DIR, att.filename)
    if not os.path.exists(path):
        abort(404)
    return send_file(path, download_name=att.original_name, as_attachment=False)


@app.route('/api/attachments/<int:att_id>', methods=['DELETE'])
@write_required
def api_attachment_delete(att_id):
    att = db.session.get(RowAttachment, att_id)
    if not att:
        abort(404)
    # Only uploader or admin can delete
    if att.uploaded_by_id != current_user.id and not current_user.is_admin:
        abort(403)
    path = os.path.join(ATTACH_DIR, att.filename)
    if os.path.exists(path):
        os.remove(path)
    db.session.delete(att)
    db.session.commit()
    return jsonify({'ok': True})


# FIX 3 — Period lock / unlock
@app.route('/api/run/<int:run_id>/lock', methods=['POST'])
@login_required
def api_lock_run(run_id):
    if not current_user.is_superadmin:
        abort(403)
    run = db.session.get(ReconRun, run_id)
    if not run:
        abort(404)
    data = request.get_json(force=True)
    lock = data.get('lock', True)
    run.locked = lock
    run.locked_by_id = current_user.id if lock else None
    run.locked_at = now_ist() if lock else None
    action_label = 'lock' if lock else 'unlock'
    log_audit(None, current_user, action_label,
              'period_lock', '',
              f'State: {run.state} | Period: {run.period} | Run ID: {run.id}')
    db.session.commit()
    return jsonify({'ok': True, 'locked': run.locked})


@app.route('/api/lock-by-period', methods=['POST'])
@admin_required
def api_lock_by_period():
    data = request.get_json() or {}
    period = data.get('period', '')
    state = data.get('state', '')
    action = data.get('action', 'lock')
    if not period or period == 'all':
        return jsonify({'ok': False, 'error': 'Select a specific period'}), 400
    q = ReconRun.query.filter(ReconRun.period == period, ReconRun.archived != True)
    if state and state != 'all':
        q = q.filter(ReconRun.state == state)
    runs = q.all()
    if not runs:
        return jsonify({'ok': False, 'error': 'No runs found for this period'}), 404
    for run in runs:
        run.locked = (action == 'lock')
        run.locked_by_id = current_user.id if action == 'lock' else None
        run.locked_at = now_ist() if action == 'lock' else None
        log_audit(None, current_user, action,
                  'period_lock', '',
                  f'State: {run.state} | Period: {run.period} | Run ID: {run.id}')
    db.session.commit()
    return jsonify({'ok': True, 'locked': action == 'lock', 'count': len(runs)})


# ---- Feature 10: vendor follow-up (Book-Keeping shoots mail; everyone sees tracking) ----
@app.route('/api/row/<int:row_id>/followup', methods=['POST'])
@login_required
def api_followup(row_id):
    row = db.session.get(ReconRow, row_id)
    if not row:
        abort(404)
    if not current_user.can_see_state(row.state_name):
        abort(403)
    note = (request.get_json(force=True).get('note') or '').strip()
    row.followup_at = now_ist()
    row.followup_by_id = current_user.id
    row.followup_count = (row.followup_count or 0) + 1
    if note:
        row.followup_note = note

    # If SMTP configured + vendor email known, actually send the reminder.
    emailed = None
    vm = db.session.get(VendorMaster, row.gstin)
    vendor_email = vm.email if vm else ''
    if _smtp_configured() and vendor_email:
        import email_util
        inv = row.books_inv or row.gstn_inv or ''
        amt = row.books_total or row.gstn_total or 0
        body = render_template('vendor_followup_email.html', vendor=row.vendor or 'Vendor',
                               gstin=row.gstin, inv=inv, amount=f'{amt:,.0f}', note=note)
        ok, msg = email_util.send_email(_smtp_cfg(), vendor_email,
            f'GST follow-up: {row.gstin} {inv}', body)
        emailed = {'ok': ok, 'to': vendor_email, 'msg': msg}

    log_audit(row.id, current_user, 'followup', 'followup', '',
              f'#{row.followup_count} by {current_user.name}'
              + (f' · emailed {vendor_email}' if emailed and emailed['ok'] else ''))
    db.session.commit()
    return jsonify({'ok': True, 'row': row.to_dict(), 'emailed': emailed,
                    'mailto': not (emailed and emailed['ok'])})


# ---- GSTR-3B Table 4 tag ----
_TABLE4_OPTIONS = {
    '4A1': '4A(1) Import of Goods',
    '4A2': '4A(2) Import of Services',
    '4A3': '4A(3) Reverse Charge (RCM)',
    '4A4': '4A(4) From ISD',
    '4A5': '4A(5) All Other ITC',
    '4B1': '4B(1) ITC Reversed — Rules 38/42/43 & Sec 17(5)',
    '4B2': '4B(2) ITC Reversed — Others',
    '4D1': '4D(1) ITC Reclaimed (reversed earlier)',
    '4D2': '4D(2) Ineligible — Sec 16(4) / PoS',
}

@app.route('/api/table4-options')
@login_required
def api_table4_options():
    return jsonify(_TABLE4_OPTIONS)


@app.route('/api/row/<int:row_id>/table4', methods=['POST'])
@write_required
def api_set_table4(row_id):
    row = ReconRow.query.get_or_404(row_id)
    val = (request.get_json(force=True).get('table4') or '').strip()
    if val and val not in _TABLE4_OPTIONS:
        return jsonify(ok=False, error='Invalid table4 value'), 400
    row.itc_table4 = val
    db.session.commit()
    return jsonify(ok=True, itc_table4=val, label=_TABLE4_OPTIONS.get(val, ''))


# ---- Zoho ITC Accept / Reject ----
@app.route('/api/row/<int:row_id>/itc-action', methods=['POST'])
@write_required
def api_itc_action(row_id):
    """Accept (ITC eligible) or Reject (ITC ineligible) a GSTR-2B row,
    and optionally sync the action to Zoho Books."""
    if not current_user.is_admin:
        abort(403)
    row = db.session.get(ReconRow, row_id)
    if not row or not current_user.can_see_state(row.state_name):
        abort(403)
    data = request.get_json(force=True)
    action = data.get('action')  # 'accept' | 'reject' | 'reopen'
    if action not in ('accept', 'reject', 'reopen'):
        return jsonify({'ok': False, 'error': 'Invalid action'}), 400

    old_status = row.status
    if action == 'accept':
        row.status = 'approved'
        row.team_reason = row.team_reason or 'ITC Eligible — Accepted'
        label = 'ITC Accepted'
    elif action == 'reject':
        row.status = 'rejected'
        row.team_reason = data.get('reason') or row.team_reason or 'ITC Ineligible — Rejected'
        label = 'ITC Rejected'
    else:
        row.status = 'open'
        label = 'Reopened'

    row.approved_by_id = current_user.id
    row.approved_at = now_ist()
    log_audit(row.id, current_user, action, 'status', old_status, row.status)
    db.session.commit()
    return jsonify({'ok': True, 'action': action, 'label': label,
                    'zoho_ok': False, 'zoho_msg': '', 'row': row.to_dict()})


# ---- Vendor Email: send + thread + log ----
def _email_contact_line():
    cc = get_setting('email_cc', 'mahesh.thakur@wiom.in, tushar.gupta@wiom.in, ap@wiom.in')
    emails = [e.strip() for e in cc.split(',') if e.strip()]
    return 'For any concerns or queries, please feel free to contact ' + ', '.join(f'@{e}' for e in emails)

def _email_default_cc():
    return get_setting('email_cc', 'mahesh.thakur@wiom.in, wiomfinance@wiom.in, tushar.gupta@wiom.in, ap@wiom.in')

def _email_sign():
    return get_setting('email_sign', 'Regards,\nFinance and Taxation Team\nOmnia Information Private Limited')

_MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

def _fmt_period(p):
    if not p:
        return ''
    parts = str(p).split('-')
    if len(parts) < 2:
        return p
    try:
        return f"{_MONTHS[int(parts[1])-1]}' {parts[0][2:]}"
    except Exception:
        return p

def _fmt_date(d):
    if not d:
        return '—'
    s = str(d).strip()[:10]          # e.g. "16-04-2026" or "2026-04-16"
    parts = s.split('-')
    if len(parts) == 3:
        if len(parts[0]) == 4:       # YYYY-MM-DD → flip
            return f'{parts[2]}-{parts[1]}-{parts[0]}'
        return s                     # already DD-MM-YYYY
    try:
        import pandas as pd
        return pd.to_datetime(d, dayfirst=True).strftime('%d-%m-%Y')
    except Exception:
        return str(d)

def _inv_table(rows_data):
    """rows_data: list of (inv_no, inv_date_str, amount_float)"""
    sep = '  ' + '─' * 62
    hdr = f"\n{sep}\n  {'Sr.':<5} {'Invoice No':<30} {'Date':<14} Amount (₹)\n{sep}"
    body = '\n'.join(
        f"  {str(i+1)+'.':<5} {(r[0] or '—'):<30} {_fmt_date(r[1]):<14} {r[2]:,.0f}"
        for i, r in enumerate(rows_data)
    )
    return hdr + '\n' + body + '\n' + sep

EMAIL_TEMPLATES = {
    'not_filed': {
        'subject': 'GST ITC Alert: {vendor} — Invoice(s) not reflecting in GSTR-2B',
    },
    'not_received': {
        'subject': 'GST Invoice Clarification: {vendor} — Invoice in GSTR-2B but not received',
    },
    'followup': {
        'subject': 'Follow-up: {vendor} — GST ITC Mismatch (Invoice {inv})',
    },
}


def _build_email_body(tpl_key, row, sender_name, note=''):
    vendor  = row.vendor or 'Vendor'
    gstin   = row.gstin or ''
    period  = _fmt_period(row.period)
    note_ln = f'Note: {note}' if note else ''

    if tpl_key == 'not_filed':
        table = _inv_table([(row.books_inv or row.gstn_inv, row.books_date or row.gstn_date,
                             row.books_total or row.gstn_total or 0)])
        body = f"""Dear M/s {vendor},

Greetings from WIOM!

We have noticed that the following invoice(s) from your side are not reflecting in our GSTR-2B for the period {period}:
{table}
GSTIN (yours): {gstin}

This is causing an ITC mismatch in our books. Request you to:
1. Verify whether GSTR-1 has been filed for the above invoice.
2. If not filed, please file GSTR-1 at the earliest so ITC reflects in next month's GSTR-2B.
3. If already filed, please share the acknowledgement/filing date for our records.

Please revert on this email at the earliest to avoid any ITC loss on our side.
{note_ln}

{_email_contact_line()}

{_email_sign()}
"""

    elif tpl_key == 'not_received':
        table = _inv_table([(row.gstn_inv or row.books_inv, row.gstn_date or row.books_date,
                             row.gstn_total or row.books_total or 0)])
        body = f"""Dear M/s {vendor},

Greetings from WIOM!

We observed that an invoice appears in our GSTR-2B for the period {period} but has not been received / booked in our system:
{table}
GSTIN (yours): {gstin}

Request you to share the original invoice copy at the earliest so we can book it in our records.
{note_ln}

{_email_contact_line()}

{_email_sign()}
"""

    else:  # followup
        inv_no = row.books_inv or row.gstn_inv or '—'
        table = _inv_table([(inv_no, row.books_date or row.gstn_date,
                             row.books_total or row.gstn_total or 0)])
        body = f"""Dear M/s {vendor},

This is a follow-up to our earlier email regarding the GST mismatch for the period {period}:
{table}
GSTIN: {gstin}

We have not yet received a response. Request you to kindly revert at the earliest.
{note_ln}

{_email_contact_line()}

{_email_sign()}
"""

    return body


def _send_vendor_mail(to, cc, subject, body, in_reply_to=''):
    import smtplib, uuid
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    cfg = _smtp_cfg()
    if not cfg.get('host') or not cfg.get('user'):
        return False, 'SMTP not configured', ''
    msg = MIMEMultipart('alternative')
    msg_id = f'<wiom-recon-{uuid.uuid4().hex}@wiom.in>'
    msg['Message-ID'] = msg_id
    msg['From'] = f"WIOM Finance and Taxation <{cfg['user']}>"
    msg['To'] = to
    if cc:
        msg['Cc'] = cc
    msg['Subject'] = subject
    if in_reply_to:
        msg['In-Reply-To'] = in_reply_to
        msg['References'] = in_reply_to
    plain = body.replace('**', '').replace('•', '-')
    html = '<pre style="font-family:Arial,sans-serif;font-size:14px;white-space:pre-wrap;">' + \
           body.replace('**', '<b>').replace('\n', '<br>') + '</pre>'
    msg.attach(MIMEText(plain, 'plain'))
    msg.attach(MIMEText(html, 'html'))
    recipients = [e.strip() for e in (to + ',' + (cc or '')).split(',') if e.strip()]
    try:
        port = int(cfg.get('port', 587))
        if port == 465:
            s = smtplib.SMTP_SSL(cfg['host'], port, timeout=15)
        else:
            s = smtplib.SMTP(cfg['host'], port, timeout=15)
            s.starttls()
        s.login(cfg['user'], cfg['password'])
        s.sendmail(cfg['user'], recipients, msg.as_string())
        s.quit()
        return True, 'Sent', msg_id
    except Exception as e:
        return False, str(e), msg_id


@app.route('/api/row/<int:row_id>/vendor-email', methods=['GET'])
@login_required
def api_vendor_email_get(row_id):
    """Return vendor email address + prior email thread for this row."""
    row = db.session.get(ReconRow, row_id)
    if not row or not current_user.can_see_state(row.state_name):
        abort(403)
    vm = db.session.get(VendorMaster, row.gstin or '')
    vendor_email = vm.email if vm else ''
    emails = VendorEmail.query.filter_by(row_id=row_id).order_by(VendorEmail.sent_at).all()
    return jsonify({
        'vendor_email': vendor_email,
        'smtp_ok': _smtp_configured(),
        'thread': [{
            'id': e.id, 'seq': e.seq, 'tpl': e.template_type,
            'to': e.to_email, 'cc': e.cc_email, 'subject': e.subject,
            'body': e.body, 'by': e.sent_by_name,
            'at': e.sent_at.strftime('%d-%b-%Y %H:%M'),
            'ok': e.ok, 'error': e.error,
        } for e in emails],
    })


@app.route('/api/row/<int:row_id>/vendor-email', methods=['POST'])
@write_required
def api_vendor_email_post(row_id):
    """Send a vendor email (new or follow-up in same thread)."""
    row = db.session.get(ReconRow, row_id)
    if not row or not current_user.can_see_state(row.state_name):
        abort(403)
    data = request.get_json(force=True)
    to_email = (data.get('to') or '').strip()
    cc_email = (data.get('cc') or '').strip()
    tpl_key = data.get('template', 'not_filed')
    note = (data.get('note') or '').strip()
    custom_body = (data.get('body') or '').strip()
    custom_subject = (data.get('subject') or '').strip()

    if not to_email:
        return jsonify({'ok': False, 'error': 'Recipient email required.'}), 400

    # Save vendor email to master if not present
    vm = db.session.get(VendorMaster, row.gstin or '')
    if vm and not vm.email and to_email:
        vm.email = to_email
        db.session.flush()

    # Build subject + body
    prior = VendorEmail.query.filter_by(row_id=row_id).order_by(VendorEmail.sent_at).all()
    seq = len(prior) + 1
    first = prior[0] if prior else None
    tpl = EMAIL_TEMPLATES.get(tpl_key, EMAIL_TEMPLATES['not_filed'])
    inv = row.books_inv or row.gstn_inv or '—'
    auto_subject = tpl['subject'].format(inv=inv, vendor=row.vendor or row.gstin or 'Vendor')
    if seq > 1:
        auto_subject = f'Follow-up #{seq-1}: ' + auto_subject
    subject = custom_subject or auto_subject
    body = custom_body or _build_email_body(tpl_key, row, current_user.name, note)
    in_reply_to = first.message_id if first else ''

    ok, err, msg_id = _send_vendor_mail(to_email, cc_email, subject, body, in_reply_to)

    ve = VendorEmail(
        row_id=row_id, to_email=to_email, cc_email=cc_email,
        subject=subject, body=body, template_type=tpl_key,
        message_id=msg_id,
        thread_message_id=first.message_id if first else msg_id,
        seq=seq, sent_by_id=current_user.id, sent_by_name=current_user.name,
        ok=ok, error=err if not ok else '')
    db.session.add(ve)

    # Update followup tracking on the row
    row.followup_at = now_ist()
    row.followup_by_id = current_user.id
    row.followup_count = (row.followup_count or 0) + 1
    row.followup_note = f'Mail #{seq} to {to_email}' + (f' — {note}' if note else '')
    log_audit(row.id, current_user, 'vendor_email', 'email',
              '', f'#{seq} → {to_email} ({tpl_key}) ok={ok}')
    db.session.commit()
    return jsonify({'ok': ok, 'error': err, 'seq': seq, 'row': row.to_dict()})


@app.route('/api/vendor-email-tracker')
@login_required
def api_vendor_email_tracker():
    """Dashboard tracker: rows with vendor emails sent."""
    q = db.session.query(VendorEmail, ReconRow).join(
        ReconRow, VendorEmail.row_id == ReconRow.id)
    if not current_user.is_admin and current_user.state_list():
        q = q.filter(ReconRow.state_name.in_(current_user.state_list()))
    # Latest email per row
    from sqlalchemy import func
    sub = db.session.query(
        VendorEmail.row_id,
        func.max(VendorEmail.sent_at).label('last_at'),
        func.count(VendorEmail.id).label('cnt')
    ).group_by(VendorEmail.row_id).subquery()
    rows = db.session.query(ReconRow, sub.c.last_at, sub.c.cnt).join(
        sub, ReconRow.id == sub.c.row_id).order_by(sub.c.last_at.desc()).limit(200).all()
    return jsonify([{
        'id': r.id, 'gstin': r.gstin, 'vendor': r.vendor,
        'inv': r.books_inv or r.gstn_inv or '—',
        'state': r.state_name, 'period': r.period,
        'status': r.status, 'mail_count': cnt,
        'last_mail': last_at.strftime('%d-%b-%Y %H:%M') if last_at else '',
        'recon_status': r.recon_status,
    } for r, last_at, cnt in rows])


@app.route('/api/gstin/<gstin>/vendor-email', methods=['GET'])
@login_required
def api_gstin_vendor_email_get(gstin):
    """Return vendor email + all their problematic rows + email thread (any row of this GSTIN)."""
    vm = db.session.get(VendorMaster, gstin.strip())
    vendor_email = vm.email if vm else ''
    rows = ReconRow.query.filter(ReconRow.gstin == gstin.strip(),
                                 ReconRow.category != 'matched').order_by(ReconRow.period).all()
    # Email history: any VendorEmail linked to any row of this GSTIN
    row_ids = [r.id for r in ReconRow.query.filter(ReconRow.gstin == gstin.strip()).with_entities(ReconRow.id).all()]
    emails = VendorEmail.query.filter(VendorEmail.row_id.in_(row_ids)).order_by(VendorEmail.sent_at).all() if row_ids else []
    return jsonify({
        'vendor_email': vendor_email,
        'smtp_ok': _smtp_configured(),
        'rows': [{'id': r.id, 'category': r.category, 'period': r.period,
                  'books_inv': r.books_inv, 'gstn_inv': r.gstn_inv,
                  'books_date': r.books_date, 'gstn_date': r.gstn_date,
                  'books_total': r.books_total or 0, 'gstn_total': r.gstn_total or 0,
                  'recon_status': r.recon_status} for r in rows],
        'thread': [{'id': e.id, 'seq': e.seq, 'tpl': e.template_type,
                    'to': e.to_email, 'cc': e.cc_email, 'subject': e.subject,
                    'body': e.body, 'by': e.sent_by_name,
                    'at': e.sent_at.strftime('%d-%b-%Y %H:%M'),
                    'ok': e.ok, 'error': e.error} for e in emails],
    })


@app.route('/api/gstin/<gstin>/vendor-email', methods=['POST'])
@write_required
def api_gstin_vendor_email_post(gstin):
    """Send a vendor-level email covering ALL their mismatched invoices for a GSTIN."""
    gstin = gstin.strip()
    data = request.get_json(force=True)
    to_email = (data.get('to') or '').strip()
    cc_email = (data.get('cc') or '').strip()
    custom_subject = (data.get('subject') or '').strip()
    custom_body = (data.get('body') or '').strip()

    if not to_email:
        return jsonify({'ok': False, 'error': 'Recipient email required.'}), 400

    vm = db.session.get(VendorMaster, gstin)
    if vm and not vm.email and to_email:
        vm.email = to_email
        db.session.flush()

    # All mismatch rows for this GSTIN
    mismatch_rows = ReconRow.query.filter(ReconRow.gstin == gstin,
                                          ReconRow.category != 'matched').order_by(ReconRow.period).all()
    if not mismatch_rows:
        return jsonify({'ok': False, 'error': 'No mismatch rows found for this GSTIN.'}), 400

    # Email thread: any prior emails for any row of this GSTIN
    row_ids_all = [r.id for r in ReconRow.query.filter(ReconRow.gstin == gstin).with_entities(ReconRow.id).all()]
    prior = VendorEmail.query.filter(VendorEmail.row_id.in_(row_ids_all)).order_by(VendorEmail.sent_at).all() if row_ids_all else []
    seq = len(prior) + 1
    first = prior[0] if prior else None
    in_reply_to = first.message_id if first else ''

    vendor_name = (vm.name if vm else None) or (mismatch_rows[0].vendor if mismatch_rows else '') or gstin
    period = mismatch_rows[0].period if mismatch_rows else ''
    subject = custom_subject or f'GST Reconciliation Mismatch — {vendor_name} ({period})'
    body = custom_body

    ok, err, msg_id = _send_vendor_mail(to_email, cc_email, subject, body, in_reply_to)

    # Log VendorEmail against first mismatch row
    ref_row = mismatch_rows[0]
    ve = VendorEmail(
        row_id=ref_row.id, to_email=to_email, cc_email=cc_email,
        subject=subject, body=body, template_type='vendor_level',
        message_id=msg_id,
        thread_message_id=first.message_id if first else msg_id,
        seq=seq, sent_by_id=current_user.id, sent_by_name=current_user.name,
        ok=ok, error=err if not ok else '')
    db.session.add(ve)
    log_audit(ref_row.id, current_user, 'vendor_email', 'email',
              '', f'Vendor-level #{seq} → {to_email} ({len(mismatch_rows)} rows) ok={ok}')
    db.session.commit()
    return jsonify({'ok': ok, 'error': err, 'seq': seq})


# ---- Feature 5: vendor drill-down (all rows for a GSTIN, all states/months) ----
@app.route('/api/vendor-rows/<gstin>')
@login_required
def api_vendor_rows(gstin):
    q = ReconRow.query.filter(ReconRow.gstin == gstin.strip())
    if not current_user.is_admin and current_user.state_list():
        q = q.filter(ReconRow.state_name.in_(current_user.state_list()))
    rows = q.order_by(ReconRow.period, ReconRow.category).all()
    vm = db.session.get(VendorMaster, gstin.strip())
    return jsonify({'gstin': gstin, 'vendor': (vm.name if vm else (rows[0].vendor if rows else '')),
                    'rows': [r.to_dict() for r in rows], 'count': len(rows)})


# ---- Feature: Vendor Drill Down page ----
@app.route('/vendor/<gstin>')
@login_required
def vendor_detail_page(gstin):
    return render_template('vendor_detail.html', gstin=gstin.strip(), states=WIOM_STATES)


# ---- Feature: ITC trend by state ----
@app.route('/api/trend/by-state')
@login_required
def api_trend_by_state():
    months = int(request.args.get('months', 6))
    q = db.session.query(ReconRow.period, ReconRow.state_name,
                         func.sum(ReconRow.books_igst + ReconRow.books_cgst + ReconRow.books_sgst).label('itc'))
    if not current_user.is_admin and current_user.state_list():
        q = q.filter(ReconRow.state_name.in_(current_user.state_list()))
    q = q.filter(ReconRow.category == 'books_only').group_by(ReconRow.period, ReconRow.state_name)
    all_periods = sorted(set(r.period for r in q.all()), reverse=True)[:months]
    all_periods = sorted(all_periods)
    agg = {}
    for period, state, itc in q.all():
        if period in all_periods:
            agg.setdefault(state, {})[period] = round(itc or 0)
    states = sorted(agg.keys())
    return jsonify({'periods': all_periods, 'states': {s: [agg[s].get(p, 0) for p in all_periods] for s in states}})


# ---- Feature: Dashboard state health (traffic light) ----
@app.route('/api/dashboard-state-health')
@login_required
def api_dashboard_state_health():
    from sqlalchemy import func
    states = current_user.state_list() if not current_user.is_admin else WIOM_STATES
    out = []
    for state in states:
        rows_q = ReconRow.query.filter(ReconRow.state_name == state)
        total = rows_q.count()
        open_n = rows_q.filter(ReconRow.status == 'open').count()
        remarked = rows_q.filter(ReconRow.status == 'remarked').count()
        approved = rows_q.filter(ReconRow.status.in_(['approved', 'resolved'])).count()
        itc_risk = db.session.query(
            func.sum(ReconRow.books_igst + ReconRow.books_cgst + ReconRow.books_sgst)
        ).filter(ReconRow.state_name == state, ReconRow.category == 'books_only').scalar() or 0
        last_run = ReconRun.query.filter(ReconRun.state == state, ReconRun.archived != True)\
            .order_by(ReconRun.created_at.desc()).first()
        last_upload = last_run.created_at.strftime('%d-%b-%Y') if last_run else None
        last_period = last_run.period if last_run else None
        if total == 0:
            light = 'grey'
        elif open_n == 0 and remarked == 0:
            light = 'green'
        elif open_n == 0:
            light = 'yellow'
        else:
            light = 'red'
        out.append({'state': state, 'total': total, 'open': open_n, 'remarked': remarked,
                    'approved': approved, 'itc_risk': round(itc_risk),
                    'last_upload': last_upload, 'last_period': last_period, 'light': light})
    return jsonify(out)


# ---- Feature: Approval reminders (stale remarked rows) ----
@app.route('/api/approval-reminders')
@login_required
def api_approval_reminders():
    since = now_ist() - timedelta(days=3)
    q = ReconRow.query.filter(ReconRow.status == 'remarked', ReconRow.remarked_at <= since)
    if not current_user.is_admin and current_user.state_list():
        q = q.filter(ReconRow.state_name.in_(current_user.state_list()))
    count = q.count()
    by_state = {}
    for r in q.all():
        by_state[r.state_name] = by_state.get(r.state_name, 0) + 1
    return jsonify({'count': count, 'by_state': by_state})


# ---- Feature: Locked periods list ----
@app.route('/api/locked-periods')
@login_required
def api_locked_periods():
    runs = ReconRun.query.filter(ReconRun.locked == True, ReconRun.archived != True).all()
    user_ids = {r.locked_by_id for r in runs if r.locked_by_id}
    users = {u.id: u.name for u in User.query.filter(User.id.in_(user_ids)).all()} if user_ids else {}
    return jsonify([{'period': r.period, 'state': r.state,
                     'locked_at': r.locked_at.strftime('%d-%b-%Y') if r.locked_at else '',
                     'locked_by': users.get(r.locked_by_id, '') if r.locked_by_id else ''}
                    for r in runs])


# ---- Feature: Run diff report ----
@app.route('/api/run/<int:run_id>/diff')
@login_required
def api_run_diff(run_id):
    run = db.session.get(ReconRun, run_id)
    if not run:
        return jsonify({'error': 'Not found'}), 404
    try:
        summary = json.loads(run.summary_json or '{}')
        diff = summary.get('diff', {})
        return jsonify({'ok': True, 'diff': diff, 'period': run.period, 'state': run.state})
    except Exception:
        return jsonify({'ok': True, 'diff': {}})


# ---- Feature: Bulk remark Excel template download ----
@app.route('/api/bulk-remark/template')
@login_required
def bulk_remark_template():
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    period = request.args.get('period', '')
    state = request.args.get('state', '')
    q = ReconRow.query
    if period:
        q = q.filter(ReconRow.period == period)
    if state:
        q = q.filter(ReconRow.state_name == state)
    if not current_user.is_admin and current_user.state_list():
        q = q.filter(ReconRow.state_name.in_(current_user.state_list()))
    rows = q.order_by(ReconRow.state_name, ReconRow.gstin).limit(5000).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Bulk Remarks'
    headers = ['row_id', 'state', 'period', 'gstin', 'vendor', 'books_inv', 'gstn_inv',
               'category', 'books_tax', 'current_status', 'current_remark', 'NEW_REMARK', 'NEW_REASON']
    yellow = PatternFill('solid', fgColor='FFFF00')
    bold = Font(bold=True)
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h)
        c.font = bold
        if h.startswith('NEW_'):
            c.fill = yellow
    reasons = ['Rate Difference', 'Invoice Not Filed', 'Period Mismatch',
               'Duplicate Invoice', 'Cancelled Invoice', 'RCM Transaction', 'Other']
    for row in rows:
        tax = (row.books_igst or 0) + (row.books_cgst or 0) + (row.books_sgst or 0)
        ws.append([row.id, row.state_name, row.period, row.gstin, row.vendor,
                   row.books_inv, row.gstn_inv, row.category, round(tax, 2),
                   row.status, row.team_remark or '', '', ''])
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['L'].width = 40
    ws.column_dimensions['M'].width = 25
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"bulk_remarks_{state or 'all'}_{period or 'all'}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


# ---- Feature: Bulk remark Excel upload ----
@app.route('/api/bulk-remark/upload', methods=['POST'])
@write_required
def bulk_remark_upload():
    import openpyxl
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file'}), 400
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    except Exception:
        return jsonify({'error': 'Invalid or corrupted Excel file. Download the template first.'}), 400
    ws = wb.active
    headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        id_col = headers.index('row_id')
        remark_col = headers.index('NEW_REMARK')
        reason_col = headers.index('NEW_REASON')
    except ValueError:
        return jsonify({'error': 'Missing columns: row_id, NEW_REMARK, NEW_REASON. Download the template first.'}), 400
    updated = skipped = 0
    for data_row in ws.iter_rows(min_row=2, values_only=True):
        row_id = data_row[id_col]
        new_remark = str(data_row[remark_col] or '').strip()
        new_reason = str(data_row[reason_col] or '').strip()
        if not row_id or not new_remark:
            skipped += 1
            continue
        row = db.session.get(ReconRow, int(row_id))
        if not row:
            skipped += 1
            continue
        if not current_user.is_admin and not current_user.can_see_state(row.state_name):
            skipped += 1
            continue
        row.team_remark = new_remark
        if new_reason:
            row.team_reason = new_reason
        if row.status == 'open':
            row.status = 'remarked'
        row.remarked_by_id = current_user.id
        row.remarked_at = now_ist()
        log_audit(row.id, current_user, 'remark', 'bulk_excel', '', new_remark)
        updated += 1
    db.session.commit()
    return jsonify({'ok': True, 'updated': updated, 'skipped': skipped})


# ---- Feature: CFO PDF (print-ready redirect) ----
@app.route('/cfo-pdf')
@admin_required
def cfo_pdf():
    return redirect(url_for('cfo_summary', _anchor='print') + '&print=1')


# ---- Feature: 4: ITC-at-risk trend (month series) ----
@app.route('/api/trend')
@login_required
def api_trend():
    q = ReconRow.query
    if not current_user.is_admin and current_user.state_list():
        q = q.filter(ReconRow.state_name.in_(current_user.state_list()))
    series = {}
    for r in q.all():
        s = series.setdefault(r.period, {'itc_risk': 0.0, 'gap': 0.0})
        if r.category == 'books_only':  # in books, not in 2B → ITC at risk
            s['itc_risk'] += (r.books_igst or 0) + (r.books_cgst or 0) + (r.books_sgst or 0)
        s['gap'] += r.total_diff or 0
    out = [{'period': p, 'itc_risk': round(v['itc_risk'], 0), 'gap': round(v['gap'], 0)}
           for p, v in sorted(series.items())]
    return jsonify(out)


@app.route('/api/cumulative')
@login_required
def api_cumulative():
    """State x period pivot of total differences — cumulative view."""
    q = db.session.query(
        ReconRow.state_name, ReconRow.period,
        func.count(ReconRow.id),
        func.sum(ReconRow.total_diff),
        func.sum(db.case((ReconRow.status.in_(['approved', 'resolved']), 1), else_=0)),
    )
    if not current_user.is_admin and current_user.state_list():
        q = q.filter(ReconRow.state_name.in_(current_user.state_list()))
    q = q.group_by(ReconRow.state_name, ReconRow.period)
    out = {}
    for state, period, cnt, diff, done in q.all():
        out.setdefault(state, {})[period] = {
            'count': cnt, 'diff': round(diff or 0, 2), 'done': done,
        }
    return jsonify(out)


@app.route('/api/vendor/<gstin>')
@login_required
def api_vendor(gstin):
    vm = db.session.get(VendorMaster, gstin.strip())
    if vm:
        if not current_user.can_see_state(vm.state_name):
            abort(403)
        return jsonify({'gstin': vm.gstin, 'name': vm.name,
                        'state': vm.state_name, 'source': vm.source})
    return jsonify({'gstin': gstin, 'name': 'Unknown', 'state': '', 'source': ''})


# ---- Feature 7: bulk import remarks from CSV ----
@app.route('/import-remarks', methods=['GET', 'POST'])
@login_required
def import_remarks():
    if request.method == 'POST':
        import csv, io as _io
        f = request.files.get('file')
        if not f or not f.filename.lower().endswith('.csv'):
            return jsonify({'ok': False, 'error': 'Upload a .csv file.'}), 400
        text = f.read().decode('utf-8-sig', errors='replace')
        reader = csv.DictReader(_io.StringIO(text))
        applied, skipped = 0, 0
        for r in reader:
            row = None
            rid = (r.get('row_id') or r.get('id') or '').strip()
            if rid.isdigit():
                row = db.session.get(ReconRow, int(rid))
            if row is None:
                g = (r.get('gstin') or '').strip()
                inv = (r.get('invoice') or r.get('books_inv') or r.get('gstn_inv') or '').strip()
                if g and inv:
                    row = ReconRow.query.filter(ReconRow.gstin == g,
                        (ReconRow.books_inv == inv) | (ReconRow.gstn_inv == inv)).first()
            if row is None or not current_user.can_see_state(row.state_name):
                skipped += 1
                continue
            remark = (r.get('remark') or '').strip()
            reason = (r.get('reason') or '').strip()
            if not (remark or reason):
                skipped += 1
                continue
            old = row.team_remark
            if remark:
                row.team_remark = remark
            if reason:
                row.team_reason = reason
            row.remarked_by_id = current_user.id
            row.remarked_at = now_ist()
            if row.status == 'open':
                row.status = 'remarked'
            log_audit(row.id, current_user, 'remark', 'team_remark(import)', old, row.team_remark)
            applied += 1
        db.session.commit()
        return jsonify({'ok': True, 'applied': applied, 'skipped': skipped})
    return render_template('import_remarks.html')


# ---- Feature 9: CFO one-page executive summary (print → PDF) ----
def _cfo_context(rows, state):
    gap = gap_from_rows(rows)
    def tax(r, side):
        if side == 'b':
            return (r.books_igst or 0) + (r.books_cgst or 0) + (r.books_sgst or 0)
        return (r.gstn_igst or 0) + (r.gstn_cgst or 0) + (r.gstn_sgst or 0)
    itc_risk = sum(tax(r, 'b') for r in rows if r.category == 'books_only')
    excess_2b = sum(tax(r, 'g') for r in rows if r.category == 'gstn_only')
    matched = [r for r in rows if r.category == 'matched']
    fully = sum(1 for r in matched if 'Fully Reconciled' in (r.recon_status or ''))
    by_state = {}
    for r in rows:
        s = by_state.setdefault(r.state_name, {'rows': 0, 'gap': 0.0, 'risk': 0.0})
        s['rows'] += 1; s['gap'] += r.total_diff or 0
        if r.category == 'books_only':
            s['risk'] += tax(r, 'b')
    done = sum(1 for r in rows if r.status in ('approved', 'resolved'))
    rejected = sum(1 for r in rows if r.status == 'rejected')
    return dict(
        scope=(state if state and state != 'all' else 'All States'),
        generated=now_ist().strftime('%d-%b-%Y %H:%M'),
        total=len(rows), matched=len(matched), fully=fully,
        books_only=sum(1 for r in rows if r.category == 'books_only' and r.status != 'rejected'),
        gstn_only=sum(1 for r in rows if r.category == 'gstn_only' and r.status != 'rejected'),
        itc_risk=round(itc_risk), excess_2b=round(excess_2b),
        done=done, pending=len(rows) - done,
        rejected=rejected,
        by_state={k: {'rows': v['rows'], 'gap': round(v['gap']), 'risk': round(v['risk'])}
                  for k, v in sorted(by_state.items())},
        top_gaps=gap[:10], breakdown=_breakdown_data(rows))


@app.route('/cfo-summary')
@admin_required
def cfo_summary():
    rows = _filtered_rows_query().all()
    return render_template('cfo_summary.html', **_cfo_context(rows, request.args.get('state')))


# ---- Feature 12: login history + DB backup (super admin) ----
@app.route('/login-history')
@superadmin_required
def login_history():
    events = LoginEvent.query.order_by(LoginEvent.created_at.desc()).limit(500).all()
    return render_template('login_history.html', events=events)


@app.route('/settings/backup')
@superadmin_required
def backup_db():
    uri = app.config['SQLALCHEMY_DATABASE_URI']
    if uri.startswith('sqlite:///'):
        path = uri.replace('sqlite:///', '', 1)
        if os.path.exists(path):
            return send_file(path, as_attachment=True,
                download_name=f"wiom_recon_backup_{now_ist().strftime('%Y%m%d_%H%M')}.db")
    return jsonify({'error': 'Backup only available for SQLite (local) deployments.'}), 400


@app.route('/settings/fix-fully-reconciled', methods=['POST'])
@superadmin_required
def fix_fully_reconciled_status():
    """One-time: set status=approved for all Fully Reconciled (Zoho Auto) rows."""
    updated = ReconRow.query.filter(
        ReconRow.recon_status.like('%Fully Reconciled%'),
        ReconRow.status == 'open'
    ).update({'status': 'approved'}, synchronize_session=False)
    db.session.commit()
    return jsonify({'ok': True, 'msg': f'{updated} rows updated to approved.'})


@app.route('/settings/clear-recon', methods=['POST'])
@superadmin_required
def clear_recon_data():
    """Wipe all recon rows/runs/comments/audit logs. Users, settings, vendor master kept."""
    from models import AuditLog, RowComment, ReconRow, ReconRun
    AuditLog.query.delete(synchronize_session=False)
    RowComment.query.delete(synchronize_session=False)
    ReconRow.query.delete(synchronize_session=False)
    ReconRun.query.delete(synchronize_session=False)
    db.session.commit()
    return jsonify({'ok': True, 'msg': 'All reconciliation data cleared. Users and settings intact.'})


@app.route('/api/recent-runs')
@login_required
def api_recent_runs():
    q = ReconRun.query
    if not current_user.is_admin and current_user.state_list():
        q = q.filter(ReconRun.state.in_(current_user.state_list()))
    runs = q.order_by(ReconRun.created_at.desc()).limit(10).all()
    return jsonify([{
        'id': r.id, 'period': r.period or '', 'state': r.state or 'All',
        'label': r.label or '', 'total_rows': r.total_rows or 0,
        'uploaded_by': r.uploaded_by.name if r.uploaded_by else '—',
        'created_at': r.created_at.strftime('%d-%b-%Y %H:%M') if r.created_at else ''
    } for r in runs])


@app.route('/api/activity')
@login_required
def api_activity():
    logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(25).all()
    # Batch fetch rows (avoid N+1)
    row_ids = list({l.row_id for l in logs if l.row_id})
    rows_by_id = {}
    if row_ids:
        for r in ReconRow.query.filter(ReconRow.id.in_(row_ids)).all():
            rows_by_id[r.id] = r
    result = []
    for l in logs:
        row = rows_by_id.get(l.row_id) if l.row_id else None
        result.append({
            'user': l.user_name or '—',
            'action': l.action or '',
            'detail': l.new_value or l.field or '',
            'gstin': row.gstin if row else '',
            'vendor': row.vendor if row else '',
            'at': l.created_at.strftime('%d-%b %H:%M') if l.created_at else ''
        })
    return jsonify(result)


@app.route('/api/gstin-suggest')
@login_required
def api_gstin_suggest():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])
    like = f'%{q}%'
    results = VendorMaster.query.filter(
        VendorMaster.gstin.like(like) | VendorMaster.name.like(like)
    ).limit(10).all()
    return jsonify([{'gstin': v.gstin, 'name': v.name or ''} for v in results])


# ----------------------------------------------------------------------
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(e):
    db.session.rollback()
    return render_template('500.html'), 500


def _summary_stats(user):
    q = ReconRow.query
    if not user.is_admin and user.state_list():
        q = q.filter(ReconRow.state_name.in_(user.state_list()))
    total = q.count()
    open_n = q.filter(ReconRow.status == 'open').count()
    remarked = q.filter(ReconRow.status == 'remarked').count()
    done = q.filter(ReconRow.status.in_(['approved', 'resolved'])).count()
    itc = 0.0
    for r in q.filter(ReconRow.category == 'books_only').all():
        itc += (r.books_igst or 0) + (r.books_cgst or 0) + (r.books_sgst or 0)
    return {'total': total, 'open': open_n, 'remarked': remarked, 'done': done,
            'itc_risk': round(itc)}


if __name__ == '__main__':
    # Daily Slack summary at 7:00 AM IST
    from apscheduler.schedulers.background import BackgroundScheduler
    import pytz
    scheduler = BackgroundScheduler(timezone=pytz.timezone('Asia/Kolkata'))
    scheduler.add_job(slack_notify.notify_pending_summary, 'cron', hour=7,  minute=0)
    scheduler.add_job(slack_notify.notify_cfo_summary,     'cron', hour=19, minute=0)
    scheduler.start()

    # Optional port override from CLI (used by the preview launcher)
    port = Config.PORT
    if len(sys.argv) > 1 and sys.argv[1].isdigit():
        port = int(sys.argv[1])
    print("\n" + "=" * 64)
    print("  WIOM Zoho Books vs GST Recon — Workflow Platform")
    print(f"  http://localhost:{port}")
    print(f"  Login: {Config.ADMIN_EMAIL} / {Config.ADMIN_PASSWORD}")
    print("=" * 64 + "\n")
    app.run(debug=False, port=port, host=Config.HOST)
