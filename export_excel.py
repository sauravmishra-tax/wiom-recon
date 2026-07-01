"""
Cumulative Excel export — builds a multi-sheet workbook from the database
(all months to date), not a single month. Mirrors the report tabs and adds
the team workflow columns (reason, remark, status, who approved).
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PINK = PatternFill('solid', fgColor='E0119D')
DARK = PatternFill('solid', fgColor='1C1C28')
HFONT = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
TITLE = Font(name='Calibri', bold=True, size=13, color='E0119D')
WFONT = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
INR = '#,##0.00'
THIN = Border(bottom=Side(style='thin', color='DDDDDD'))


def _sheet(wb, name, color, title, headers):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = color
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1, title).font = TITLE
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.fill = PINK; cell.font = HFONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'A4'
    return ws


def _row(ws, r, values, money_from=None):
    for c, v in enumerate(values, 1):
        cell = ws.cell(r, c, v)
        cell.border = THIN
        if money_from and c >= money_from and isinstance(v, (int, float)):
            cell.number_format = INR


def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_cumulative_excel(rows, gap_rows, scope_label):
    """rows: list of ReconRow ; gap_rows: list of dicts from gap aggregation."""
    wb = Workbook()
    wb.remove(wb.active)

    matched = [r for r in rows if r.category == 'matched']
    fully = [r for r in matched if 'Fully Reconciled' in (r.recon_status or '')]
    cross_only = [r for r in matched if 'Fully Reconciled' not in (r.recon_status or '')]
    books = [r for r in rows if r.category == 'books_only']
    gstn = [r for r in rows if r.category == 'gstn_only']

    def tax(r, side):
        if side == 'b':
            return (r.books_igst or 0) + (r.books_cgst or 0) + (r.books_sgst or 0)
        return (r.gstn_igst or 0) + (r.gstn_cgst or 0) + (r.gstn_sgst or 0)

    # ---- Expert Cross-Match ----
    h = ['S.No', 'GSTIN', 'Vendor', 'State', 'Period', 'Type', 'Books Inv', '2B Inv',
         'Books Date', '2B Date', 'Books Taxable', '2B Taxable', 'Taxable Diff',
         'Books Tax', '2B Tax', 'Tax Diff', 'Books Total', '2B Total', 'Total Diff',
         'Recon Status', 'Reason', 'Team Remark', 'Workflow Status', 'Remarked By', 'Approved By']
    ws = _sheet(wb, 'Expert Cross-Match', 'E0119D',
                f'WIOM · CUMULATIVE Expert Cross-Match · {scope_label}', h)
    for i, r in enumerate(cross_only, 1):
        _row(ws, i + 3, [i, r.gstin, r.vendor, r.state_name, r.period, r.txn_type,
            r.books_inv, r.gstn_inv, r.books_date, r.gstn_date,
            r.books_taxable, r.gstn_taxable, (r.books_taxable or 0) - (r.gstn_taxable or 0),
            tax(r, 'b'), tax(r, 'g'), tax(r, 'b') - tax(r, 'g'),
            r.books_total, r.gstn_total, r.total_diff, r.recon_status,
            r.team_reason, r.team_remark, r.status,
            r.remarked_by.name if r.remarked_by else '', r.approved_by.name if r.approved_by else ''],
            money_from=11)
    _widths(ws, [6, 20, 30, 13, 9, 10, 18, 18, 12, 12, 14, 14, 13, 13, 13, 13, 14, 14, 13, 24, 22, 34, 14, 16, 16])

    # ---- Books Only / GSTR-2B Only ----
    for name, color, items, side, title in [
        ('Books Only Detail', 'CC0000', books, 'b', 'IN BOOKS, NOT IN GSTR-2B (ITC at risk)'),
        ('GSTR-2B Only Detail', '2980B9', gstn, 'g', 'IN GSTR-2B, NOT IN BOOKS (unbooked)'),
    ]:
        hh = ['S.No', 'GSTIN', 'Vendor', 'State', 'Period', 'Invoice No', 'Type', 'Date',
              'Taxable', 'IGST', 'CGST', 'SGST', 'Total Tax', 'Total Amount',
              'Cross Status', 'Reason', 'Team Remark', 'Workflow Status', 'Remarked By', 'Approved By']
        ws = _sheet(wb, name, color, f'WIOM · CUMULATIVE {title} · {scope_label}', hh)
        for i, r in enumerate(items, 1):
            inv = r.books_inv if side == 'b' else r.gstn_inv
            dt = r.books_date if side == 'b' else r.gstn_date
            taxable = r.books_taxable if side == 'b' else r.gstn_taxable
            ig = r.books_igst if side == 'b' else r.gstn_igst
            cg = r.books_cgst if side == 'b' else r.gstn_cgst
            sg = r.books_sgst if side == 'b' else r.gstn_sgst
            tot = r.books_total if side == 'b' else r.gstn_total
            _row(ws, i + 3, [i, r.gstin, r.vendor, r.state_name, r.period, inv, r.txn_type, dt,
                taxable, ig, cg, sg, tax(r, side), tot, r.recon_status,
                r.team_reason, r.team_remark, r.status,
                r.remarked_by.name if r.remarked_by else '', r.approved_by.name if r.approved_by else ''],
                money_from=9)
        _widths(ws, [6, 20, 30, 13, 9, 18, 10, 12, 14, 13, 13, 13, 14, 15, 22, 22, 34, 14, 16, 16])

    # ---- Fully Reconciled ----
    hh = ['S.No', 'GSTIN', 'Vendor', 'State', 'Period', 'Invoice No', 'Type', 'Date',
          'Taxable', 'Total Tax', 'Total Amount', 'ITC Status', 'Team Remark', 'Workflow Status']
    ws = _sheet(wb, 'Fully Reconciled Detail', '00A36C',
                f'WIOM · CUMULATIVE Fully Reconciled · {scope_label}', hh)
    for i, r in enumerate(fully, 1):
        _row(ws, i + 3, [i, r.gstin, r.vendor, r.state_name, r.period, r.books_inv, r.txn_type,
            r.books_date, r.books_taxable, tax(r, 'b'), r.books_total, r.recon_status,
            r.team_remark, r.status], money_from=9)
    _widths(ws, [6, 20, 30, 13, 9, 18, 10, 12, 14, 14, 15, 22, 34, 14])

    # ---- GSTIN Gap Analysis ----
    hh = ['S.No', 'GSTIN', 'Vendor', 'State', 'Books #', 'Books Taxable', 'Books Tax', 'Books Total',
          '2B #', '2B Taxable', '2B Tax', '2B Total', 'Taxable Gap', 'Tax Gap', 'Total Gap',
          'Gap %', 'Risk Level', 'Remark', 'Action']
    ws = _sheet(wb, 'GSTIN Gap Analysis', 'FF6600',
                f'WIOM · CUMULATIVE GSTIN Gap Analysis · {scope_label}', hh)
    for i, g in enumerate(gap_rows, 1):
        _row(ws, i + 3, [i, g['gstin'], g['vendor'], g['state'], g['b_cnt'], g['b_taxable'],
            g['b_tax'], g['b_total'], g['g_cnt'], g['g_taxable'], g['g_tax'], g['g_total'],
            g['taxable_gap'], g['tax_gap'], g['total_gap'], g['gap_pct'], g['risk'],
            g['remark'], g['action']], money_from=6)
        ws.cell(i + 3, 16).number_format = '0.0%'
    _widths(ws, [6, 20, 30, 13, 8, 14, 13, 14, 8, 14, 13, 14, 14, 13, 14, 9, 12, 26, 26])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
